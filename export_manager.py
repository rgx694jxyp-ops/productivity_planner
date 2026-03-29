"""
export_manager.py
-----------------
Exports everything to Excel.
Every export function returns raw bytes ready for st.download_button().
"""

import io
import os
import tempfile
from datetime import datetime

from database import (
    get_orders, get_order, get_submissions, get_order_hours_worked,
    get_clients, get_client_trends, get_employees, get_uph_history,
    get_coaching_notes, get_assignments,
)
from orders import get_order_progress


def _new_workbook():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl.Workbook()
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip3 install openpyxl")


def _header_row(ws, headers: list[str], row: int = 1):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="FF0F2D52")
    font = Font(bold=True, color="FFFFFFFF", size=10, name="Arial")
    aln  = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = fill; c.font = font; c.alignment = aln
    ws.row_dimensions[row].height = 18


def _autofit(ws, max_w: int = 40):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_w)


def _wb_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Order export ──────────────────────────────────────────────────────────────

def export_order(order_id: str) -> bytes:
    """Full export for one order: summary + daily submissions + employee breakdown."""
    wb  = _new_workbook()
    wb.remove(wb.active)

    order    = get_order(order_id)
    progress = get_order_progress(order_id)
    subs     = get_submissions(order_id=order_id)
    assigns  = get_assignments(order_id=order_id, active_only=False)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    _header_row(ws, ["Field", "Value"])
    summary_rows = [
        ("Order Number",     order.get("order_number", "")),
        ("Description",      order.get("description", "")),
        ("Client",           order.get("clients", {}).get("name", "") if order.get("clients") else ""),
        ("Status",           order.get("status", "")),
        ("Total Units",      order.get("total_units", 0)),
        ("Units Completed",  progress.get("units_completed", 0)),
        ("Units Remaining",  progress.get("units_remaining", 0)),
        ("% Complete",       f"{progress.get('pct_complete', 0)}%"),
        ("Hours Worked",     progress.get("hours_worked", 0)),
        ("Target Date",      order.get("target_date", "")),
        ("Created",          str(order.get("created_at", ""))[:10]),
        ("Completed",        str(order.get("completed_at", ""))[:10] if order.get("completed_at") else ""),
    ]
    for i, (field, val) in enumerate(summary_rows, start=2):
        ws.cell(i, 1, field)
        ws.cell(i, 2, val)
    _autofit(ws)

    # ── Daily submissions sheet ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Daily Submissions")
    _header_row(ws2, ["Date", "Employee ID", "Units", "UPH", "Hours Worked", "Source File"])
    for i, s in enumerate(subs, start=2):
        ws2.cell(i, 1, s.get("work_date", ""))
        ws2.cell(i, 2, s.get("emp_id", ""))
        ws2.cell(i, 3, s.get("units", 0))
        ws2.cell(i, 4, s.get("uph", 0))
        ws2.cell(i, 5, s.get("hours_worked", 0))
        ws2.cell(i, 6, s.get("source_file", ""))
    _autofit(ws2)

    # ── Employee breakdown sheet ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Employee Breakdown")
    _header_row(ws3, ["Employee ID", "Total Units", "Total Hours", "Avg UPH", "Days Worked"])
    from collections import defaultdict
    emp_agg: dict[str, dict] = defaultdict(lambda: {"units": 0.0, "hours": 0.0, "uphs": [], "days": set()})
    for s in subs:
        eid = s.get("emp_id", "")
        emp_agg[eid]["units"]  += float(s.get("units") or 0)
        emp_agg[eid]["hours"]  += float(s.get("hours_worked") or 0)
        if s.get("uph"): emp_agg[eid]["uphs"].append(float(s["uph"]))
        emp_agg[eid]["days"].add(s.get("work_date", ""))
    for i, (eid, agg) in enumerate(emp_agg.items(), start=2):
        avg_uph = round(sum(agg["uphs"]) / len(agg["uphs"]), 2) if agg["uphs"] else 0
        ws3.cell(i, 1, eid)
        ws3.cell(i, 2, agg["units"])
        ws3.cell(i, 3, round(agg["hours"], 2))
        ws3.cell(i, 4, avg_uph)
        ws3.cell(i, 5, len(agg["days"]))
    _autofit(ws3)

    return _wb_to_bytes(wb)


# ── Client export ─────────────────────────────────────────────────────────────

def export_client(client_id: str) -> bytes:
    """Full client history: all orders + trend data."""
    wb = _new_workbook()
    wb.remove(wb.active)

    clients = get_clients()
    client  = next((c for c in clients if c["id"] == client_id), {})
    orders  = get_orders(client_id=client_id)
    trends  = get_client_trends(client_id)

    # ── Orders sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Orders")
    _header_row(ws, ["Order #", "Description", "Status", "Total Units",
                     "Completed", "% Done", "Created", "Completed On"])
    for i, o in enumerate(orders, start=2):
        total     = float(o.get("total_units") or 0)
        completed = float(o.get("units_completed") or 0)
        pct       = f"{round(completed/total*100,1)}%" if total > 0 else "0%"
        ws.cell(i, 1, o.get("order_number", ""))
        ws.cell(i, 2, o.get("description", ""))
        ws.cell(i, 3, o.get("status", ""))
        ws.cell(i, 4, total)
        ws.cell(i, 5, completed)
        ws.cell(i, 6, pct)
        ws.cell(i, 7, str(o.get("created_at", ""))[:10])
        ws.cell(i, 8, str(o.get("completed_at", ""))[:10] if o.get("completed_at") else "")
    _autofit(ws)

    # ── Trends sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Productivity Trends")
    _header_row(ws2, ["Period", "Avg UPH", "Total Units", "Orders Completed"])
    for i, t in enumerate(trends, start=2):
        ws2.cell(i, 1, t.get("period", ""))
        ws2.cell(i, 2, t.get("avg_uph", 0))
        ws2.cell(i, 3, t.get("total_units", 0))
        ws2.cell(i, 4, t.get("orders_completed", 0))
    _autofit(ws2)

    return _wb_to_bytes(wb)


# ── Employee export ───────────────────────────────────────────────────────────

def export_employee(emp_id: str) -> bytes:
    """Full employee history: UPH trend + coaching notes."""
    wb = _new_workbook()
    wb.remove(wb.active)

    history = get_uph_history(emp_id, days=365)
    notes   = get_coaching_notes(emp_id)

    ws = wb.create_sheet("UPH History")
    _header_row(ws, ["Date", "UPH", "Units", "Hours Worked", "Department"])
    for i, h in enumerate(history, start=2):
        ws.cell(i, 1, h.get("work_date", ""))
        ws.cell(i, 2, h.get("uph", 0))
        ws.cell(i, 3, h.get("units", 0))
        ws.cell(i, 4, h.get("hours_worked", 0))
        ws.cell(i, 5, h.get("department", ""))
    _autofit(ws)

    ws2 = wb.create_sheet("Coaching Notes")
    _header_row(ws2, ["Date", "Note", "Created By"])
    for i, n in enumerate(notes, start=2):
        ws2.cell(i, 1, str(n.get("created_at", ""))[:10])
        ws2.cell(i, 2, n.get("note", ""))
        ws2.cell(i, 3, n.get("created_by", ""))
    _autofit(ws2)

    return _wb_to_bytes(wb)


# ── Full data export ──────────────────────────────────────────────────────────

def export_all_orders() -> bytes:
    """All orders across all clients — summary view."""
    wb  = _new_workbook()
    ws  = wb.active
    ws.title = "All Orders"
    _header_row(ws, ["Client", "Order #", "Description", "Status",
                     "Total Units", "Completed", "Remaining",
                     "% Done", "Hours Worked", "Target Date"])
    orders = get_orders()
    for i, o in enumerate(orders, start=2):
        prog = get_order_progress(o["id"])
        ws.cell(i, 1,  (o.get("clients") or {}).get("name", ""))
        ws.cell(i, 2,  o.get("order_number", ""))
        ws.cell(i, 3,  o.get("description", ""))
        ws.cell(i, 4,  o.get("status", ""))
        ws.cell(i, 5,  prog.get("total_units", 0))
        ws.cell(i, 6,  prog.get("units_completed", 0))
        ws.cell(i, 7,  prog.get("units_remaining", 0))
        ws.cell(i, 8,  f"{prog.get('pct_complete', 0)}%")
        ws.cell(i, 9,  get_order_hours_worked(o["id"]))
        ws.cell(i, 10, o.get("target_date", ""))
    _autofit(ws)
    return _wb_to_bytes(wb)
