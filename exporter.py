"""
exporter.py
-----------
The publishing chapter: turning in-memory data into real files.

Produces:
  1. A styled Excel workbook (.xlsx) with all report sheets
  2. Optional PDF export via matplotlib (cross-platform, no Outlook needed)

Requires: pip install openpyxl matplotlib
"""

import os
from datetime import datetime

from settings  import Settings
from error_log import ErrorLog
from ranker    import build_top_bottom_summary


# ── Colour palette (Excel ARGB hex strings) ──────────────────────────────────
NAVY      = "FF1F497D"
STEEL     = "FFBDD7EE"
WHITE     = "FFFFFFFF"
HIGHLIGHT = {
    "top":    "FFC6EFCE",
    "bottom": "FFFFC7CE",
    "target": "FFFFEB9C",
    None:     None,
}


def export_excel(
    top_performers:    list[dict],
    dept_report:       dict[str, list[dict]],   # {dept: [rows]}
    dept_trends:       list[dict],
    weekly_summary:    list[dict],
    history:           list[dict],
    settings:          Settings,
    error_log:         ErrorLog,
) -> str:
    """
    Writes a fully styled Excel workbook and returns the file path.
    Returns "" if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.chart import LineChart, BarChart, Reference
        from openpyxl.chart.series import SeriesLabel
    except ImportError:
        print("  [Warning] openpyxl not installed — skipping Excel export.")
        print("            Run: pip install openpyxl")
        return ""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove the default blank sheet

    # ── Build each sheet ──────────────────────────────────────────────────────
    _write_top_performers(wb, top_performers)
    _write_department_performance(wb, dept_report, settings)
    _write_department_trends(wb, dept_trends)
    _write_weekly_summary(wb, weekly_summary)
    _write_historical_data(wb, history)

    # ── Save ─────────────────────────────────────────────────────────────────
    stamp    = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_dir  = settings.get_output_dir()
    out_path = os.path.join(out_dir, f"DPD_Report_{stamp}.xlsx")

    try:
        wb.save(out_path)
        print(f"  ✓  Excel report saved → {out_path}")
        return out_path
    except IOError as e:
        error_log.log("ExportExcel", 0, "export error", str(e), out_path)
        print(f"  ✗  Could not save Excel file: {e}")
        return ""


def export_pdf(
    dept_report: dict[str, list[dict]],
    settings:    Settings,
    error_log:   ErrorLog,
) -> str:
    """
    Exports a PDF summary using matplotlib — works on macOS, Windows, and Linux.
    Returns the saved file path, or "" on failure.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")   # non-interactive backend — no display required
        import matplotlib.pyplot as plt
        import matplotlib.gridspec as gridspec
    except ImportError:
        print("  [Warning] matplotlib not installed — skipping PDF export.")
        print("            Run: pip install matplotlib")
        return ""

    stamp    = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_dir  = settings.get_output_dir()
    out_path = os.path.join(out_dir, f"DPD_Report_{stamp}.pdf")

    from matplotlib.backends.backend_pdf import PdfPages

    try:
        with PdfPages(out_path) as pdf:
            for dept, rows in dept_report.items():
                if not rows:
                    continue
                fig, ax = plt.subplots(figsize=(11, 8.5))
                ax.axis("off")
                ax.set_title(dept, fontsize=16, fontweight="bold", pad=20)

                col_labels = ["Shift", "Rank", "Name", "Avg UPH", "Records"]
                table_data = [
                    [r.get("Shift", ""),
                     r.get("Shift Rank", ""),
                     r.get("Employee Name", ""),
                     f"{r.get('Average UPH', 0):.2f}",
                     r.get("Record Count", "")]
                    for r in rows
                ]

                tbl = ax.table(
                    cellText=table_data,
                    colLabels=col_labels,
                    loc="center",
                    cellLoc="left",
                )
                tbl.auto_set_font_size(False)
                tbl.set_fontsize(9)
                tbl.scale(1, 1.4)

                # Apply row colours
                highlight_map = {"top": "#C6EFCE", "bottom": "#FFC7CE", "target": "#FFEB9C"}
                for i, row in enumerate(rows):
                    hl = row.get("_highlight")
                    if hl and hl in highlight_map:
                        for j in range(len(col_labels)):
                            tbl[i + 1, j].set_facecolor(highlight_map[hl])

                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        print(f"  ✓  PDF report saved → {out_path}")
        return out_path

    except Exception as e:
        error_log.log("ExportPDF", 0, "export error", str(e), out_path)
        print(f"  ✗  Could not save PDF: {e}")
        return ""


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_top_performers(wb, rows: list[dict]):
    if not rows:
        return
    ws = wb.create_sheet("Top Performers")
    headers = ["Rank", "Department", "Shift", "Employee Name", "Average UPH", "Record Count"]
    _write_header_row(ws, headers, row=1)
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.get("Rank", ""))
        ws.cell(i, 2, r.get("Department", ""))
        ws.cell(i, 3, r.get("Shift", ""))
        ws.cell(i, 4, r.get("Employee Name", ""))
        ws.cell(i, 5, round(float(r.get("Average UPH", 0) or 0), 2))
        ws.cell(i, 6, r.get("Record Count", ""))
    _auto_width(ws)


def _write_department_performance(wb, dept_report: dict, settings: Settings):
    ws = wb.create_sheet("Department Performance")
    ws.sheet_properties.tabColor = "1F497D"

    current_row = 1
    for dept, rows in dept_report.items():
        if not rows:
            continue

        # Department header band
        _write_band(ws, current_row, dept, col_span=5)
        current_row += 1

        # Column headers
        col_headers = ["Shift", "Rank", "Employee Name", "Average UPH", "Record Count"]
        _write_header_row(ws, col_headers, row=current_row, fill_hex="FFBDD7EE",
                          font_color="FF1F497D")
        data_start = current_row + 1
        current_row += 1

        # Data rows with highlighting
        for row in rows:
            hl = row.get("_highlight")
            ws.cell(current_row, 1, row.get("Shift", ""))
            ws.cell(current_row, 2, row.get("Shift Rank", ""))
            ws.cell(current_row, 3, row.get("Employee Name", ""))
            uph_cell = ws.cell(current_row, 4, round(float(row.get("Average UPH", 0) or 0), 2))
            uph_cell.number_format = "0.00"
            ws.cell(current_row, 5, row.get("Record Count", ""))

            if hl and HIGHLIGHT.get(hl):
                fill = _solid_fill(HIGHLIGHT[hl])
                for col in range(1, 6):
                    ws.cell(current_row, col).fill = fill

            current_row += 1

        # Top/Bottom summary block
        summary = build_top_bottom_summary(rows, n=3)
        current_row = _write_summary_block(ws, summary, current_row, settings)
        current_row += 3   # gap between departments

    _auto_width(ws)


def _write_department_trends(wb, rows: list[dict]):
    if not rows:
        return
    ws = wb.create_sheet("Department Trends")
    headers = ["Month", "Department", "Average UPH", "Record Count"]
    _write_header_row(ws, headers, row=1)
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.get("Month", ""))
        ws.cell(i, 2, r.get("Department", ""))
        c = ws.cell(i, 3, r.get("Average UPH", 0))
        c.number_format = "0.00"
        ws.cell(i, 4, r.get("Record Count", ""))
    _auto_width(ws)


def _write_weekly_summary(wb, rows: list[dict]):
    if not rows:
        return
    ws = wb.create_sheet("Weekly Summary")
    headers = ["Department", "Week", "Avg UPH", "Total Units", "Record Count"]
    _write_header_row(ws, headers, row=1)
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.get("Department", ""))
        ws.cell(i, 2, r.get("Week", ""))
        c = ws.cell(i, 3, r.get("Avg UPH", 0))
        c.number_format = "0.00"
        ws.cell(i, 4, r.get("Total Units", 0))
        ws.cell(i, 5, r.get("Record Count", 0))
    _auto_width(ws)


def _write_historical_data(wb, rows: list[dict]):
    if not rows:
        return
    ws = wb.create_sheet("Historical Data")
    headers = list(rows[0].keys())
    _write_header_row(ws, headers, row=1)
    for i, row in enumerate(rows, start=2):
        for j, key in enumerate(headers, start=1):
            ws.cell(i, j, row.get(key, ""))
    _auto_width(ws)


def _write_summary_block(ws, summary: dict, start_row: int, settings: Settings) -> int:
    """Write the Top/Bottom summary block and return the next available row."""
    try:
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        return start_row

    # Summary header
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    c = ws.cell(start_row, 1, "Summary")
    c.font = Font(bold=True, color="FFFFFFFF")
    c.fill = _solid_fill("FF1F497D")
    start_row += 1

    for section, label_color in [("top", "FFC6EFCE"), ("bottom", "FFFFC7CE")]:
        section_rows = summary.get(section, [])
        label = f"Top {len(section_rows)}" if section == "top" else f"Bottom {len(section_rows)}"
        label += " performers (by average UPH)"

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        c = ws.cell(start_row, 1, label)
        c.font = Font(bold=True)
        c.fill = _solid_fill(label_color)
        start_row += 1

        for rank, row in enumerate(section_rows, start=1):
            ws.cell(start_row, 1, rank)
            ws.cell(start_row, 3, row.get("Employee Name", ""))
            uph_cell = ws.cell(start_row, 4, round(float(row.get("Average UPH", 0) or 0), 2))
            uph_cell.number_format = "0.00"
            ws.cell(start_row, 5, row.get("Record Count", ""))
            fill = _solid_fill(label_color)
            for col in range(1, 6):
                ws.cell(start_row, col).fill = fill
            start_row += 1

    return start_row


# ── Low-level openpyxl helpers ────────────────────────────────────────────────

def _write_header_row(ws, headers: list[str], row: int,
                      fill_hex: str = "FFC8C8C8", font_color: str = "FF000000"):
    try:
        from openpyxl.styles import PatternFill, Font
        fill = _solid_fill(fill_hex)
        font = Font(bold=True, color=font_color)
    except ImportError:
        fill, font = None, None

    for col, text in enumerate(headers, start=1):
        c = ws.cell(row, col, text)
        if fill:
            c.fill = fill
        if font:
            c.font = font


def _write_band(ws, row: int, text: str, col_span: int = 5):
    """Write a full-width department header band."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
        c = ws.cell(row, 1, text)
        c.font = Font(bold=True, size=13, color="FFFFFFFF")
        c.fill = _solid_fill("FF1F497D")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24
    except ImportError:
        ws.cell(row, 1, text)


def _solid_fill(argb_hex: str):
    try:
        from openpyxl.styles import PatternFill
        # Accept both AARRGGBB and RRGGBB
        h = argb_hex.lstrip("#")
        if len(h) == 6:
            h = "FF" + h
        return PatternFill("solid", fgColor=h)
    except ImportError:
        return None


def _auto_width(ws, max_width: int = 50):
    """Set each column width based on the widest cell value.
    MergedCell objects have no column_letter, so we skip them safely."""
    try:
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return

    for col in ws.columns:
        max_len    = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
