"""
app.py — Productivity Planner
Run with:   streamlit run app.py

╔════════════════════════════════════════════════════════════════════════════════╗
║                         CODE STRUCTURE MAP                                     ║
╚════════════════════════════════════════════════════════════════════════════════╝

IMPORTS & SETUP
  - Standard library imports
  - Streamlit & external dependencies
  - Database & module imports with error handling
  - Data loader imports

CORE CACHE SYSTEM (Lines ~50-120)
  - Tenant ID helpers (_tid)
  - Raw cache functions (_raw_cached_*)
  - Wrapper functions (_cached_*) for clean API

SESSION & AUTH HELPERS (Lines ~120-190)
  - Sign out, cache clearing
  - Success messages, logging
  - Error tracking

EMAIL BACKGROUND SYSTEM (Lines ~190-370)
  - Background email scheduler thread
  - Email logging and sending pipeline

APP INITIALIZATION & STATE (Lines ~370-800)
  - Main setup (_init)
  - Plan management
  - Sidebar navigation (render_sidebar)

CORE BUSINESS LOGIC (Lines ~800-1000)
  - Database availability check
  - Risk calculation (_calc_risk_level_shared) - SHARED ACROSS ALL VIEWS

PAGE FUNCTIONS (Lines ~1000-6000, Ordered by Navigation)
  - page_supervisor()      👔 Supervisor View (daily dashboard)
  - page_dashboard()       📊 Dashboard (risk view)
  - page_import()          📁 Import Data (3-step pipeline)
  - page_employees()       👥 Employees (individual profiles)
  - page_productivity()    📈 Productivity (UPH, trends, coaching, labor cost)
  - page_email()           📧 Email Setup (SMTP, schedules, send)
  - page_settings()        ⚙️ Settings (goals, display, audit)

IMPORT HELPER FUNCTIONS (Lines ~1600-1900)
  - Step 1: Upload CSV files
  - Step 2: Column mapping
  - Step 3: Data pipeline & validation

EMPLOYEE PAGE HELPERS (Lines ~2300-2700)
  - Employee history view
  - AI coaching suggestions
  - Manual coaching notes

REPORT BUILDERS (Lines ~4500-4800)
  - Period report generation
  - HTML email body construction

CRITICAL HELPERS (Lines ~5500-5900)
  - CSV parsing
  - Column auto-detection
  - Access control
  - Login flow

SUBSCRIPTION & CHECKOUT (Lines ~5900-6160)
  - Session timeout checking
  - Stripe integration
  - Subscription page

MAIN ENTRY POINT (Line ~6160)
  - main() function
  - Router logic
  - Startup sequence
"""

import io, os, sys, csv, json, tempfile, traceback, threading, time, math, re, html as _html_mod
from datetime import datetime, date
import pandas as pd

import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ════════════════════════════════════════════════════════════════════════════════
# IMPORTS FROM INTERNAL MODULES
# ════════════════════════════════════════════════════════════════════════════════

# ── Guard: show friendly error if supabase not installed ─────────────────────
try:
    from database import (
        get_employees, get_employee, upsert_employee,
        get_uph_history, get_avg_uph,
        add_coaching_note, get_coaching_notes, delete_coaching_note, archive_coaching_notes,
        get_client as _get_db_client,
        batch_store_uph_history, get_all_uph_history,
        batch_upsert_employees,
    )
    from export_manager import (export_employee)
    DB_AVAILABLE = True
except RuntimeError as e:
    DB_AVAILABLE = False
    DB_ERROR     = str(e)
except ImportError as e:
    DB_AVAILABLE = False
    _ie = str(e)
    if "supabase" in _ie.lower():
        DB_ERROR = "supabase library not installed. Run:  pip3 install supabase"
    else:
        DB_ERROR = f"Import error: {_ie}"
except Exception as e:
    DB_AVAILABLE = False
    DB_ERROR     = f"Unexpected error loading database module: {type(e).__name__}: {e}"

# Backward-compatible DB client shim. Prevents NameError in production if
# _get_db_client was not imported due to an older deployment/import path.
try:
    _ = _get_db_client  # type: ignore[name-defined]  # assignment suppresses Streamlit magic rendering
except NameError:
    def _get_db_client():
        from database import get_client as _db_get_client
        return _db_get_client()

from data_loader import (REQUIRED_FIELDS,
                         auto_detect as _dl_auto_detect,
                         parse_csv_bytes as _dl_parse_csv)


# ════════════════════════════════════════════════════════════════════════════════
# CACHE & SESSION SYSTEM
# ════════════════════════════════════════════════════════════════════════════════

def _tid() -> str:
    return st.session_state.get("tenant_id", "")

@st.cache_data(ttl=120, show_spinner=False)
def _raw_cached_employees(_tid_key: str = ""):
    if not DB_AVAILABLE: return []
    try:
        result = get_employees()
        return result or []
    except Exception: return []

@st.cache_data(ttl=300, show_spinner=False)
def _raw_cached_targets(_tid_key: str = "") -> dict:
    try:
        from goals import get_all_targets
        return get_all_targets()
    except Exception: return {}

@st.cache_data(ttl=60, show_spinner=False)
def _raw_cached_active_flags(_tid_key: str = "") -> dict:
    try:
        from goals import get_active_flags
        return get_active_flags()
    except Exception: return {}

@st.cache_data(ttl=60, show_spinner=False)
def _raw_cached_uph_history(_tid_key: str = ""):
    try: return get_all_uph_history(days=0)
    except Exception: return []

@st.cache_data(ttl=60, show_spinner=False)
def _raw_cached_all_coaching_notes(_tid_key: str = ""):
    try:
        from database import get_client as _get_sb, _tq
        sb = _get_sb()
        r  = _tq(sb.table("coaching_notes").select("emp_id").eq("archived", False)).execute()
        return {row["emp_id"] for row in (r.data or [])}
    except Exception:
        return set()

@st.cache_data(ttl=30, show_spinner=False)
def _raw_cached_coaching_notes_for(emp_id: str, _tid_key: str = "") -> list:
    try: return get_coaching_notes(emp_id)
    except Exception: return []

# ── Tenant-aware wrappers (call these everywhere) ────────────────────────────
def _cached_employees():        return _raw_cached_employees(_tid())
def _cached_targets():          return _raw_cached_targets(_tid())
def _cached_active_flags():     return _raw_cached_active_flags(_tid())
def _cached_uph_history():      return _raw_cached_uph_history(_tid())
def _cached_all_coaching_notes(): return _raw_cached_all_coaching_notes(_tid())
def _cached_coaching_notes_for(emp_id: str): return _raw_cached_coaching_notes_for(emp_id, _tid())


# ════════════════════════════════════════════════════════════════════════════════
# SESSION & STATE MANAGEMENT
# ════════════════════════════════════════════════════════════════════════════════

def _full_sign_out():
    """Wipe ALL session state so the next user starts completely fresh."""
    _bust_cache()
    # Preserve only Streamlit internal keys (prefixed with _ from Streamlit itself)
    # Clear everything else — pipeline data, navigation, uploads, etc.
    keys_to_keep = set()  # nothing to keep
    for k in list(st.session_state.keys()):
        if k not in keys_to_keep:
            del st.session_state[k]


def _bust_cache():
    """Call after any write to force fresh data on next render."""
    _raw_cached_employees.clear()
    _raw_cached_targets.clear()
    _raw_cached_uph_history.clear()
    _raw_cached_all_coaching_notes.clear()
    _raw_cached_active_flags.clear()
    _raw_cached_coaching_notes_for.clear()
    # Also invalidate derived subscription plan cache used by sidebar badges.
    st.session_state.pop("_current_plan", None)


def _success_then_rerun(msg: str, delay: float = 0):
    """Show a success toast and rerun immediately."""
    st.toast(msg, icon="✅")
    st.rerun()


def _tenant_log_path(base_name: str) -> str:
    """Return a tenant-specific log file path."""
    import os as _os
    _dir = _os.path.dirname(__file__)
    tid = st.session_state.get("tenant_id", "")
    if tid:
        return _os.path.join(_dir, f"{base_name}_{tid}.log")
    return _os.path.join(_dir, f"{base_name}.log")

def _audit(action: str, detail: str = ""):
    """Append a timestamped entry to tenant-specific audit log."""
    try:
        from datetime import datetime as _dt
        log_path = _tenant_log_path("dpd_audit")
        entry    = f"{_dt.now().strftime('%Y-%m-%d %H:%M:%S')} | {action} | {detail}\n"
        with open(log_path, "a", encoding="utf-8") as _f:
            _f.write(entry)
    except Exception:
        pass


def _log_app_error(category: str, message: str, detail: str = "",
                   severity: str = "error"):
    """Log an error to the DB error_reports table. Never raises."""
    try:
        from database import log_error
        _email = st.session_state.get("user_email", "")
        log_error(
            category=category,
            message=message,
            detail=detail,
            user_email=_email,
            severity=severity,
        )
    except Exception:
        # Last resort: print to server console
        print(f"[APP_ERROR] [{severity}] [{category}] {message}")


# ════════════════════════════════════════════════════════════════════════════════
# EMAIL & BACKGROUND JOBS
# ════════════════════════════════════════════════════════════════════════════════

# ── Background email scheduler ────────────────────────────────────────────────
# Runs in a daemon thread so emails fire at the scheduled time even when no
# user is actively viewing the app. Started only once via a module-level flag.

def _email_log(msg: str):
    """Append a line to tenant-specific email scheduler log."""
    try:
        import os as _eos
        from datetime import datetime as _edt
        _p = _tenant_log_path("dpd_email_scheduler")
        with open(_p, "a", encoding="utf-8") as _f:
            _f.write(f"{_edt.now().strftime('%Y-%m-%d %H:%M:%S')} | {msg}\n")
    except Exception:
        pass


def _email_timezone_options() -> list[str]:
    return [
        "America/New_York", "America/Chicago", "America/Denver", "America/Phoenix",
        "America/Los_Angeles", "America/Anchorage", "Pacific/Honolulu",
        "America/Toronto", "America/Vancouver", "America/Winnipeg", "America/Halifax",
        "Europe/London", "Europe/Paris", "Europe/Berlin", "Europe/Madrid",
        "Europe/Rome", "Europe/Amsterdam", "Europe/Zurich", "Europe/Stockholm",
        "Europe/Warsaw", "Europe/Istanbul",
        "Asia/Dubai", "Asia/Kolkata", "Asia/Bangkok", "Asia/Singapore",
        "Asia/Shanghai", "Asia/Tokyo", "Asia/Seoul",
        "Australia/Sydney", "Australia/Melbourne", "Pacific/Auckland",
        "UTC",
    ]


def _run_scheduled_reports_for_tenant(tenant_id: str = "", force_now: bool = False,
                                      schedule_names: list[str] | None = None) -> list[dict]:
    """Send scheduled reports for one tenant and return per-schedule results."""
    from email_engine import (
        get_schedules_due_now, get_schedules, send_report_email,
        mark_schedule_sent, load_email_config,
    )
    from settings import Settings
    from goals import load_goals
    from database import get_subscription, get_client as _sched_client

    tid = tenant_id or st.session_state.get("tenant_id", "")
    if not tid:
        return []

    schedule_filter = set(schedule_names or [])
    tz = Settings(tenant_id=tid).get("timezone", "")
    schedules = get_schedules(tid)
    due = [s for s in schedules if s.get("active")] if force_now else get_schedules_due_now(timezone=tz, tenant_id=tid)
    if schedule_filter:
        due = [s for s in due if s.get("name", "") in schedule_filter]
    if not due:
        return []

    tenant_cfg = load_email_config(tenant_id=tid)
    global_recips = [
        r.get("email", "").strip()
        for r in (tenant_cfg.get("recipients") or [])
        if r.get("email")
    ]
    try:
        sb = _sched_client()
        emp_resp = sb.table("employees").select("department").eq("tenant_id", tid).execute()
        depts = sorted({(row.get("department") or "").strip() for row in (emp_resp.data or []) if row.get("department")})
    except Exception:
        depts = []
    targets = (load_goals(tenant_id=tid) or {}).get("dept_targets", {})
    plan_name = ((get_subscription(tid) or {}).get("plan") or "starter").lower()

    results = []
    for sched in due:
        send_to = sched.get("recipients", []) or global_recips
        if not send_to:
            results.append({"name": sched.get("name", ""), "ok": False, "error": "No recipients configured", "recipients": []})
            continue

        period = sched.get("report_period", "Prior day")
        if period == "Custom":
            try:
                d_start = date.fromisoformat(sched.get("date_start", date.today().isoformat()))
                d_end = date.fromisoformat(sched.get("date_end", d_start.isoformat()))
            except Exception:
                d_start, d_end = _resolve_period_dates("Prior day")
        else:
            d_start, d_end = _resolve_period_dates(period)

        xl_data, default_subject, body = _build_period_report(
            d_start, d_end, "All departments", depts, [], targets,
            tenant_id=tid, plan_name=plan_name,
        )
        subject = (sched.get("subject_tpl") or "").strip() or default_subject
        ok, err = send_report_email(send_to, subject, body, xl_data, tenant_id=tid)
        if ok:
            mark_schedule_sent(sched.get("name", ""), timezone=tz, tenant_id=tid)
            _email_log(f"[{tid[:8]}] SENT '{sched.get('name')}' to {send_to} (tz={tz or 'not set'})")
        else:
            _email_log(f"[{tid[:8]}] FAILED '{sched.get('name')}': {err}")
        results.append({
            "name": sched.get("name", ""),
            "ok": ok,
            "error": err,
            "recipients": send_to,
        })
    return results


def _bg_send_scheduled_emails():
    """Called by the background thread every 60 s to fire due email schedules.
    Iterates over ALL tenants with email configs — no st.session_state needed."""
    try:
        from database    import SUPABASE_URL as _BG_URL, SUPABASE_KEY as _BG_KEY
        from supabase    import create_client as _bg_create

        # Create a fresh Supabase client (can't use session-cached one in bg thread)
        sb = _bg_create(_BG_URL, _BG_KEY)

        # Get all tenant email configs that have schedules
        try:
            r = sb.table("tenant_email_config").select("tenant_id, schedules").execute()
            tenant_configs = r.data or []
        except Exception as _e:
            _email_log(f"Could not fetch tenant email configs: {_e}")
            return

        for tc in tenant_configs:
            tid = tc.get("tenant_id", "")
            scheds = tc.get("schedules") or []
            if not tid or not scheds:
                continue

            try:
                results = _run_scheduled_reports_for_tenant(tenant_id=tid, force_now=False)
                if not results and scheds:
                    _email_log(f"[{tid[:8]}] Has {len(scheds)} schedule(s) but none are due now")
            except Exception as _te2:
                _email_log(f"[{tid[:8]}] Tenant error: {_te2}")
                try:
                    from database import log_error
                    log_error("email", f"Tenant processing error: {_te2}",
                              severity="error", tenant_id=tid)
                except Exception:
                    pass
    except Exception as _te:
        _email_log(f"Thread error: {_te}")


_email_thread_started = False

def _start_email_thread():
    """Start a background thread that checks email schedules every 60 s. Idempotent."""
    global _email_thread_started
    if _email_thread_started:
        return
    _email_thread_started = True

    import time as _time

    def _loop():
        _time.sleep(30)   # brief startup delay so the app finishes loading first
        _email_log("Background email thread started")
        while True:
            try:
                _bg_send_scheduled_emails()
            except Exception as _e:
                _email_log(f"Unhandled loop error: {_e}")
            _time.sleep(60)

    t = threading.Thread(target=_loop, daemon=True, name="email-scheduler")
    t.start()
    return t


# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title = "Productivity Planner",
    page_icon  = "📦",
    layout     = "wide",
    initial_sidebar_state = "expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────

st.html("""<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  /* ── Global font ── */
  html, body, [class*="css"], .stApp { font-family: 'Inter', sans-serif !important; }

  /* ── Hide Streamlit chrome ── */
  #MainMenu, footer,
  [data-testid="stDecoration"], [data-testid="stStatusWidget"] { visibility: hidden !important; display: none !important; }
  /* Keep header visible so the sidebar toggle button works */
  header[data-testid="stHeader"] { background: transparent !important; }

  /* ── Page background ── */
  .stApp { background: #F7F9FC !important; }
  .main .block-container { padding-top: 1.8rem; padding-bottom: 3rem; max-width: 1200px; }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] { background: #0F2D52 !important; border-right: none !important; }
  [data-testid="stSidebar"] > div { background: #0F2D52 !important; }
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] div { color: #CBD8E8 !important; }
  [data-testid="stSidebar"] .stRadio label { font-size: 13px !important; padding: 6px 0 !important; font-weight: 500 !important; }
  [data-testid="stSidebar"] hr { border-color: #1A4A8A !important; opacity: 0.6; }

  /* ── Headings ── */
  h1 { color: #0F2D52 !important; font-weight: 600 !important; font-size: 1.6rem !important; letter-spacing: -0.02em; }
  h2 { color: #0F2D52 !important; font-weight: 600 !important; }
  h3 { color: #1A4A8A !important; font-weight: 500 !important; }
  /* Subheader (used heavily in the app) */
  [data-testid="stHeading"] { color: #0F2D52 !important; }

  /* ── Body text — force dark readable colour everywhere outside sidebar ── */
  .main p, .main span, .main label, .main div,
  .block-container p, .block-container span, .block-container label {
    color: #1A2D42;
  }
  .main .stCaption, .main small, .main caption { color: #5A7A9C !important; }

  /* ── Buttons — covers both st.button and st.form_submit_button ── */
  .stButton > button,
  [data-testid="stFormSubmitButton"] > button {
    border-radius: 6px !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    transition: background 0.15s, color 0.15s;
  }
  /* ── ALL buttons: force readable text regardless of Streamlit version ── */
  /* Targets every possible button selector Streamlit uses */
  .stButton > button,
  [data-testid="stFormSubmitButton"] > button,
  [data-testid="stBaseButton-primary"],
  [data-testid="stBaseButton-secondary"] {
    border-radius: 6px !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    transition: background 0.15s, color 0.15s;
  }

  /* Primary: navy bg, ALWAYS white text */
  .stButton > button[kind="primary"],
  [data-testid="stFormSubmitButton"] > button,
  [data-testid="stFormSubmitButton"] > button[kind="primary"],
  button[data-testid="stBaseButton-primary"] {
    background-color: #0F2D52 !important;
    color: #ffffff !important;
    border: none !important;
  }
  .stButton > button[kind="primary"]:hover,
  [data-testid="stFormSubmitButton"] > button:hover,
  button[data-testid="stBaseButton-primary"]:hover {
    background-color: #1A4A8A !important;
    color: #ffffff !important;
  }
  /* Force white on the inner p/span inside primary buttons */
  .stButton > button[kind="primary"] p,
  .stButton > button[kind="primary"] span,
  [data-testid="stFormSubmitButton"] > button p,
  [data-testid="stFormSubmitButton"] > button span,
  button[data-testid="stBaseButton-primary"] p,
  button[data-testid="stBaseButton-primary"] span {
    color: #ffffff !important;
  }

  /* Secondary: white bg, navy text */
  .stButton > button[kind="secondary"],
  .stButton > button:not([kind="primary"]),
  button[data-testid="stBaseButton-secondary"] {
    background-color: #ffffff !important;
    color: #0F2D52 !important;
    border: 1px solid #C5D4E8 !important;
  }
  .stButton > button[kind="secondary"]:hover,
  .stButton > button:not([kind="primary"]):hover,
  button[data-testid="stBaseButton-secondary"]:hover {
    background-color: #E8F0F9 !important;
    color: #0F2D52 !important;
    border-color: #1A4A8A !important;
  }
  .stButton > button:not([kind="primary"]) p,
  .stButton > button:not([kind="primary"]) span,
  button[data-testid="stBaseButton-secondary"] p,
  button[data-testid="stBaseButton-secondary"] span {
    color: #0F2D52 !important;
  }

  /* ── Tabs ── */
  .stTabs [data-baseweb="tab-list"] { border-bottom: 2px solid #E2EBF4 !important; gap: 2px; }
  .stTabs [data-baseweb="tab"] {
    font-size: 12px !important; font-weight: 500 !important;
    color: #5A7A9C !important; padding: 7px 14px !important;
    border-radius: 4px 4px 0 0 !important; background: transparent !important;
  }
  .stTabs [aria-selected="true"] {
    color: #0F2D52 !important; border-bottom: 2px solid #0F2D52 !important;
    background: transparent !important;
  }

  /* ── Metrics ── */
  [data-testid="stMetric"] { background: #ffffff; border: 1px solid #E2EBF4; border-radius: 8px; padding: 14px 18px; }
  [data-testid="stMetricLabel"] > div { font-size: 11px !important; font-weight: 600 !important; color: #5A7A9C !important; text-transform: uppercase; letter-spacing: 0.05em; }
  [data-testid="stMetricValue"] > div { font-size: 26px !important; font-weight: 600 !important; color: #0F2D52 !important; }
  [data-testid="stMetricDelta"] { font-size: 12px !important; }

  /* ── Alert / notification boxes ── */
  /* SUCCESS — green background, dark green text */
  [data-testid="stAlert"][data-baseweb="notification"][kind="positive"],
  div.stSuccess, div.element-container div.stAlert.stSuccess,
  [data-testid="stNotification"][kind="success"] {
    background-color: #EAF4EC !important;
    border-left: 4px solid #2E7D32 !important;
    border-radius: 0 8px 8px 0 !important;
  }
  div.stSuccess *, div.stSuccess p, div.stSuccess span,
  [data-testid="stAlert"][data-baseweb="notification"][kind="positive"] *,
  [data-testid="stAlert"][data-baseweb="notification"][kind="positive"] p { color: #1B5E20 !important; }

  /* WARNING — pale orange background, dark orange text — NO yellow-on-yellow */
  [data-testid="stAlert"][data-baseweb="notification"][kind="warning"],
  div.stWarning, div.element-container div.stAlert.stWarning,
  [data-testid="stNotification"][kind="warning"] {
    background-color: #FFF3E0 !important;
    border-left: 4px solid #E65100 !important;
    border-radius: 0 8px 8px 0 !important;
  }
  div.stWarning *, div.stWarning p, div.stWarning span,
  [data-testid="stAlert"][data-baseweb="notification"][kind="warning"] *,
  [data-testid="stAlert"][data-baseweb="notification"][kind="warning"] p { color: #BF360C !important; }

  /* ERROR — pale red background, dark red text */
  [data-testid="stAlert"][data-baseweb="notification"][kind="error"],
  div.stError, div.element-container div.stAlert.stError,
  [data-testid="stNotification"][kind="error"] {
    background-color: #FDECEA !important;
    border-left: 4px solid #C62828 !important;
    border-radius: 0 8px 8px 0 !important;
  }
  div.stError *, div.stError p, div.stError span,
  [data-testid="stAlert"][data-baseweb="notification"][kind="error"] *,
  [data-testid="stAlert"][data-baseweb="notification"][kind="error"] p { color: #7F0000 !important; }

  /* INFO — pale blue background, dark navy text */
  [data-testid="stAlert"][data-baseweb="notification"][kind="info"],
  div.stInfo, div.element-container div.stAlert.stInfo,
  [data-testid="stNotification"][kind="info"] {
    background-color: #E8F0F9 !important;
    border-left: 4px solid #0F2D52 !important;
    border-radius: 0 8px 8px 0 !important;
  }
  div.stInfo *, div.stInfo p, div.stInfo span,
  [data-testid="stAlert"][data-baseweb="notification"][kind="info"] *,
  [data-testid="stAlert"][data-baseweb="notification"][kind="info"] p { color: #0D2440 !important; }

  /* Broad fallback: any alert icon or text should never be invisible */
  [data-testid="stAlert"] svg { opacity: 0.8; }

  /* ── Inputs — white background, dark text, no yellow tint ── */
  .stTextInput input, .stTextInput textarea,
  .stNumberInput input,
  .stTextArea textarea {
    background: #ffffff !important;
    color: #1A2D42 !important;
    border: 1px solid #C5D4E8 !important;
    border-radius: 6px !important;
  }
  .stTextInput input::placeholder,
  .stTextArea textarea::placeholder { color: #8AAAC4 !important; }
  .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
    border-color: #1A4A8A !important;
    box-shadow: 0 0 0 2px rgba(26,74,138,0.12) !important;
    caret-color: #0F2D52 !important;
  }
  .stTextInput input, .stNumberInput input {
    caret-color: #0F2D52 !important;
  }

  /* ── Selectbox ── */
  [data-baseweb="select"] > div {
    background: #ffffff !important;
    border-color: #C5D4E8 !important;
    border-radius: 6px !important;
    color: #1A2D42 !important;
  }
  [data-baseweb="select"] span { color: #1A2D42 !important; }
  [data-baseweb="popover"] { background: #ffffff !important; }
  [data-baseweb="menu"] li { color: #1A2D42 !important; background: #ffffff !important; }
  [data-baseweb="menu"] li:hover { background: #E8F0F9 !important; }

  /* ── Radio buttons ── */
  .stRadio > label { color: #1A2D42 !important; }
  .stRadio [data-baseweb="radio"] span { color: #1A2D42 !important; }

  /* ── Checkbox / Toggle ── */
  .stCheckbox label, .stToggle label { color: #1A2D42 !important; }

  /* ── Slider ── */
  .stSlider label { color: #1A2D42 !important; }
  .stSlider [data-baseweb="slider"] [role="slider"] { background: #0F2D52 !important; }

  /* ── Number input label ── */
  .stNumberInput label { color: #1A2D42 !important; }

  /* ── Expander ── */
  [data-testid="stExpander"] {
    border: 1px solid #E2EBF4 !important;
    border-radius: 8px !important;
    background: #ffffff !important;
  }
  /* Collapsed header */
  [data-testid="stExpander"] summary,
  [data-testid="stExpander"] summary span,
  [data-testid="stExpander"] summary p { color: #0F2D52 !important; font-weight: 500 !important; }
  [data-testid="stExpander"] summary:hover { background: #F0F5FB !important; }
  /* Hover on open expander: keep light blue bg but force text BLACK (not white) */
  [data-testid="stExpander"][open] summary:hover,
  [data-testid="stExpander"][open] summary:hover *,
  [data-testid="stExpander"][open] summary:hover span,
  [data-testid="stExpander"][open] summary:hover p,
  [data-testid="stExpander"][open] summary:hover strong,
  [data-testid="stExpander"][open] summary:hover b { color: #0F2D52 !important; }
  details[open] > summary:hover,
  details[open] > summary:hover * { color: #0F2D52 !important; }
  /* Open state — header row goes dark blue; force its text white */
  [data-testid="stExpander"][open] summary,
  [data-testid="stExpander"][open] summary span,
  [data-testid="stExpander"][open] summary p,
  [data-testid="stExpander"][open] summary strong,
  [data-testid="stExpander"][open] summary b,
  [data-testid="stExpander"][open] summary *,
  details[open] > summary,
  details[open] > summary * { color: #ffffff !important; }
  /* ALL content inside expanders — always dark readable text */
  [data-testid="stExpander"] p,
  [data-testid="stExpander"] span,
  [data-testid="stExpander"] td,
  [data-testid="stExpander"] th,
  [data-testid="stExpander"] li,
  [data-testid="stExpander"] code,
  [data-testid="stExpander"] .stMarkdown p,
  [data-testid="stExpander"] .stMarkdown span { color: #1A2D42 !important; }
  /* Markdown tables inside expanders */
  [data-testid="stExpander"] table td,
  [data-testid="stExpander"] table th { color: #1A2D42 !important; background: transparent !important; }

  /* ── File uploader ── */
  [data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 2px dashed #C5D4E8 !important;
    border-radius: 8px !important;
  }
  [data-testid="stFileUploader"] span,
  [data-testid="stFileUploader"] p,
  [data-testid="stFileUploader"] small,
  [data-testid="stFileUploader"] div { color: #1A2D42 !important; }
  [data-testid="stFileUploaderDropzoneInstructions"],
  [data-testid="stFileUploaderDropzoneInstructions"] * { color: #5A7A9C !important; }
  /* Dropzone instruction text — "Drag and drop files here · Limit 200MB" */
  [data-testid="stFileUploaderDropzone"],
  [data-testid="stFileUploaderDropzone"] small,
  [data-testid="stFileUploaderDropzone"] span,
  [data-testid="stFileUploaderDropzone"] p,
  [data-testid="stFileUploaderDropzone"] div,
  [data-testid="stFileUploaderDropzone"] label,
  [data-testid="stFileUploaderDropzone"] button span { color: #5A7A9C !important; }
  /* The specific "Drag and drop" + "Limit" lines use these internal classes */
  .stFileUploaderDropzone span { color: #5A7A9C !important; }
  [data-baseweb="file-uploader"] span,
  [data-baseweb="file-uploader"] p,
  [data-baseweb="file-uploader"] small { color: #5A7A9C !important; }
  /* Uploaded file list — filename and size text */
  [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
  [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] span,
  [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] p,
  [data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] small { color: #0F2D52 !important; }
  /* The file name specifically */
  [data-testid="stFileUploaderFileName"] { color: #0F2D52 !important; font-weight: 500 !important; }
  [data-testid="stFileUploaderFileSize"] { color: #5A7A9C !important; }
  /* Catch-all: any text inside the uploader that might go white */
  section[data-testid="stFileUploader"] * { color: #1A2D42 !important; }
  section[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzoneInstructions"] * { color: #5A7A9C !important; }

  /* ── Dataframe ── */
  [data-testid="stDataFrame"] { border: 1px solid #E2EBF4 !important; border-radius: 8px !important; }

  /* ── Progress bar ── */
  .stProgress > div > div { background: #0F2D52 !important; border-radius: 4px; }
  /* Progress bar status text */
  [data-testid="stProgress"] p,
  [data-testid="stProgress"] span,
  .stProgress p,
  .stProgress span,
  .stProgress ~ div p,
  .stProgress ~ div span { color: #ffffff !important; }
  /* The text label that sits above/below the bar */
  [data-testid="stProgressBar"] + div,
  [data-testid="stProgressBar"] + div p,
  [data-testid="stProgressBar"] + div span { color: #ffffff !important; }

  /* ── Date input ── */
  .stDateInput input { background: #ffffff !important; color: #1A2D42 !important; border: 1px solid #C5D4E8 !important; border-radius: 6px !important; }
  .stDateInput label { color: #1A2D42 !important; }

  /* ── Divider ── */
  hr { border-color: #E2EBF4 !important; }

  /* ── Caption text everywhere outside sidebar ── */
  .main .stCaption p { color: #5A7A9C !important; }

  /* ── Multiselect ── */
  [data-baseweb="tag"] { background: #E8F0F9 !important; }
  [data-baseweb="tag"] span { color: #0F2D52 !important; }

  /* ── Time input ── */
  .stTimeInput input { background: #ffffff !important; color: #1A2D42 !important; }
</style>
""")


# ── Session state ─────────────────────────────────────────────────────────────

def _init():
    defaults = {
        "uploaded_sessions":  [],   # list of {filename, rows, mapping, timestamp}
        "submission_plan":    None, # pending plan awaiting confirmation
        "split_overrides":    {},   # emp_id -> list of split assignments
        # Step navigation
        "import_step":        1,    # 1=upload 2=map(optional) 3=pipeline
        "alloc_rows":         [],   # processed rows from pipeline
        "alloc_date":         None, # work date from pipeline
        "chart_months":       12,
        "trend_weeks":        4,
        "target_uph":         0.0,
        "top_pct":            10,
        "bot_pct":            10,
        "smart_merge":        True,
        # Productivity pipeline state (kept for backward compat)
        "raw_rows":           [],
        "csv_headers":        [],
        "mapping":            {},
        "mapping_ready":      False,
        "history":            [],
        "top_performers":     [],
        "dept_report":        {},
        "dept_trends":        [],
        "weekly_summary":     [],
        "employee_rolling_avg": [],
        "employee_risk":       [],
        "goal_status":        [],
        "trend_data":         {},
        "pipeline_done":      False,
        "pipeline_error":     "",
        "warnings":           [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()


# ── Sidebar ───────────────────────────────────────────────────────────────────

def _get_current_plan() -> str:
    """Return current subscription plan: 'starter', 'pro', 'business', or 'admin'."""
    # Check admin bypass first
    _user_email = st.session_state.get("user_email", "").lower()
    try:
        _admin_emails = [e.strip().lower() for e in st.secrets.get("ADMIN_EMAILS", "").split(",") if e.strip()]
        if _user_email in _admin_emails:
            return "admin"
    except Exception:
        pass
    # Refresh cached plan periodically so open tabs don't retain stale access.
    _cached_plan = st.session_state.get("_current_plan")
    _cached_ts = float(st.session_state.get("_current_plan_ts", 0) or 0)
    if _cached_plan and (time.time() - _cached_ts) < 60:
        return _cached_plan
    try:
        from database import get_subscription
        sub = get_subscription()
        plan = sub.get("plan", "starter") if sub else "starter"
        st.session_state["_current_plan"] = plan
        st.session_state["_current_plan_ts"] = time.time()
        return plan
    except Exception:
        return "starter"


def _plan_rank(plan: str) -> int:
    return {"starter": 1, "pro": 2, "business": 3, "admin": 99}.get((plan or "").lower(), 1)


def _has_plan(min_plan: str) -> bool:
    return _plan_rank(_get_current_plan()) >= _plan_rank(min_plan)


def _plan_gate(min_plan: str, feature_name: str) -> bool:
    if _has_plan(min_plan):
        return True
    st.info(f"{feature_name} is available on **{min_plan.capitalize()}** and above.")
    st.caption("Upgrade in Settings → Subscription to unlock this feature.")
    return False


def render_sidebar() -> str:
    _plan = _get_current_plan()

    with st.sidebar:
        st.markdown("""
<div style="padding:8px 0 20px;">
  <div style="font-size:19px;font-weight:700;color:#fff;letter-spacing:-.02em;line-height:1.15;">
    📦 Productivity<br>Planner
  </div>
</div>""", unsafe_allow_html=True)

        st.divider()

        # Build nav with stable internal keys so routing does not depend on
        # emoji rendering/encoding.
        nav_items = [
            ("supervisor", "👔  Supervisor"),
            ("dashboard", "📊  Dashboard"),
            ("import", "📁  Import Data"),
            ("employees", "👥  Employees"),
            ("productivity", "📈  Productivity"),
            ("email", "📧  Email Setup"),
        ]
        nav_items.append(("settings", "⚙️  Settings"))

        nav_keys = [k for k, _ in nav_items]
        nav_labels = {k: lbl for k, lbl in nav_items}

        goto = str(st.session_state.pop("goto_page", "") or "").lower()
        if goto:
            if goto in nav_keys:
                st.session_state["_current_page_key"] = goto
            elif "supervisor" in goto:
                st.session_state["_current_page_key"] = "supervisor"
            elif "dashboard" in goto:
                st.session_state["_current_page_key"] = "dashboard"
            elif "import" in goto:
                st.session_state["_current_page_key"] = "import"
            elif "employee" in goto:
                st.session_state["_current_page_key"] = "employees"
            elif "productivity" in goto:
                st.session_state["_current_page_key"] = "productivity"
            elif "email" in goto:
                st.session_state["_current_page_key"] = "email"
            elif "setting" in goto:
                st.session_state["_current_page_key"] = "settings"

        current_key = st.session_state.get("_current_page_key", nav_keys[0])
        if current_key not in nav_keys:
            current_key = nav_keys[0]

        page = st.radio(
            "Navigation",
            nav_keys,
            index=nav_keys.index(current_key),
            format_func=lambda k: nav_labels.get(k, k.title()),
            label_visibility="collapsed",
        )
        st.session_state["_current_page_key"] = page

        st.divider()
        if st.button("↺ Refresh data", use_container_width=True, key="sb_refresh"):
            _bust_cache()
            st.rerun()

        # ── Plan badge ──────────────────────────────────────────────────
        _plan_display = _plan.capitalize() if _plan != "admin" else "Admin"
        _plan_color = {"starter": "#888", "pro": "#1E90FF", "business": "#FFD700", "admin": "#FF6347"}.get(_plan, "#888")
        st.markdown(
            f'<div style="font-size:10px;color:{_plan_color};font-weight:700;margin-bottom:4px;">'
            f'Plan: {_plan_display}</div>',
            unsafe_allow_html=True,
        )
        try:
            from database import get_employee_count, get_employee_limit
            _emp_count = get_employee_count()
            _emp_limit = get_employee_limit()
            _limit_str = "unlimited" if _emp_limit == -1 else str(_emp_limit)
            st.markdown(
                f'<div style="font-size:10px;color:#7FA8CC;margin-bottom:8px;">'
                f'Employees: {_emp_count}/{_limit_str}</div>',
                unsafe_allow_html=True,
            )
        except Exception:
            pass

        # ── User info + logout ─────────────────────────────────────────────
        user_name = st.session_state.get("user_name", "")
        if user_name:
            _safe_name = _html_mod.escape(user_name)
            st.markdown(
                f'<div style="font-size:11px;color:#7FA8CC;margin-bottom:4px;">'
                f'Signed in as<br><b style="color:#CBD8E8;">{_safe_name}</b></div>',
                unsafe_allow_html=True,
            )
        if st.button("Sign out", use_container_width=True, key="sb_signout"):
            _full_sign_out()
            st.rerun()

        st.markdown(
            '<div style="font-size:10px;color:#3D5A7A;line-height:1.7;">'
            'Productivity Planner · v3.0</div>',
            unsafe_allow_html=True,
        )
    return page


# ── DB gate ───────────────────────────────────────────────────────────────────

def require_db() -> bool:
    if not DB_AVAILABLE:
        st.error(f"Database not available: {DB_ERROR}")
        st.info("Run `pip3 install supabase` in your terminal then restart the app.")
        return False
    return True


# ── Risk calculation (shared by Supervisor View and Dashboard) ────────────────

def _calc_risk_level(emp, history):
    """Calculate performance risk level: 🔴 High, 🟡 Medium, 🟢 Low.
    Combines: trend + under-goal streak + variance.
    Returns: (risk_level, risk_score, details_dict)
    
    Used by: page_supervisor, page_dashboard, page_productivity.
    """
    risk_score = 0.0
    details = {
        "trend_score": 0,
        "streak_score": 0,
        "variance_score": 0,
    }
    
    trend = emp.get("trend", "insufficient_data")
    if trend == "down":
        risk_score += 4
        details["trend_score"] = 4
    elif trend == "flat":
        risk_score += 1
        details["trend_score"] = 1
    elif trend == "up":
        risk_score -= 2
        details["trend_score"] = -2
    
    emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
    target_uph = emp.get("Target UPH", "—")
    under_goal_streak = 0
    
    if history and target_uph != "—":
        try:
            target = float(target_uph)
            emp_history = [r for r in history
                           if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
            if emp_history:
                sorted_hist = sorted(emp_history,
                                    key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                for r in reversed(sorted_hist):
                    try:
                        uph_val = float(r.get("UPH", 0) or 0)
                        if uph_val < target:
                            under_goal_streak += 1
                        else:
                            break
                    except (ValueError, TypeError):
                        break
        except (ValueError, TypeError):
            pass
    
    if under_goal_streak >= 7:
        risk_score += 5
        details["streak_score"] = 5
    elif under_goal_streak >= 5:
        risk_score += 4
        details["streak_score"] = 4
    elif under_goal_streak >= 3:
        risk_score += 2.5
        details["streak_score"] = 2.5
    elif under_goal_streak >= 1:
        risk_score += 0.5
        details["streak_score"] = 0.5
    
    variance = 0.0
    uph_values = []
    if history:
        emp_history = [r for r in history
                       if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
        for r in emp_history:
            try:
                uph_val = float(r.get("UPH", 0) or 0)
                if uph_val > 0:
                    uph_values.append(uph_val)
            except (ValueError, TypeError):
                pass
    
    if len(uph_values) >= 3:
        avg_uph = sum(uph_values) / len(uph_values)
        variance = sum((x - avg_uph) ** 2 for x in uph_values) / len(uph_values)
        std_dev = variance ** 0.5
        coeff_variation = (std_dev / avg_uph * 100) if avg_uph > 0 else 0
        
        if coeff_variation > 30:
            risk_score += 3
            details["variance_score"] = 3
        elif coeff_variation > 20:
            risk_score += 1.5
            details["variance_score"] = 1.5
        elif coeff_variation > 10:
            risk_score += 0.5
            details["variance_score"] = 0.5
        
        details["variance_pct"] = round(coeff_variation, 1)
    
    details["under_goal_streak"] = under_goal_streak
    details["total_score"] = round(risk_score, 1)
    
    if risk_score >= 7:
        return "🔴 High", risk_score, details
    elif risk_score >= 4:
        return "🟡 Medium", risk_score, details
    else:
        return "🟢 Low", risk_score, details


# Alias: some code paths call this as _calc_risk_level_shared
_calc_risk_level_shared = _calc_risk_level


# ════════════════════════════════════════════════════════════════════════════════
# PAGE: SUPERVISOR VIEW
# Daily one-screen view: department health, top risks, trending alerts, actions
# CRITICAL: This is the primary daily-use screen
# ════════════════════════════════════════════════════════════════════════════════

def page_supervisor():
    """One-screen supervisor view: risks, trends, context, actions. Daily-use focused."""
    st.title("👔 Supervisor View")
    st.caption("One-screen daily overview: team health, top risks, trends, and next steps.")

    if not require_db(): return

    # Load data
    if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
        with st.spinner("Loading supervisor data…"):
            try:
                from app import _build_archived_productivity
                _build_archived_productivity()
            except Exception:
                pass

    gs = st.session_state.get("goal_status", [])
    history = st.session_state.get("history", [])
    
    if not gs:
        st.info("No productivity data. Run Import Data to get started.")
        return

    # ────────────────────────────────────────────────────────────────────────────
    # SECTION 1: DEPARTMENT HEALTH SUMMARY
    # ────────────────────────────────────────────────────────────────────────────
    
    st.subheader("📊 Team Health by Department")
    depts = sorted(set(r.get("Department", "") for r in gs if r.get("Department")))
    dept_cols = st.columns(len(depts) if depts else 1)
    
    dept_summary = {}
    for dept in depts:
        dept_emps = [r for r in gs if r.get("Department") == dept]
        on_goal = len([r for r in dept_emps if r.get("goal_status") == "on_goal"])
        below_goal = len([r for r in dept_emps if r.get("goal_status") == "below_goal"])
        total = len(dept_emps)
        health_pct = round((on_goal / total * 100) if total > 0 else 0)
        
        dept_summary[dept] = {
            "on_goal": on_goal,
            "below_goal": below_goal,
            "total": total,
            "health_pct": health_pct,
        }
    
    for col, dept in zip(dept_cols, depts):
        s = dept_summary[dept]
        health_color = "#28a745" if s["health_pct"] >= 75 else "#ffc107" if s["health_pct"] >= 50 else "#dc3545"
        col.metric(
            f"**{dept}**",
            f"{s['on_goal']}/{s['total']} on goal",
            f"{s['health_pct']}%",
            label_visibility="visible"
        )
    
    st.divider()
    
    # ────────────────────────────────────────────────────────────────────────────
    # SECTION 2: TOP RISKS (5-10 employees)
    # ────────────────────────────────────────────────────────────────────────────
    
    st.subheader("🔴 Top Risks — Action Required Today")
    
    # Filter for below_goal and calculate risk
    below_goal = [r for r in gs if r.get("goal_status") == "below_goal"]
    
    if not below_goal:
        st.success("✅ All employees meeting goals!")
    else:
        risk_list = []
        for emp in below_goal:
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            risk_level, risk_score, risk_details = _calc_risk_level_shared(emp, history)
            
            # Get flagged info for context tags
            flagged_info = {}
            try:
                from goals import get_employee_flags
                all_flags = get_employee_flags()
                flagged_info = next((f for f in all_flags if f.get("emp_id") == emp_id), {})
            except Exception:
                pass
            
            risk_list.append({
                "name": emp.get("Employee", emp.get("Employee Name", "Unknown")),
                "department": emp.get("Department", ""),
                "emp_id": emp_id,
                "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
                "target_uph": emp.get("Target UPH", "—"),
                "trend": emp.get("trend", "—"),
                "change_pct": emp.get("change_pct", 0.0),
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_details": risk_details,
                "context_tags": flagged_info.get("context_tags", []),
            })
        
        # Sort by risk score (highest first) and take top 10
        risk_list.sort(key=lambda x: x["risk_score"], reverse=True)
        top_risks = risk_list[:10]
        
        # Display each top risk
        for i, emp in enumerate(top_risks, 1):
            try:
                target = float(emp["target_uph"]) if emp["target_uph"] != "—" else 0
                gap = target - emp["avg_uph"]
                gap_str = f"-{gap:.1f} UPH" if gap > 0 else "—"
            except (ValueError, TypeError):
                gap_str = "—"
            
            # Build summary row
            col1, col2, col3, col4, col5 = st.columns([1, 2, 1.5, 1, 2])
            col1.write(f"**#{i}**")
            col2.write(f"**{emp['name']}** · {emp['department']}")
            col3.write(f"{emp['risk_level']} ({emp['risk_score']:.1f})")
            col4.write(f"{emp['avg_uph']:.1f} / {emp['target_uph']}")
            col5.write(f"{emp['trend']} {emp['change_pct']:+.0f}%")
            
            # Expandable details
            with st.expander(f"View context & actions for {emp['name']}", expanded=False):
                # Context tags
                if emp['context_tags']:
                    tag_cols = st.columns(len(emp['context_tags']))
                    for tag_col, tag in zip(tag_cols, emp['context_tags']):
                        tag_col.markdown(f"📌 {tag}")
                    st.write("")
                
                # Risk breakdown
                d = emp['risk_details']
                rc1, rc2, rc3 = st.columns(3)
                rc1.metric("Trend", f"{d.get('trend_score', 0):+.0f}pt", "down trend" if d.get("trend_score", 0) >= 4 else "")
                rc2.metric("Streak", f"{d.get('streak_score', 0):.1f}pt", f"{d.get('under_goal_streak', 0)} days below")
                rc3.metric("Variance", f"{d.get('variance_score', 0):.1f}pt", f"{d.get('variance_pct', 0):.0f}% CV")
                
                st.write("")
                st.markdown("**👉 Suggested Next Actions:**")
                
                # Build context-aware actions
                actions = []
                
                # Context-specific actions (take priority)
                if "Equipment issues" in emp['context_tags']:
                    actions.append("🔧 **Fix the equipment first** — Resolve tool/system issues before coaching.")
                if "New employee" in emp['context_tags']:
                    actions.append("📋 **Structured onboarding check** — Ensure proper training & mentoring.")
                if "Cross-training" in emp['context_tags']:
                    actions.append("🎓 **Support transition** — Provide extra mentoring during skill-building.")
                if "Shift change" in emp['context_tags']:
                    actions.append("⏰ **Allow adjustment time** — Follow up in 2 weeks for impact.")
                if "Short staffed" in emp['context_tags']:
                    actions.append("👥 **Increase capacity** — Hiring or redistribution may help faster.")
                
                # Trend-based actions
                if not any(tag in emp['context_tags'] for tag in ["Equipment issues", "Short staffed"]):
                    if emp["trend"] == "down":
                        actions.append("💬 **Identify obstacles** — Ask what's changed. Check for workload/personal issues.")
                    elif emp["trend"] == "flat":
                        actions.append("📈 **Break the plateau** — Try new training, task rotation, or peer mentoring.")
                
                # Gap-based actions
                try:
                    target = float(emp["target_uph"]) if emp["target_uph"] != "—" else 0
                    if target > 0:
                        gap = target - emp["avg_uph"]
                        if gap > 5 and "New employee" not in emp['context_tags']:
                            actions.append("📊 **Major gap** — Structured improvement plan with weekly check-ins.")
                        elif gap > 2 and "New employee" not in emp['context_tags']:
                            actions.append("🤝 **1-on-1 coaching** — Discuss goals and what support they need.")
                except (ValueError, TypeError):
                    pass
                
                # Default if no actions were generated
                if not actions:
                    actions.append("🤝 **1-on-1 conversation** — Discuss performance, barriers, and support.")
                
                for action in actions:
                    st.write(action)
    
    st.divider()
    
    # ────────────────────────────────────────────────────────────────────────────
    # SECTION 3: TRENDING DOWN ALERTS
    # ────────────────────────────────────────────────────────────────────────────
    
    st.subheader("⚠️ Trending Down — Proactive Check-In Recommended")
    trending_down = [r for r in gs if r.get("trend") == "down" and r.get("goal_status") != "below_goal"]
    
    if trending_down:
        st.caption(f"⚠️ {len(trending_down)} employee(s) showing downward trend but NOT yet below goal.")
        for emp in trending_down[:5]:  # Show top 5 only
            col1, col2, col3 = st.columns([2, 1, 2])
            col1.write(f"**{emp.get('Employee Name', 'Unknown')}** ({emp.get('Department', '')})")
            col2.write(f"↓ {emp.get('change_pct', 0):+.0f}%")
            col3.write(f"{emp.get('Average UPH', 0):.1f} / {emp.get('Target UPH', '—')}")
    else:
        st.success("✅ No concerning trends detected.")
    
    st.divider()
    
    # ────────────────────────────────────────────────────────────────────────────
    # SECTION 4: BUSINESS IMPACT (What it costs to not coach)
    # ────────────────────────────────────────────────────────────────────────────
    
    st.subheader("💰 This Week's Cost Impact")
    
    # Try to get hourly wage from settings
    try:
        from settings import Settings as _ImpactS
        _avg_wage = _ImpactS().get("avg_hourly_wage", 18.0)
    except Exception:
        _avg_wage = 18.0
    
    # Calculate labor cost for below-goal employees
    cost_by_emp = []
    for emp in below_goal:
        emp_name = emp.get("Employee Name", emp.get("Employee", "Unknown"))
        target = emp.get("Target UPH", "—")
        avg_uph = emp.get("Average UPH", 0)
        
        if target == "—" or target is None or not avg_uph:
            continue
        
        try:
            target = float(target)
            avg_uph = float(avg_uph)
        except (ValueError, TypeError):
            continue
        
        if target <= 0:
            continue
        
        # Assume 40-hour work week
        hours_week = 40.0
        expected_units = target * hours_week
        actual_units = avg_uph * hours_week
        unit_diff = actual_units - expected_units
        
        if unit_diff >= 0:  # Only negative impacts count
            continue
        
        cost_per_unit = _avg_wage / target if target > 0 else 0
        weekly_cost = abs(unit_diff * cost_per_unit)
        
        cost_by_emp.append({
            "name": emp_name,
            "department": emp.get("Department", ""),
            "weekly_cost": weekly_cost,
            "gap": target - avg_uph,
        })
    
    if cost_by_emp:
        cost_by_emp.sort(key=lambda x: x["weekly_cost"], reverse=True)
        top_3_cost = cost_by_emp[:3]
        total_top_3_cost = sum(e["weekly_cost"] for e in top_3_cost)
        
        st.error(f"🔴 **Top 3 underperformers costing ${total_top_3_cost:,.0f} this week**")
        
        for i, emp in enumerate(top_3_cost, 1):
            c1, c2, c3 = st.columns([2, 1, 1.5])
            c1.write(f"**{i}. {emp['name']}** ({emp['department']})")
            c2.metric("Gap", f"-{emp['gap']:.1f} UPH")
            c3.metric("Weekly Cost", f"${emp['weekly_cost']:,.0f}")
        
        st.caption(f"📣 **Key insight:** Improving these 3 people alone would save ${total_top_3_cost:,.0f}/week (${ total_top_3_cost * 52:,.0f}/year).")
    else:
        st.success("✅ No significant labor cost impact detected.")
    
    st.divider()
    
    # ────────────────────────────────────────────────────────────────────────────
    # SECTION 5: QUICK COACHING TIPS (Generic)
    # ────────────────────────────────────────────────────────────────────────────
    
    st.subheader("💡 Quick Coaching Tips")
    with st.expander("How to approach coaching conversations", expanded=False):
        st.markdown("""
**For High Risk (🔴):**
- Schedule immediate 1-on-1
- Ask: "What obstacles are you facing?"
- Identify specific, measurable targets
- Plan weekly check-in reviews

**For Medium Risk (🟡):**
- Proactive check-in within 3-5 days
- Discuss trends: "I've noticed a dip the last few days"
- Offer support: extra training, peer mentoring, resources
- Set clear improvement goals

**For Trending Down (but still on goal):**
- Friendly conversation to prevent further decline
- Ask about workload, personal factors, team dynamics
- Recognize effort while addressing trend

**Universal best practices:**
- Focus on behaviors & solutions, not blame
- Use data to support conversation
- Celebrate wins when they happen
- Make support concrete: training, tools, time
""")
    
    st.divider()
    
    # Footer with quick links
    st.caption("📚 Want to dive deeper? Visit **📈 Productivity** page for detailed analytics and risk breakdown.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: IMPORT DATA
# Clean 3-step flow. Each step only runs when active — no tab lag.
# Step 1: Upload CSV(s)
# Step 2: Map columns (per file, auto-detected)
# Step 3: Run pipeline — employees registered, UPH calculated, data ready
# ══════════════════════════════════════════════════════════════════════════════

def page_dashboard():
    """Dashboard: At-a-glance risk view of all employees."""
    st.title("📊 Dashboard")
    st.caption("Risk-based priority view of all employees. Filter by risk level or department.")

    if not require_db(): return

    if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
        with st.spinner("Loading dashboard data…"):
            try:
                from app import _build_archived_productivity
                _build_archived_productivity()
            except Exception:
                pass

    gs = st.session_state.get("goal_status", [])
    history = st.session_state.get("history", [])
    
    if not gs:
        st.info("No productivity data. Run Import Data to get started.")
        return

    # Filter for below_goal employees
    below_goal = [r for r in gs if r.get("goal_status") == "below_goal"]
    
    if not below_goal:
        st.success("✅ All employees are meeting their goals!")
        st.divider()
        # Show summary of on-goal employees
        on_goal = [r for r in gs if r.get("goal_status") != "below_goal"]
        if on_goal:
            st.subheader("On Target")
            st.metric("Employees meeting goals", len(on_goal))
        return

    # Score all below-goal employees using shared risk calculator
    risk_list = []
    for emp in below_goal:
        emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
        risk_level, risk_score, risk_details = _calc_risk_level(emp, history)
        risk_list.append({
            "name": emp.get("Employee", emp.get("Employee Name", "Unknown")),
            "department": emp.get("Department", ""),
            "emp_id": emp_id,
            "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
            "target_uph": emp.get("Target UPH", "—"),
            "trend": emp.get("trend", "—"),
            "change_pct": emp.get("change_pct", 0.0),
            "risk_level": risk_level,
            "risk_score": risk_score,
            "risk_details": risk_details,
        })

    # Filter controls
    col1, col2, col3 = st.columns([2, 2, 1])
    
    risk_filter = col1.multiselect(
        "Filter by risk level",
        ["🔴 High", "🟡 Medium", "🟢 Low"],
        default=["🔴 High", "🟡 Medium", "🟢 Low"],
        key="dash_risk_filter"
    )
    
    dept_options = sorted(set(r["department"] for r in risk_list if r["department"]))
    dept_filter = col2.multiselect(
        "Filter by department",
        ["All departments"] + dept_options,
        default=["All departments"],
        key="dash_dept_filter"
    )
    
    sort_by = col3.selectbox(
        "Sort by",
        ["Risk (High → Low)", "UPH (Low → High)", "Streak (Longest)"],
        key="dash_sort"
    )

    # Apply filters
    filtered = [r for r in risk_list if r["risk_level"] in risk_filter]
    if "All departments" not in dept_filter:
        filtered = [r for r in filtered if r["department"] in dept_filter]

    # Sort
    if sort_by == "Risk (High → Low)":
        filtered.sort(key=lambda x: x["risk_score"], reverse=True)
    elif sort_by == "UPH (Low → High)":
        filtered.sort(key=lambda x: x["avg_uph"])
    elif sort_by == "Streak (Longest)":
        filtered.sort(key=lambda x: x["risk_details"]["under_goal_streak"], reverse=True)

    # Summary metrics
    st.divider()
    high_risk = len([r for r in filtered if r["risk_level"] == "🔴 High"])
    med_risk = len([r for r in filtered if r["risk_level"] == "🟡 Medium"])
    low_risk = len([r for r in filtered if r["risk_level"] == "🟢 Low"])
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Below Goal", len(filtered))
    m2.metric("🔴 High Risk", high_risk)
    m3.metric("🟡 Medium Risk", med_risk)
    m4.metric("🟢 Low Risk", low_risk)

    st.divider()

    # Display as scrollable table
    st.subheader("Performance Priority List")
    
    if not filtered:
        st.info("No employees match the selected filters.")
        return

    # Create table data
    table_data = []
    for r in filtered:
        table_data.append({
            "Risk": r["risk_level"],
            "Name": r["name"],
            "Dept": r["department"],
            "Current UPH": r["avg_uph"],
            "Target": r["target_uph"],
            "Trend": r["trend"],
            "Streak": f"{r['risk_details']['under_goal_streak']} days",
            "Score": r["risk_score"],
        })

    df = pd.DataFrame(table_data)
    
    # Color code by risk
    def _color_risk(val):
        if "🔴" in str(val):
            return "background-color: #ffcccc; color: #8b0000"
        elif "🟡" in str(val):
            return "background-color: #fff9e6; color: #ff6600"
        elif "🟢" in str(val):
            return "background-color: #e6ffe6; color: #008000"
        return ""

    styled = df.style.applymap(_color_risk, subset=["Risk"])
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # Export button
    csv_buf = df.to_csv(index=False)
    st.download_button(
        "⬇️ Download priority list",
        csv_buf,
        f"priority_list_{date.today()}.csv",
        "text/csv",
        key="dash_export"
    )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: IMPORT DATA
# ══════════════════════════════════════════════════════════════════════════════

def page_import():
    st.title("📁 Import Data")

    if not require_db(): return

    step = st.session_state.import_step

    # ── Reset button — shown whenever something is in progress ───────────────
    if (step > 1 or st.session_state.get("uploaded_sessions") or
            st.session_state.get("alloc_rows") or st.session_state.pipeline_done):
        if st.button("↺ Start over", key="import_reset", type="secondary"):
            keys_to_clear = [
                "uploaded_sessions", "import_step", "alloc_rows",
                "pipeline_done", "top_performers", "goal_status", "dept_report",
                "dept_trends", "weekly_summary", "history", "mapping",
                "_archived_loaded",
            ]
            for k in keys_to_clear:
                st.session_state.pop(k, None)
            # Clear all au_ / ao_ / snap_ widget keys
            for k in list(st.session_state.keys()):
                if k.startswith(("au_", "ao_", "snap_", "alloc_sel_")):
                    del st.session_state[k]
            st.session_state.import_step = 1
            _bust_cache()
            st.rerun()

    # ── Step indicator ────────────────────────────────────────────────────────
    s1c = "#0F2D52" if step >= 1 else "#C5D4E8"
    s2c = "#2E7D32" if step >= 3 else ("#0F2D52" if step >= 2 else "#C5D4E8")
    st.markdown(f"""
<div style="display:flex;gap:8px;align-items:center;margin-bottom:1.5rem;">
  <div style="background:{s1c};color:#fff;border-radius:50%;width:28px;height:28px;
              display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;">1</div>
  <span style="color:{s1c};font-size:13px;font-weight:500;">Upload</span>
  <span style="color:#C5D4E8;margin:0 4px;">──────</span>
  <div style="background:{s2c};color:#fff;border-radius:50%;width:28px;height:28px;
              display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;">2</div>
  <span style="color:{s2c};font-size:13px;font-weight:500;">Process</span>
</div>""", unsafe_allow_html=True)

    if step == 1:
        _import_step1()
    elif step == 2:
        _import_step2()
    elif step == 3:
        _import_step3()


def _import_step1():
    """Step 1 — upload one or more CSV files."""
    st.subheader("Upload your CSV file(s)")
    st.caption("Upload one or more CSV files. Columns are auto-detected — you'll only see the mapping step if we can't find required fields.")

    files = st.file_uploader(
        "Drag files here or click Browse",
        type=["csv"],
        accept_multiple_files=True,
        key="import_uploader",
    )

    if not files:
        st.info("Waiting for files…")
        return

    _MAX_FILE_MB = 50
    pending = []
    for f in files:
        # ── File size guard ───────────────────────────────────────────
        f.seek(0, 2)
        size_mb = f.tell() / (1024 * 1024)
        f.seek(0)
        if size_mb > _MAX_FILE_MB:
            st.error(f"**{f.name}** is {size_mb:.1f} MB — max allowed is {_MAX_FILE_MB} MB. Split into smaller files.")
            continue

        try:
            raw_bytes = f.read()
        except Exception as _read_err:
            st.error(f"Could not read **{f.name}**: {_read_err}")
            _log_app_error("import", f"File read error ({f.name}): {_read_err}")
            continue

        rows, headers = _parse_csv(raw_bytes)
        if not headers:
            st.error(
                f"**{f.name}** could not be parsed as a valid CSV header row. "
                "Make sure row 1 contains column names (e.g., EmployeeID, EmployeeName, Department, UPH/Units/HoursWorked) "
                "and that the file is comma-separated."
            )
            continue
        if not rows:
            st.error(
                f"**{f.name}** has headers but no usable data rows. "
                "Add at least one employee row beneath the header and remove blank trailing lines."
            )
            continue
        pending.append({
            "filename":  f.name,
            "rows":      rows,
            "headers":   headers,
            "row_count": len(rows),
        })
        st.success(f"✓ **{f.name}** — {len(rows):,} rows, {len(headers)} columns")

    if pending:
        if st.button("Continue →", type="primary", use_container_width=True):
            sessions = [
                {**p, "mapping": {}, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M")}
                for p in pending
            ]
            # Auto-detect columns for each file
            all_auto = True
            for s in sessions:
                headers = s.get("headers", list(s.get("rows",[{}])[0].keys()) if s.get("rows") else [])
                auto = _auto_detect(headers)
                has_id   = bool(auto.get("EmployeeID"))
                has_name = bool(auto.get("EmployeeName"))
                has_uph  = bool(auto.get("UPH")) or (bool(auto.get("Units")) and bool(auto.get("HoursWorked")))
                if has_id and has_name and has_uph:
                    s["mapping"] = {
                        "Date":        auto.get("Date", ""),
                        "EmployeeID":  auto.get("EmployeeID", ""),
                        "EmployeeName":auto.get("EmployeeName", ""),
                        "Department":  auto.get("Department", ""),
                        "Shift":       auto.get("Shift", ""),
                        "UPH":         auto.get("UPH", ""),
                        "Units":       auto.get("Units", ""),
                        "HoursWorked": auto.get("HoursWorked", ""),
                    }
                else:
                    all_auto = False

            st.session_state.uploaded_sessions = sessions
            st.session_state.submission_plan  = None
            st.session_state.split_overrides  = {}
            if all_auto:
                # Skip mapping — go straight to pipeline
                st.session_state.import_step = 3
            else:
                st.session_state.import_step = 2
            st.rerun()


def _import_step2():
    """Step 2 — map columns for each file."""
    sessions = st.session_state.uploaded_sessions
    if not sessions:
        st.session_state.import_step = 1
        st.rerun()
        return

    st.subheader("Map your columns")
    st.caption("We've auto-detected the best match for each field. Check them and adjust anything that looks wrong.")

    all_mapped = True

    for idx, s in enumerate(sessions):
        headers = s.get("headers", list(s.get("rows",[{}])[0].keys()) if s.get("rows") else [])
        auto    = _auto_detect(headers)
        options = ["— not in this file —"] + headers

        with st.container():
            _safe_fn = _html_mod.escape(s["filename"])
            st.markdown(
                f'<div style="background:#F0F5FB;border-radius:8px;padding:14px 16px;margin-bottom:8px;">'
                f'<span style="font-size:14px;font-weight:600;color:#0F2D52;">{_safe_fn}</span>'
                f'<span style="font-size:12px;color:#5A7A9C;margin-left:12px;">{s["row_count"]:,} rows</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # ── UPH Source Selection (OUTSIDE form for immediate rerun) ──
            m = s.get("mapping") or {}
            _map_has_uph = bool(m.get("UPH"))
            _map_has_calc = bool(m.get("Units") and m.get("HoursWorked"))
            uph_src_key = f"uph_src_{idx}"
            _default_src = "Already have UPH column" if (_map_has_uph and not _map_has_calc) else "Calculate: Units ÷ Hours"
            saved_src   = st.session_state.get(uph_src_key, _default_src)
            
            # Clear stale values when source changes
            _prev_key = f"uph_src_prev_{idx}"
            _prev_src = st.session_state.get(_prev_key, saved_src)
            if _prev_src != saved_src:
                if "Already" in saved_src:
                    st.session_state.pop(f"fm_{idx}_Units_un", None)
                    st.session_state.pop(f"fm_{idx}_HoursWorked_h", None)
                else:
                    st.session_state.pop(f"fm_{idx}_UPH_u", None)
            st.session_state[_prev_key] = saved_src
            
            uph_src = st.radio(
                "UPH source",
                ["Calculate: Units ÷ Hours", "Already have UPH column"],
                index=1 if "Already" in saved_src else 0,
                key=uph_src_key,
                horizontal=True,
            )

            ca, cb = st.columns(2)

            def _sel(label, field, req, col, extra=""):
                cur = m.get(field) or auto.get(field, "")
                idx2 = options.index(cur) if cur in options else 0
                dot  = "🔴" if req else "⚪"
                v    = col.selectbox(f"{dot} {label}", options, index=idx2,
                                     key=f"fm_{idx}_{field}_{extra}")
                return "" if v.startswith("—") else v

            with ca:
                d_date = _sel("Date (optional — use work date picker if absent)",
                              "Date",         False, ca)
                d_eid  = _sel("Employee ID",  "EmployeeID",   True,  ca)
                d_name = _sel("Employee Name","EmployeeName", True,  ca)
                d_dept = _sel("Department",   "Department",   False, ca)

            with cb:
                d_shift = _sel("Shift",                    "Shift",       False, cb)

                if "Already" in uph_src:
                    st.caption("Using your existing UPH column.")
                    d_uph   = _sel("UPH column",   "UPH",         True,  cb, "u")
                    d_units = ""
                    d_hrs   = ""
                else:
                    st.caption("UPH will be calculated from Units ÷ Hours Worked.")
                    d_uph   = ""
                    d_units = _sel("Units",        "Units",       True,  cb, "un")
                    d_hrs   = _sel("Hours Worked", "HoursWorked", True,  cb, "h")

                # Validation — Date is optional (falls back to work date picker)
                missing = [
                    f for f, v in [
                        ("Employee ID", d_eid),
                        ("Employee Name", d_name),
                    ] if not v
                ]

                if "Already" in uph_src:
                    uph_ok = bool(d_uph)
                    if not uph_ok:
                        uph_msg = "Select your UPH column above."
                    else:
                        uph_msg = ""
                else:
                    uph_ok = bool(d_units and d_hrs)
                    if not d_units and not d_hrs:
                        uph_msg = "Select both Units and Hours Worked columns."
                    elif not d_units:
                        uph_msg = "Select the Units column."
                    elif not d_hrs:
                        uph_msg = "Select the Hours Worked column."
                    else:
                        uph_msg = ""

                can_confirm = not missing and uph_ok

                if missing:
                    st.warning(f"Still needed: {', '.join(missing)}")
                if uph_msg:
                    st.warning(uph_msg)
                # Auto-confirm if all fields detected with no conflicts
                _auto_confirmed = (can_confirm and
                                   not s.get("mapping") and
                                   all(auto.get(f) for f in ["EmployeeID","EmployeeName","Units"]) and
                                   len([v for v in auto.values() if v]) >= 4)
                if _auto_confirmed:
                    sessions[idx]["mapping"] = {
                        "Date": d_date, "EmployeeID": d_eid, "EmployeeName": d_name,
                        "Department": d_dept, "Shift": d_shift,
                        "UPH": d_uph, "Units": d_units, "HoursWorked": d_hrs,
                    }
                    st.session_state.uploaded_sessions = sessions
                    st.success(f"✓ Auto-confirmed: {s['filename']}")
                    st.rerun()

                if can_confirm and not _auto_confirmed:
                    st.info("✓ All fields mapped — ready to confirm.")

                confirmed = st.button(
                    f"Confirm mapping for {s['filename']}",
                    type="primary",
                    use_container_width=True,
                    disabled=_auto_confirmed,
                    key=f"confirm_mapping_{idx}",
                )

                if confirmed:
                    if not can_confirm:
                        st.warning("Fix the issues above before confirming.")
                    else:
                        sessions[idx]["mapping"] = {
                            "Date":        d_date,
                            "EmployeeID":  d_eid,
                            "EmployeeName":d_name,
                            "Department":  d_dept,
                            "Shift":       d_shift,
                            "UPH":         d_uph,
                            "Units":       d_units,
                            "HoursWorked": d_hrs,
                        }
                        st.session_state.uploaded_sessions = sessions
                        st.rerun()

            # Show confirmation tick if mapping saved
            if s.get("mapping") and s["mapping"].get("EmployeeID"):
                st.success(f"✓ Mapping confirmed for {s['filename']}")
            else:
                all_mapped = False

    st.divider()
    col1, col2 = st.columns(2)

    if col1.button("← Back to upload", use_container_width=True):
        st.session_state.import_step = 1
        st.rerun()

    if all_mapped:
        if col2.button("Continue to pipeline →", type="primary", use_container_width=True):
            st.session_state.import_step = 3
            st.rerun()
    else:
        col2.info("Confirm all file mappings above to continue.")


def _import_step3():
    """Step 3 — run the pipeline. Registers employees, calculates UPH, stores history."""
    sessions = st.session_state.uploaded_sessions
    if not sessions:
        st.session_state.import_step = 1
        st.rerun()
        return

    st.subheader("Run the pipeline")

    # Summary of what's loaded
    total_rows = sum(s["row_count"] for s in sessions)
    st.markdown(
        f'<div style="background:#E8F0F9;border-radius:8px;padding:14px 16px;margin-bottom:1rem;">'
        f'<span style="color:#0F2D52;font-weight:600;">{len(sessions)} file(s) ready</span>'
        f'<span style="color:#5A7A9C;font-size:12px;margin-left:12px;">{total_rows:,} total rows</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    for s in sessions:
        m = s.get("mapping",{})
        st.caption(f"📄 **{s['filename']}** — {s['row_count']:,} rows · mapped to {sum(1 for v in m.values() if v)} fields")

    st.divider()

    # Check if any session has a Date column mapped
    _has_date_col = any(s.get("mapping",{}).get("Date") for s in sessions)

    if _has_date_col:
        st.info("📅 Date column detected — work dates come from your CSV. No date picker needed.")
        work_date = date.today()   # placeholder only; rows use their own date
    else:
        st.markdown("**Work date for this import**")
        st.caption("Your CSV has no Date column mapped — all rows will be recorded under this date.")
        work_date = st.date_input("Work date", value=date.today(), label_visibility="collapsed")

    if st.button("▶  Run pipeline now", type="primary", use_container_width=True):
        all_rows    = []
        all_mapping = {}
        for s in sessions:
            all_rows.extend(s["rows"])
            if not all_mapping and s.get("mapping"):
                all_mapping = s["mapping"]

        bar = st.progress(0, text="Registering employees…")

        # Register all employees — one batch upsert instead of N individual calls
        id_col    = all_mapping.get("EmployeeID","EmployeeID")
        name_col  = all_mapping.get("EmployeeName","EmployeeName")
        dept_col  = all_mapping.get("Department","Department")
        shift_col = all_mapping.get("Shift","Shift")
        seen_emps = {}
        name_fixed_count = 0
        uph_rejected_count = 0
        neg_value_fixed_count = 0
        max_reasonable_uph = 500.0
        for row in all_rows:
            eid = str(row.get(id_col,"")).strip()
            if eid and eid not in seen_emps:
                _safe_name, _name_flagged = _sanitize_employee_name(row.get(name_col, ""), eid)
                _safe_dept = _normalize_label_text(row.get(dept_col, ""), max_len=40)
                _safe_shift = _normalize_label_text(row.get(shift_col, ""), max_len=30)
                if _name_flagged or _safe_name != str(row.get(name_col, "")).strip():
                    name_fixed_count += 1
                seen_emps[eid] = {
                    "emp_id":     eid,
                    "name":       _safe_name,
                    "department": _safe_dept,
                    "shift":      _safe_shift,
                }
        if seen_emps:
            # Check employee limit before importing
            try:
                from database import can_add_employees, get_employee_count, get_employee_limit
                _el = get_employee_limit()
                if _el != -1:  # not unlimited
                    _existing = get_employee_count()
                    _existing_ids = {
                        str(e.get("emp_id", "")).strip()
                        for e in (_cached_employees() or [])
                        if str(e.get("emp_id", "")).strip()
                    }
                    _new_ids = [eid for eid in seen_emps.keys() if eid not in _existing_ids]
                    _new_unique = len(_new_ids)
                    if _existing + _new_unique > _el and _el > 0:
                        _plan = _get_current_plan()
                        st.error(
                            f"Employee limit reached. Your **{_plan.capitalize()}** plan allows "
                            f"**{_el}** employees and you have **{_existing}**. "
                            f"This import adds **{_new_unique}** new employee(s). "
                            f"Upgrade your plan in Settings → Subscription."
                        )
                        return
            except Exception:
                pass  # don't block import if limit check fails
            try:
                batch_upsert_employees(list(seen_emps.values()))
            except Exception as _e:
                st.warning(f"Employee sync warning: {_e} — allocation will continue.")
                _log_app_error("pipeline", f"Employee sync error: {_e}", detail=traceback.format_exc(), severity="warning")
            _bust_cache()

        bar.progress(25, text="Processing rows…")

        # Run productivity pipeline
        try:
            from data_processor import process_data
            from ranker         import rank_employees, build_department_report, calculate_employee_risk
            from trends         import calculate_department_trends, build_weekly_summary, calculate_employee_rolling_average
            from error_log      import ErrorLog
            from goals          import analyse_trends, build_goal_status

            class _PS:
                def get(self, k, d=None): return st.session_state.get(k, d)
                def get_output_dir(self): return tempfile.gettempdir()
                def get_dept_target_uph(self, d):
                    t = _cached_targets().get(d, 0)
                    return float(t) if t else float(st.session_state.get("target_uph",0) or 0)
                def all_mappings(self): return all_mapping

            ps  = _PS()
            log = ErrorLog(tempfile.gettempdir())

            processed = process_data(all_rows, all_mapping, ps, log)

            # User-friendly cleanup pass: normalize employee labels and discard
            # unrealistic/non-finite UPH values before ranking.
            _proc_name_col = all_mapping.get("EmployeeName") or "EmployeeName"
            _proc_id_col = all_mapping.get("EmployeeID") or "EmployeeID"
            _proc_dept_col = all_mapping.get("Department") or "Department"
            _proc_uph_col = all_mapping.get("UPH") or "UPH"
            for _row in processed:
                _eid = str(_row.get(_proc_id_col, "")).strip()
                _raw_name = _row.get(_proc_name_col, "")
                _safe_name, _flagged = _sanitize_employee_name(_raw_name, _eid)
                if _flagged or _safe_name != str(_raw_name).strip():
                    name_fixed_count += 1
                _row[_proc_name_col] = _safe_name
                _row[_proc_dept_col] = _normalize_label_text(_row.get(_proc_dept_col, ""), max_len=40)

                _raw_uph = _row.get(_proc_uph_col, "")
                if str(_raw_uph).strip() != "":
                    try:
                        _uph = float(_raw_uph)
                        if (not math.isfinite(_uph)) or _uph < 0 or _uph > max_reasonable_uph:
                            _row[_proc_uph_col] = ""
                            uph_rejected_count += 1
                        else:
                            _row[_proc_uph_col] = round(_uph, 4)
                    except (ValueError, TypeError):
                        _row[_proc_uph_col] = ""
                        uph_rejected_count += 1

            if name_fixed_count or uph_rejected_count:
                st.warning(
                    "Data cleanup applied for readability: "
                    f"{name_fixed_count} employee label(s) normalized, "
                    f"{uph_rejected_count} invalid UPH value(s) ignored."
                )

            bar.progress(40, text="Preparing import data…")

            existing = st.session_state.history
            existing.extend(processed)
            st.session_state.history = existing

            bar.progress(60, text="Storing UPH history…")

            # Aggregate per employee — handle multiple files with different column names
            # Aggregate per (emp_id, order_number, date) so that:
            # - employees who worked on multiple orders get pre-split rows
            # - employees who worked across multiple dates get per-date rows
            # - if no order/date columns are mapped we fall back to one row per emp
            from collections import defaultdict

            # key = (emp_id, order_number_or_"", date_str)
            combo_agg = defaultdict(lambda: {
                "units": 0.0, "hours": 0.0, "uphs": [], "dept": "", "name": ""
            })

            # Track per (employee, date) for UPH history — one record per day per employee
            emp_date_totals = defaultdict(lambda: {
                "units": 0.0, "hours": 0.0, "uphs": [], "dept": "", "name": ""
            })

            # Determine whether any session has a date column mapped
            has_date_col  = any(s.get("mapping",{}).get("Date")  for s in sessions)

            for sess in sessions:
                s_mapping  = sess.get("mapping") or all_mapping
                s_rows     = sess.get("rows", [])
                s_id_col   = s_mapping.get("EmployeeID",   "EmployeeID")
                s_name_col = s_mapping.get("EmployeeName", "EmployeeName")
                s_dept_col = s_mapping.get("Department",   "Department")
                s_date_col = s_mapping.get("Date",         "")
                s_u_col    = s_mapping.get("Units",        "Units")
                s_h_col    = s_mapping.get("HoursWorked",  "HoursWorked")
                s_uph_col  = s_mapping.get("UPH",          "UPH")

                for row in s_rows:
                    eid = str(row.get(s_id_col, "")).strip()
                    if not eid:
                        continue

                    try:
                        units_val = float(row.get(s_u_col, 0) or 0)
                        if not math.isfinite(units_val):
                            units_val = 0.0
                    except (ValueError, TypeError):
                        units_val = 0.0
                    try:
                        hours_val = float(row.get(s_h_col, 0) or 0)
                        if not math.isfinite(hours_val):
                            hours_val = 0.0
                    except (ValueError, TypeError):
                        hours_val = 0.0
                    if units_val < 0:
                        units_val = 0.0
                        neg_value_fixed_count += 1
                    if hours_val < 0:
                        hours_val = 0.0
                        neg_value_fixed_count += 1
                    raw_uph = row.get(s_uph_col, None)

                    # Date: use mapped column value, fall back to work_date picker
                    row_date = work_date.isoformat()
                    if s_date_col and row.get(s_date_col):
                        raw_d = str(row[s_date_col]).strip()[:10]
                        try:
                            datetime.strptime(raw_d, "%Y-%m-%d")
                            row_date = raw_d
                        except ValueError:
                            pass

                    key = (eid, "", row_date)
                    combo_agg[key]["units"] += units_val
                    combo_agg[key]["hours"] += hours_val
                    _valid_uph_val = None
                    if raw_uph:
                        try:
                            _uph_val = float(raw_uph)
                            if math.isfinite(_uph_val) and 0 <= _uph_val <= max_reasonable_uph:
                                _valid_uph_val = _uph_val
                            else:
                                uph_rejected_count += 1
                        except (ValueError, TypeError):
                            uph_rejected_count += 1
                    if _valid_uph_val is not None:
                        combo_agg[key]["uphs"].append(_valid_uph_val)

                    name_val, _name_flagged = _sanitize_employee_name(row.get(s_name_col, ""), eid)
                    if _name_flagged:
                        name_fixed_count += 1
                    dept_val = _normalize_label_text(row.get(s_dept_col, ""), max_len=40)
                    if name_val:
                        combo_agg[key]["name"]                    = name_val
                        emp_date_totals[(eid, row_date)]["name"]  = name_val
                    if dept_val:
                        combo_agg[key]["dept"]                    = dept_val
                        emp_date_totals[(eid, row_date)]["dept"]  = dept_val

                    emp_date_totals[(eid, row_date)]["units"] += units_val
                    emp_date_totals[(eid, row_date)]["hours"] += hours_val
                    if _valid_uph_val is not None:
                        emp_date_totals[(eid, row_date)]["uphs"].append(_valid_uph_val)

            if neg_value_fixed_count:
                st.warning(f"Adjusted {neg_value_fixed_count} negative unit/hour value(s) to 0.")

            # Build alloc_rows — one entry per (emp, date) combo
            alloc_rows = []
            for (eid, _unused, row_date), agg in combo_agg.items():
                if agg["uphs"]:
                    uph = round(sum(agg["uphs"]) / len(agg["uphs"]), 2)
                elif agg["hours"] > 0:
                    uph = round(agg["units"] / agg["hours"], 2)
                else:
                    uph = 0.0
                alloc_rows.append({
                    "emp_id":    eid,
                    "name":      agg["name"] or eid,
                    "dept":      agg["dept"],
                    "units":     float(round(agg["units"])),
                    "hours":     round(agg["hours"], 2),
                    "uph":       uph,
                    "date":      row_date,
                })
            alloc_rows.sort(key=lambda r: (r.get("dept",""), r.get("name",""), r.get("date","")))

            # Canonical date for the alloc page
            wd_str    = work_date.isoformat()
            csv_dates = sorted({r["date"] for r in alloc_rows if r["date"]})
            canon_date = csv_dates[0] if (has_date_col and csv_dates) else wd_str

            # Resolve any blank departments from the employees table
            _emp_dept_map = {e["emp_id"]: e.get("department","") for e in (_cached_employees() or [])}

            # Store UPH history in background thread so user doesn't wait
            uph_batch = []
            for (eid, uph_date), agg in emp_date_totals.items():
                if agg["uphs"]:
                    uph = round(sum(agg["uphs"]) / len(agg["uphs"]), 2)
                elif agg["hours"] > 0:
                    uph = round(min(agg["units"] / agg["hours"], 9999), 2)
                else:
                    uph = 0.0
                if not math.isfinite(uph):
                    uph = 0.0
                units_total = round(agg["units"])
                hours_total = round(agg["hours"], 2)
                if not math.isfinite(float(units_total)):
                    units_total = 0
                if not math.isfinite(float(hours_total)):
                    hours_total = 0.0
                dept = agg["dept"] or _emp_dept_map.get(eid, "")
                uph_batch.append({
                    "emp_id":       eid,
                    "work_date":    uph_date,
                    "uph":          uph,
                    "units":        units_total,
                    "hours_worked": hours_total,
                    "department":   dept,
                })
            # Avoid inserting exact duplicates when users import the same file again.
            _dup_skipped = 0
            try:
                _tenant_id = st.session_state.get("tenant_id", "")
                _dates = sorted({r.get("work_date") for r in uph_batch if r.get("work_date")})
                _date_min = _dates[0] if _dates else ""
                _date_max = _dates[-1] if _dates else ""
                _emp_ids = sorted({r.get("emp_id") for r in uph_batch if r.get("emp_id")})
                _existing_keys = set()
                if _tenant_id and _date_min and _date_max and _emp_ids:
                    _sb = _get_db_client()
                    _res = (
                        _sb.table("uph_history")
                        .select("emp_id, work_date, uph, units, hours_worked")
                        .eq("tenant_id", _tenant_id)
                        .gte("work_date", _date_min)
                        .lte("work_date", _date_max)
                        .in_("emp_id", _emp_ids)
                        .execute()
                    )
                    for _er in (_res.data or []):
                        _existing_keys.add((
                            str(_er.get("emp_id", "")),
                            str(_er.get("work_date", "")),
                            round(float(_er.get("uph") or 0), 4),
                            round(float(_er.get("units") or 0), 4),
                            round(float(_er.get("hours_worked") or 0), 4),
                        ))

                _filtered_batch = []
                for _r in uph_batch:
                    _key = (
                        str(_r.get("emp_id", "")),
                        str(_r.get("work_date", "")),
                        round(float(_r.get("uph") or 0), 4),
                        round(float(_r.get("units") or 0), 4),
                        round(float(_r.get("hours_worked") or 0), 4),
                    )
                    if _key in _existing_keys:
                        _dup_skipped += 1
                        continue
                    _filtered_batch.append(_r)
                uph_batch = _filtered_batch
            except Exception:
                pass

            # Store UPH history synchronously so data is in DB before pipeline completes
            try:
                _bg_tid = st.session_state.get("tenant_id", "")
                if _bg_tid:
                    uph_batch = [{**r, "tenant_id": _bg_tid} for r in uph_batch]
                batch_store_uph_history(uph_batch)
                if _dup_skipped:
                    st.info(f"Skipped {_dup_skipped} duplicate UPH row(s) already in history.")
            except Exception as _uph_err:
                st.warning(f"UPH history storage warning: {_uph_err}")
                _log_app_error("pipeline", f"UPH history storage failed: {_uph_err}",
                               detail=traceback.format_exc(), severity="warning")

            _bust_cache()

            # Rebuild productivity from full DB (includes all past imports)
            # so that a second import doesn't lose the first import's data.
            bar.progress(90, text="Rebuilding full productivity view…")
            _full_ok = _build_archived_productivity()
            _ranked_count = len(st.session_state.get("top_performers", []))
            if _full_ok:
                st.session_state["_archived_last_refresh_ts"] = time.time()
            if not _full_ok:
                # Fallback path: only compute heavy analytics when archived rebuild fails.
                ranked = rank_employees(existing, all_mapping, ps, log)
                targets = _cached_targets()
                trend_data = analyse_trends(existing, all_mapping, weeks=st.session_state.trend_weeks)
                goal_status = build_goal_status(ranked, targets, trend_data)
                dept_report = build_department_report(ranked, ps, log)
                dept_trends = calculate_department_trends(existing, all_mapping, ps, log)
                weekly = build_weekly_summary(existing, all_mapping, ps, log)
                rolling_avg = calculate_employee_rolling_average(existing, all_mapping, ps, log)
                risk_scores = calculate_employee_risk(existing, all_mapping, ps, log)
                _ranked_count = len(ranked)

                # Fallback: use only current import's data
                st.session_state.update({
                    "top_performers":    ranked,
                    "dept_report":       dept_report,
                    "dept_trends":       dept_trends,
                    "weekly_summary":    weekly,
                    "employee_rolling_avg": rolling_avg,
                    "employee_risk":     risk_scores,
                    "goal_status":       goal_status,
                    "trend_data":        trend_data,
                    "pipeline_done":     True,
                    "_archived_loaded":  False,
                })

            st.session_state.update({
                "mapping":           all_mapping,
                "alloc_rows":        alloc_rows,
                "alloc_date":        canon_date,
                "alloc_has_date":    has_date_col,
            })
            bar.progress(100, text="Done!")
            _unique_emp_count = len({r["emp_id"] for r in alloc_rows})
            st.toast(f"✓ {_ranked_count} employees ranked · {_unique_emp_count} employees processed", icon="✅")
            st.session_state.goto_page = "productivity"
            st.rerun()

        except Exception as _pipe_err:
            _tb = traceback.format_exc()
            st.error("Pipeline error:")
            st.code(_tb)
            _log_app_error("pipeline", str(_pipe_err), detail=_tb)

    st.divider()
    col1, col2 = st.columns(2)

    if col1.button("← Back to mapping", use_container_width=True):
        st.session_state.import_step = 2
        st.rerun()

    if st.session_state.pipeline_done and st.session_state.alloc_rows:
        _uc = len({r["emp_id"] for r in st.session_state.alloc_rows})
        col2.success(f"✓ {_uc} employees processed — view **📈 Productivity**")

    if st.button("↺ Start fresh import", use_container_width=True):
        st.session_state.uploaded_sessions = []
        st.session_state.import_step       = 1
        st.session_state.alloc_rows        = []
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EMPLOYEES
# ══════════════════════════════════════════════════════════════════════════════

def _build_coaching_recommendations():
    """Generate smart coaching recommendations based on employee performance data."""
    gs = st.session_state.get("goal_status", [])
    if not gs:
        return []

    recommendations = []
    for r in gs:
        name = r.get("Employee", "")
        dept = r.get("Department", "")
        uph = r.get("Avg UPH")
        target = r.get("Target UPH")
        trend_dir = r.get("Trend", "")
        vs_target = r.get("vs Target")

        if not name:
            continue

        try:
            uph = float(uph) if uph and uph not in ("—", None, "") else None
        except (ValueError, TypeError):
            uph = None
        try:
            target = float(target) if target and target not in ("—", None, "", 0) else None
        except (ValueError, TypeError):
            target = None

        if uph is None:
            continue

        rec = {"name": name, "dept": dept, "uph": uph, "target": target,
               "priority": "low", "actions": [], "status": ""}

        # Determine gap
        gap_pct = 0
        if target and target > 0:
            gap_pct = round(((uph - target) / target) * 100, 1)

        # Rule-based coaching logic
        if target and gap_pct < -20:
            rec["priority"] = "high"
            rec["status"] = f"{gap_pct}% below target"
            rec["actions"].append(f"Schedule one-on-one coaching session — {name} is significantly below the {dept} target of {target} UPH.")
            rec["actions"].append("Review workstation setup and process efficiency for immediate improvements.")
            rec["actions"].append("Consider pairing with a top performer for peer mentoring.")
            if trend_dir and "declining" in str(trend_dir).lower():
                rec["actions"].append("URGENT: Performance is declining. Investigate potential issues (equipment, training, engagement).")
        elif target and gap_pct < -10:
            rec["priority"] = "medium"
            rec["status"] = f"{gap_pct}% below target"
            rec["actions"].append(f"Monitor closely — {name} is moderately below the {target} UPH target.")
            rec["actions"].append("Provide targeted training on efficiency techniques.")
            if trend_dir and "improving" in str(trend_dir).lower():
                rec["actions"].append("Positive sign: trend is improving. Continue current support.")
            else:
                rec["actions"].append("Set a 2-week improvement checkpoint to track progress.")
        elif target and gap_pct < 0:
            rec["priority"] = "low"
            rec["status"] = f"{gap_pct}% below target"
            rec["actions"].append(f"Slightly below target. Encourage {name} and recognize effort.")
            rec["actions"].append("Small adjustments to workflow may close the gap.")
        elif target and gap_pct >= 20:
            rec["priority"] = "star"
            rec["status"] = f"+{gap_pct}% above target"
            rec["actions"].append(f"Top performer! Consider {name} for peer mentor or team lead role.")
            rec["actions"].append("Recognize achievement publicly to boost team morale.")
        elif target and gap_pct >= 0:
            rec["priority"] = "low"
            rec["status"] = f"+{gap_pct}% above target"
            rec["actions"].append("Meeting or exceeding target. Keep up the good work!")
        else:
            rec["status"] = "No target set"
            rec["actions"].append(f"Set a department target for {dept} to enable performance tracking.")

        # Trend-based additions
        if trend_dir and "declining" in str(trend_dir).lower() and rec["priority"] != "high":
            rec["priority"] = "medium" if rec["priority"] == "low" else rec["priority"]
            rec["actions"].append("Note: Performance trend is declining — check in with employee.")

        recommendations.append(rec)

    # Sort: high first, then medium, then low, then star
    priority_order = {"high": 0, "medium": 1, "low": 2, "star": 3}
    recommendations.sort(key=lambda x: priority_order.get(x["priority"], 99))
    return recommendations


def page_employees():
    st.title("👥 Employees")
    if not require_db(): return

    _plan = _get_current_plan()
    EMP_OPTS = ["Employee History", "Performance Journal"]
    if _plan in ("pro", "business", "admin"):
        EMP_OPTS.append("🤖 Coaching Insights")
    if "emp_view" not in st.session_state or st.session_state.emp_view not in EMP_OPTS:
        st.session_state.emp_view = "Employee History"

    chosen = st.session_state.emp_view
    cols   = st.columns(len(EMP_OPTS))
    for i, opt in enumerate(EMP_OPTS):
        if cols[i].button(opt, key=f"ev_{opt}", use_container_width=True,
                           type="primary" if chosen == opt else "secondary"):
            st.session_state.emp_view = opt
            st.rerun()

    st.divider()
    try:
        if   chosen == "Employee History": _emp_history()
        elif chosen == "Performance Journal":   _emp_coaching()
        elif chosen == "🤖 Coaching Insights":  _emp_ai_coaching()
    except Exception as e:
        st.error(f"Error: {e}")
        _log_app_error("employees", f"Employee page error: {e}", detail=traceback.format_exc())



@st.fragment
def _emp_history():
    st.subheader("Employee UPH history")
    emps = _cached_employees()
    if not emps:
        st.info("No employees yet. Go to **Import Data** to upload a CSV — employees are created automatically from your data.")
        return

    # ── Department filter → populates employee dropdown ───────────────────────
    depts    = sorted({e.get("department","") for e in emps if e.get("department")})
    dept_sel = st.selectbox("Filter by department", ["All departments"] + depts, key="eh_dept")

    if dept_sel == "All departments":
        filtered_emps = emps
    else:
        filtered_emps = [e for e in emps if e.get("department","") == dept_sel]

    if not filtered_emps:
        st.info("No employees in that department.")
        return

    # Employee dropdown: Name — Department — ID
    emp_opts = {
        f"{_normalize_label_text(e.get('name',''))} — {_normalize_label_text(e.get('department','') or 'No dept', max_len=28)} — {e['emp_id']}": e["emp_id"]
        for e in filtered_emps
    }
    chosen = st.selectbox("Select employee", list(emp_opts.keys()), key="eh_emp")
    emp_id = emp_opts[chosen]

    from datetime import timedelta as _tdelta
    dc1, dc2 = st.columns(2)
    _def_from = (date.today() - _tdelta(days=90)).strftime("%m/%d/%Y")
    _def_to   = date.today().strftime("%m/%d/%Y")
    from_str  = dc1.text_input("From", value=st.session_state.get("eh_from", _def_from),
                                key="eh_from_input", placeholder="MM/DD/YYYY")
    to_str    = dc2.text_input("To",   value=st.session_state.get("eh_to",   _def_to),
                                key="eh_to_input",   placeholder="MM/DD/YYYY")
    try:
        from_date_h = datetime.strptime(from_str.strip(), "%m/%d/%Y").date()
    except Exception:
        from_date_h = date.today() - _tdelta(days=90)
    try:
        to_date_h = datetime.strptime(to_str.strip(), "%m/%d/%Y").date()
    except Exception:
        to_date_h = date.today()
    st.session_state["eh_from"] = from_str
    st.session_state["eh_to"]   = to_str
    days = max(1, (to_date_h - from_date_h).days)
    from_iso = from_date_h.isoformat()
    to_iso   = to_date_h.isoformat()

    # Query uph_history within date range
    try:
        _sb_h = _get_db_client()
        _r_h  = _sb_h.table("uph_history").select("*") \
                      .eq("emp_id", emp_id) \
                      .gte("work_date", from_iso) \
                      .lte("work_date", to_iso) \
                      .order("work_date").execute()
        history = _r_h.data or []
    except Exception:
        history = []

    # Fallback: derive from unit_submissions if uph_history empty
    if not history:
        try:
            from collections import defaultdict as _dh
            _sb3 = _get_db_client()
            _r3  = _sb3.table("unit_submissions").select("*") \
                        .eq("emp_id", emp_id) \
                        .gte("work_date", from_iso) \
                        .lte("work_date", to_iso) \
                        .order("work_date").execute()
            _day = _dh(lambda: {"units": 0.0, "hours": 0.0})
            for _s in (_r3.data or []):
                _dk = _s.get("work_date", "")
                _day[_dk]["units"] += float(_s.get("units") or 0)
                _day[_dk]["hours"] += float(_s.get("hours_worked") or 0)
            history = [{"work_date": _dk,
                        "uph":   round(_v["units"] / _v["hours"], 2) if _v["hours"] > 0 else 0,
                        "units": round(_v["units"]),
                        "hours_worked": round(_v["hours"], 2)}
                       for _dk, _v in sorted(_day.items())]
        except Exception:
            pass

    if not history:
        st.info("No history yet for this employee.")
        return

    uphvals = [float(h.get("uph") or 0) for h in history if h.get("uph")]
    avg_uph = round(sum(uphvals) / len(uphvals), 2) if uphvals else None
    st.metric(f"Avg UPH ({from_str} – {to_str})", f"{avg_uph:.2f}" if avg_uph else "No data")

    df = pd.DataFrame([{
        "Date":  h.get("work_date", ""),
        "UPH":   round(float(h.get("uph") or 0), 2),
        "Units": int(h.get("units", 0) or 0),
        "Hours": round(float(h.get("hours_worked") or 0), 2),
    } for h in history]).sort_values("Date")

    # Remove rows with 0 or non-finite UPH before charting to avoid Infinity warnings
    import math as _math
    df_chart = df[df["UPH"].apply(lambda x: x > 0 and _math.isfinite(x))]
    if not df_chart.empty:
        st.line_chart(df_chart.set_index("Date")[["UPH"]], use_container_width=True)
    else:
        st.info("No valid UPH data to chart for this date range.")
    st.dataframe(df, use_container_width=True, hide_index=True)

    if st.button("⬇️ Export employee history"):
        with st.spinner("Generating…"):
            data = export_employee(emp_id)
        st.download_button(f"⬇️ Download {emp_id}_history.xlsx", data,
                           f"employee_{emp_id}_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@st.fragment
@st.fragment
def _emp_ai_coaching():
    """Show AI-powered coaching recommendations based on performance data."""
    st.subheader("🤖 Coaching Insights")
    st.caption("Smart recommendations based on employee performance, goals, and trends.")

    if not st.session_state.get("pipeline_done") and not st.session_state.get("_archived_loaded"):
        _build_archived_productivity()

    recs = _build_coaching_recommendations()
    if not recs:
        st.info("No coaching data available. Import productivity data and set department goals first.")
        return

    # Summary metrics
    _high = sum(1 for r in recs if r["priority"] == "high")
    _med  = sum(1 for r in recs if r["priority"] == "medium")
    _stars = sum(1 for r in recs if r["priority"] == "star")
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("🔴 Urgent", _high)
    mc2.metric("🟡 Monitor", _med)
    mc3.metric("🟢 On Track", sum(1 for r in recs if r["priority"] == "low"))
    mc4.metric("⭐ Stars", _stars)

    st.markdown("---")

    # Filter
    _filter = st.radio("Show", ["All", "🔴 Urgent only", "🟡 Needs attention", "⭐ Top performers"],
                        horizontal=True, key="coaching_filter")

    for rec in recs:
        if _filter == "🔴 Urgent only" and rec["priority"] != "high":
            continue
        if _filter == "🟡 Needs attention" and rec["priority"] not in ("high", "medium"):
            continue
        if _filter == "⭐ Top performers" and rec["priority"] != "star":
            continue

        # Priority badge
        _badge = {"high": "🔴", "medium": "🟡", "low": "🟢", "star": "⭐"}.get(rec["priority"], "")
        _uph_str = f" — {rec['uph']} UPH" if rec["uph"] else ""
        _target_str = f" (target: {rec['target']})" if rec["target"] else ""

        with st.expander(f"{_badge} **{rec['name']}** · {rec['dept']}{_uph_str}{_target_str} · {rec['status']}"):
            for action in rec["actions"]:
                st.markdown(f"→ {action}")

    st.markdown("---")
    st.caption("Recommendations are generated from UPH data, department targets, and performance trends. "
               "Review and adapt based on your direct knowledge of each employee.")


def _emp_coaching():
    emps  = _cached_employees()
    flags = _cached_active_flags()
    if not emps:
        st.info("No employees yet. Go to **Import Data** to upload a CSV — employees are created automatically from your data.")
        return

    # ── Build employee list — all employees, annotate who has notes / is flagged
    emp_ids_with_notes = _cached_all_coaching_notes()
    all_depts = sorted({e.get("department","") for e in emps if e.get("department")})

    # ── Manager Action List ──────────────────────────────────────────────────
    # Auto-generates follow-up actions for trending-down or below-goal employees
    gs = st.session_state.get("goal_status", [])
    if "dismissed_actions" not in st.session_state:
        st.session_state.dismissed_actions = set()

    action_items = []
    for r in gs:
        eid  = str(r.get("EmployeeID", r.get("Employee Name", "")))
        name = r.get("Employee Name", "")
        dept = r.get("Department", "")
        trend      = r.get("trend", "")
        goal_st    = r.get("goal_status", "")
        change_pct = r.get("change_pct", 0)
        avg_uph    = r.get("Average UPH", 0)
        target     = r.get("Target UPH", "—")

        reasons = []
        if trend == "down":
            reasons.append(f"Trending down ({change_pct:+.1f}%)")
        if goal_st == "below_goal":
            reasons.append(f"Below goal (UPH {avg_uph:.1f} vs target {target})")

        if reasons:
            action_key = f"{eid}|{'|'.join(reasons)}"
            if action_key not in st.session_state.dismissed_actions:
                action_items.append({
                    "eid": eid, "name": name, "dept": dept,
                    "reasons": reasons, "key": action_key,
                    "trend": trend, "goal_st": goal_st,
                })

    if action_items:
        with st.expander(f"📋 **Manager Action List** — {len(action_items)} follow-up(s) needed", expanded=True):
            st.caption("Employees trending down or below goal. Complete the follow-up, then dismiss the action.")
            for ai in action_items:
                ac1, ac2, ac3 = st.columns([3, 5, 1])
                badge = ""
                if ai["trend"] == "down": badge += " ↓"
                if ai["goal_st"] == "below_goal": badge += " ⚠️"
                ac1.markdown(f"**{ai['name']}**{badge}")
                ac2.caption(f"{ai['dept']} · {' · '.join(ai['reasons'])}")
                if ac3.button("✓", key=f"dismiss_{ai['key']}", help="Mark as done"):
                    st.session_state.dismissed_actions.add(ai["key"])
                    st.rerun()

            if st.button("Clear all completed actions", key="clear_dismissed", type="secondary"):
                st.session_state.dismissed_actions.clear()
                st.rerun()
    else:
        if gs:
            st.success("✓ No follow-ups needed — all employees are on track.")

    # ── Top bar: dept filter ──────────────────────────────────────────────────
    dept_sel = st.selectbox("Department", ["All departments"] + all_depts, key="cn_dept",
                             label_visibility="collapsed")
    filtered_emps = (emps if dept_sel == "All departments"
                     else [e for e in emps if e.get("department","") == dept_sel])

    st.divider()

    # ── Main two-column layout ────────────────────────────────────────────────
    col_list, col_detail = st.columns([2, 3], gap="large")

    with col_list:
        st.caption(f"**{dept_sel}** · {len(filtered_emps)} employee(s)")
        roster = []
        for e in filtered_emps:
            indicators = ""
            if e["emp_id"] in flags:           indicators += "🚩 "
            if e["emp_id"] in emp_ids_with_notes:
                _nc = len(_cached_coaching_notes_for(e["emp_id"]))
                indicators += f"📝{_nc}" if _nc else "📝"
            # Add trend indicator from goal_status
            _gs_match = next((r for r in gs if str(r.get("EmployeeID","")) == e["emp_id"]), None)
            if _gs_match:
                _t = _gs_match.get("trend","")
                if _t == "down": indicators += " ↓"
                elif _t == "up": indicators += " ↑"
            roster.append({
                " ": indicators.strip(),
                "Name": e["name"],
                "Dept": e.get("department",""),
            })
        df_roster = pd.DataFrame(roster)
        sel = st.dataframe(
            df_roster,
            use_container_width=True,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
        )
        sel_rows = sel.selection.rows if sel and sel.selection else []
        if sel_rows:
            selected_emp = filtered_emps[sel_rows[0]]
            st.session_state["cn_selected_emp"] = selected_emp["emp_id"]
        else:
            st.session_state.pop("cn_selected_emp", None)
            selected_emp = None

    with col_detail:
        if not selected_emp:
            st.info("← Select an employee from the list")
        else:
            emp_id   = selected_emp["emp_id"]
            emp_name = selected_emp["name"]
            emp_dept = selected_emp.get("department","")

            # ── Header ────────────────────────────────────────────────────────
            is_flagged = emp_id in flags
            hc1, hc2 = st.columns([4, 1])
            hc1.markdown(f"### {emp_name}")
            hc1.caption(emp_dept)

            # Show trend + goal summary inline
            _emp_gs = next((r for r in gs if str(r.get("EmployeeID","")) == str(emp_id)), None)
            if _emp_gs:
                _trend_icon = {"up": "↑", "down": "↓", "flat": "→"}.get(_emp_gs.get("trend",""), "—")
                _chg = _emp_gs.get("change_pct", 0)
                _chg_str = f"+{_chg:.1f}%" if _chg > 0 else f"{_chg:.1f}%"
                _uph = _emp_gs.get("Average UPH", 0)
                _tgt = _emp_gs.get("Target UPH", "—")
                _gs_color = {"on_goal": "green", "below_goal": "red"}.get(_emp_gs.get("goal_status",""), "gray")
                st.markdown(
                    f"<span style='font-size:13px;'>"
                    f"Avg UPH: <strong>{_uph:.1f}</strong> · "
                    f"Target: <strong>{_tgt}</strong> · "
                    f"Trend: {_trend_icon} {_chg_str}"
                    f"</span>",
                    unsafe_allow_html=True)

            if is_flagged:
                flag_info = flags.get(emp_id, {})
                st.warning(f"🚩 Flagged for performance tracking since **{flag_info.get('flagged_on','')}**")
                if st.button("✓ Remove flag", key="cn_unflag", type="secondary"):
                    from goals import unflag_employee
                    unflag_employee(emp_id)
                    _raw_cached_active_flags.clear()
                    for r in st.session_state.get("goal_status", []):
                        if str(r.get("EmployeeID","")) == str(emp_id):
                            r["flagged"] = False
                    st.rerun()

            # ── Add entry ────────────────────────────────────────────────────
            if "cn_note_val" not in st.session_state: st.session_state.cn_note_val = ""
            if "cn_by_val"   not in st.session_state: st.session_state.cn_by_val   = ""
            note_text  = st.text_area("Add a journal entry", height=90, key="cn_note",
                                       value=st.session_state.cn_note_val,
                                       placeholder="Performance observation, coaching session, follow-up…")
            nc1, nc2 = st.columns([2, 3])
            created_by = nc2.text_input("Your name (optional)", key="cn_by",
                                         value=st.session_state.cn_by_val,
                                         placeholder="Your name")
            if nc1.button("💾 Save entry", type="primary", use_container_width=True):
                if note_text.strip():
                    add_coaching_note(emp_id, note_text.strip(), created_by.strip())
                    _raw_cached_coaching_notes_for.clear()
                    _raw_cached_all_coaching_notes.clear()
                    st.session_state.cn_note_val = ""
                    st.session_state.cn_by_val   = ""
                    st.success("✓ Entry saved.")
                    st.rerun()
                else:
                    st.warning("Write something before saving the entry.")

            # ── Past entries ─────────────────────────────────────────────────
            notes = _cached_coaching_notes_for(emp_id)
            st.divider()
            if notes:
                st.caption(f"{len(notes)} journal entry/entries on file")
                for n in notes:
                    with st.container():
                        nc1, nc2 = st.columns([10, 1])
                        _n_date = _html_mod.escape(str(n.get('created_at',''))[:10])
                        _n_by = _html_mod.escape(n.get('created_by',''))
                        _n_text = _html_mod.escape(n.get('note',''))
                        nc1.markdown(
                            f"<div style='background:#F7F9FC;border-left:3px solid #0F2D52;"
                            f"border-radius:4px;padding:8px 12px;margin-bottom:4px;'>"
                            f"<span style='color:#5A7A9C;font-size:11px;'>"
                            f"{_n_date}"
                            f"{'  ·  ' + _n_by if _n_by else ''}</span>"
                            f"<br><span style='color:#1A2D42;font-size:13px;'>{_n_text}</span>"
                            f"</div>",
                            unsafe_allow_html=True)
                        note_id = n.get("id")
                        if note_id and nc2.button("🗑", key=f"del_{note_id}", help="Delete"):
                            delete_coaching_note(str(note_id))
                            _raw_cached_coaching_notes_for.clear()
                            _raw_cached_all_coaching_notes.clear()
                            st.rerun()

                # Actions row
                st.divider()
                ac1, ac2 = st.columns(2)
                if ac1.button("⬇️ Export journal", key="cn_export", use_container_width=True):
                    buf = io.BytesIO()
                    pd.DataFrame([{
                        "Date": str(n.get("created_at",""))[:10],
                        "Note": n.get("note",""),
                        "By":   n.get("created_by",""),
                    } for n in notes]).to_excel(buf, index=False, sheet_name="Journal")
                    buf.seek(0)
                    ac1.download_button(f"⬇️ Download", buf.read(),
                                        f"{emp_id}_journal_{date.today()}.xlsx",
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="cn_dl", use_container_width=True)
                if ac2.button("📦 Archive all entries", key="cn_archive",
                               use_container_width=True, type="secondary"):
                    archive_coaching_notes(emp_id)
                    _raw_cached_coaching_notes_for.clear()
                    _raw_cached_all_coaching_notes.clear()
                    st.session_state.pop("cn_selected_emp", None)
                    st.rerun()
            else:
                st.caption("No journal entries yet — add one above.")




# ══════════════════════════════════════════════════════════════════════════════
# PAGE: PRODUCTIVITY (UPH rankings, goals, tracker — existing functionality)
# ══════════════════════════════════════════════════════════════════════════════



def _build_archived_productivity():
    """
    Build productivity session state from DB. Queries aggregate data directly
    to avoid fetching thousands of raw rows.
    """
    from collections import defaultdict
    from datetime import datetime as _dt

    emps     = {e["emp_id"]: e for e in (_cached_employees() or [])}
    emp_dept = {eid: e.get("department","") for eid, e in emps.items()}
    emp_name = {eid: e.get("name", eid)    for eid, e in emps.items()}

    # Don't bail if employees table is empty — UPH history has dept/emp info
    try:
        sb = _get_db_client()
    except NameError:
        from database import get_client as _db_get_client
        sb = _db_get_client()

    # ── Per-employee aggregates (avg UPH, total units, record count) ──────────
    # Use running sum+count instead of growing lists to keep memory constant.
    emp_agg        = defaultdict(lambda: {"uph_sum": 0.0, "units": 0.0, "count": 0})
    month_dept_agg = defaultdict(lambda: defaultdict(lambda: {"uph_sum": 0.0, "uph_count": 0, "units": 0.0}))
    week_dept_agg  = defaultdict(lambda: defaultdict(lambda: {"units": 0.0, "uph_sum": 0.0, "uph_count": 0}))
    # Per-employee daily UPH for rolling average
    emp_daily      = defaultdict(list)  # eid -> [(date, uph)]
    # Per-employee weekly UPH for trend analysis
    emp_week_uph   = defaultdict(lambda: defaultdict(list))  # emp_id -> week_str -> [uph values]

    page_size = 1000
    offset    = 0
    total_rows = 0
    _last_err  = None
    while True:
        try:
            from database import _tq as _tq_fn
            r = _tq_fn(sb.table("uph_history").select(
                "emp_id, work_date, uph, units, department"
            )).order("work_date").range(offset, offset + page_size - 1).execute()
            batch = r.data or []
        except Exception as _pe:
            _last_err = repr(_pe)
            break
        for row in batch:
            eid   = row.get("emp_id","")
            uph   = float(row.get("uph") or 0)
            units = float(row.get("units") or 0)
            dept  = emp_dept.get(eid) or row.get("department") or "Unknown"
            # Backfill dept/name from UPH history when employees table is empty
            if eid not in emp_dept and dept:
                emp_dept[eid] = dept
            if eid not in emp_name:
                emp_name[eid] = eid  # fallback to ID
            wd    = (row.get("work_date") or "")[:10]
            month = wd[:7]
            if uph > 0:
                emp_agg[eid]["uph_sum"] += uph
                emp_agg[eid]["count"]   += 1
            emp_agg[eid]["units"] += units
            emp_daily[eid].append((wd, uph))
            if month and uph > 0:
                month_dept_agg[month][dept]["uph_sum"]   += uph
                month_dept_agg[month][dept]["uph_count"] += 1
                month_dept_agg[month][dept]["units"]     += units
            if wd:
                try:
                    d  = _dt.strptime(wd, "%Y-%m-%d")
                    wk = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
                    week_dept_agg[wk][dept]["units"] += units
                    if uph > 0:
                        week_dept_agg[wk][dept]["uph_sum"]   += uph
                        week_dept_agg[wk][dept]["uph_count"] += 1
                        emp_week_uph[eid][wk].append(uph)
                except Exception:
                    pass
        total_rows += len(batch)
        if len(batch) < page_size:
            break
        offset += page_size

    # Fallback to unit_submissions if uph_history empty
    if total_rows == 0:
        offset = 0
        while True:
            try:
                r = _tq_fn(sb.table("unit_submissions").select(
                    "emp_id, work_date, units, hours_worked"
                )).order("work_date").range(offset, offset + page_size - 1).execute()
                batch = r.data or []
            except Exception:
                break
            for row in batch:
                eid   = row.get("emp_id","")
                units = float(row.get("units") or 0)
                hours = float(row.get("hours_worked") or 0)
                uph   = round(units / hours, 2) if hours > 0 else 0
                dept  = emp_dept.get(eid, "Unknown")
                wd    = (row.get("work_date") or "")[:10]
                month = wd[:7]
                if uph > 0:
                    emp_agg[eid]["uph_sum"] += uph
                    emp_agg[eid]["count"]   += 1
                emp_agg[eid]["units"] += units
                emp_daily[eid].append((wd, uph))
                if month and uph > 0:
                    month_dept_agg[month][dept]["uph_sum"]   += uph
                    month_dept_agg[month][dept]["uph_count"] += 1
                    month_dept_agg[month][dept]["units"]     += units
                if wd:
                    try:
                        d  = _dt.strptime(wd, "%Y-%m-%d")
                        wk = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
                        week_dept_agg[wk][dept]["units"]     += units
                        if uph > 0:
                            week_dept_agg[wk][dept]["uph_sum"]   += uph
                            week_dept_agg[wk][dept]["uph_count"] += 1
                    except Exception:
                        pass
            total_rows += len(batch)
            if len(batch) < page_size:
                break
            offset += page_size

    # Calculate employee rolling averages
    employee_rolling_avg = []
    for eid, dates_uph in emp_daily.items():
        if not dates_uph:
            continue
        df = pd.DataFrame(dates_uph, columns=['Date', 'UPH'])
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        df = df.sort_values('Date')
        df = df.set_index('Date')
        df['7DayRollingAvg'] = df['UPH'].rolling('7D', min_periods=1).mean()
        df['14DayRollingAvg'] = df['UPH'].rolling('14D', min_periods=1).mean()
        df = df.reset_index()
        for _, row in df.iterrows():
            employee_rolling_avg.append({
                'Date': row['Date'].strftime('%Y-%m-%d'),
                'Employee': emp_name.get(eid, eid),
                'UPH': round(row['UPH'], 2) if pd.notna(row['UPH']) else None,
                '7DayRollingAvg': round(row['7DayRollingAvg'], 2),
                '14DayRollingAvg': round(row['14DayRollingAvg'], 2)
            })
    employee_rolling_avg.sort(key=lambda r: (r["Employee"], r["Date"]))

    if not emp_agg:
        st.session_state["_archived_loaded"] = False
        # Show what we found for debugging
        if hasattr(st, "session_state"):
            st.session_state["_arch_debug"] = f"uph_history rows: {total_rows}, unit_submissions fallback ran: {total_rows == 0}"
        return False

    # ── Build ranked list ─────────────────────────────────────────────────────
    ranked = []
    for i, (eid, agg) in enumerate(
            sorted(emp_agg.items(),
                   key=lambda x: x[1]["uph_sum"] / max(x[1]["count"], 1),
                   reverse=True), 1):
        avg_uph = round(agg["uph_sum"] / max(agg["count"], 1), 2)
        ranked.append({
            "Rank":          i,
            "Department":    emp_dept.get(eid, ""),
            "Shift":         "",
            "Employee Name": emp_name.get(eid, eid),
            "Average UPH":   avg_uph,
            "Record Count":  agg["count"],
            "EmployeeID":    eid,
            "goal_status":   "no_goal",
            "trend":         "insufficient_data",
            "flagged":       False,
            "change_pct":    0,
            "Target UPH":    "—",
            "vs Target":     "—",
        })

    # ── dept_trends ───────────────────────────────────────────────────────────
    dept_trends = []
    for month in sorted(month_dept_agg):
        for dept, vals in month_dept_agg[month].items():
            if vals["uph_count"] > 0:
                dept_trends.append({
                    "Month":       month,
                    "Department":  dept,
                    "Average UPH": round(vals["uph_sum"] / vals["uph_count"], 2),
                    "Count":       vals["uph_count"],
                })

    # ── weekly_summary ────────────────────────────────────────────────────────
    weekly_summary = []
    for wk in sorted(week_dept_agg):
        for dept, vals in week_dept_agg[wk].items():
            weekly_summary.append({
                "Week":        wk,
                "Month":       wk[:7],
                "Department":  dept,
                "Total Units": round(vals["units"]),
                "Avg UPH":     round(vals["uph_sum"] / max(vals["uph_count"], 1), 2),
            })

    # ── dept_report ───────────────────────────────────────────────────────────
    dept_report = {}
    for r2 in ranked:
        dept_report.setdefault(r2.get("Department",""), []).append(r2)

    # ── Build trend_data from per-employee weekly UPH ─────────────────────────
    _trend_weeks = 4
    try:
        import streamlit as _st2
        _trend_weeks = _st2.session_state.get("trend_weeks", 4)
    except Exception:
        pass
    all_weeks_sorted = sorted(
        {wk for emp_wks in emp_week_uph.values() for wk in emp_wks}
    )
    recent_weeks = all_weeks_sorted[-_trend_weeks:] if len(all_weeks_sorted) >= _trend_weeks else all_weeks_sorted

    trend_data = {}
    for eid, week_map in emp_week_uph.items():
        week_avgs = []
        for w in recent_weeks:
            vals = week_map.get(w, [])
            if vals:
                week_avgs.append({"week": w, "avg_uph": round(sum(vals) / len(vals), 2)})

        if len(week_avgs) < 2:
            direction  = "insufficient_data"
            change_pct = 0.0
        else:
            first = week_avgs[0]["avg_uph"]
            last  = week_avgs[-1]["avg_uph"]
            change_pct = round(((last - first) / first * 100) if first else 0, 1)
            if change_pct >= 3:
                direction = "up"
            elif change_pct <= -3:
                direction = "down"
            else:
                direction = "flat"

        trend_data[eid] = {
            "name":       emp_name.get(eid, eid),
            "dept":       emp_dept.get(eid, ""),
            "direction":  direction,
            "weeks":      week_avgs,
            "change_pct": change_pct,
        }

    # ── Apply goals ───────────────────────────────────────────────────────────
    try:
        from goals import build_goal_status as _bgs
        _arch_gs = _bgs(ranked, _cached_targets(), trend_data)
    except Exception:
        _arch_gs = ranked

    st.session_state.update({
        "top_performers":  ranked,
        "goal_status":     _arch_gs,
        "dept_report":     dept_report,
        "dept_trends":     dept_trends,
        "weekly_summary":  weekly_summary,
        "employee_rolling_avg": employee_rolling_avg,
        "employee_risk": [],
        "trend_data":      trend_data,
        "pipeline_done":   True,
        "_archived_loaded": True,
        "_archived_last_refresh_ts": time.time(),
    })
    return True


def page_productivity():
    st.title("📈 Productivity")
    st.caption("UPH rankings, department goals, trend charts, and performance tracking.")

    try:
        from goals          import (analyse_trends, build_goal_status, get_all_targets,
                                    set_dept_target, flag_employee, unflag_employee,
                                    add_note, get_active_flags, load_goals, save_goals)
        from ranker         import build_department_report
        from error_log      import ErrorLog
        from exporter       import export_excel
    except ImportError as e:
        st.error(f"Productivity module error: {e}"); return

    # Nav buttons (no tabs — avoids running all content simultaneously)
    _plan_now = _get_current_plan()
    _all_opts = ["🎯 Dept Goals", "📊 Goal Status", "📈 Trends", "📉 Rolling Avg", "📅 Weekly", "💰 Labor Cost", "📋 Priority List", "🧑‍🏫 Coaching"]
    if _plan_now in ("pro", "business", "admin"):
        PROD_OPTS = _all_opts
    else:
        # Starter tier: core weekly output + ranking-focused view.
        PROD_OPTS = ["📅 Weekly", "📋 Priority List"]

    if "prod_view" not in st.session_state:
        st.session_state.prod_view = PROD_OPTS[0]

    if st.session_state.prod_view not in PROD_OPTS:
        st.session_state.prod_view = PROD_OPTS[0]

    chosen_prod = st.session_state.prod_view
    cols = st.columns(len(PROD_OPTS))
    for i, opt in enumerate(PROD_OPTS):
        if cols[i].button(opt, key=f"pv_{i}", use_container_width=True,
                          type="primary" if chosen_prod == opt else "secondary"):
            st.session_state.prod_view = opt
            st.rerun()

    st.divider()

    _targets_snapshot = _cached_targets()

    class _PS:
        def get(self, k, d=None): return st.session_state.get(k, d)
        def get_output_dir(self): return tempfile.gettempdir()
        def get_dept_target_uph(self, d):
            t = _targets_snapshot.get(d, 0)
            return float(t) if t else 0.0
        def all_mappings(self): return st.session_state.get("mapping", {})

    # ── Helper: re-apply goals to current pipeline data ──────────────────────
    def _reapply_goals():
        if not st.session_state.pipeline_done:
            return
        try:
            targets  = _cached_targets()
            tw       = st.session_state.get("trend_weeks", 4)
            history  = st.session_state.get("history", [])
            mapping  = st.session_state.get("mapping", {})
            trend_data = analyse_trends(history, mapping, weeks=tw) if history else {}
            goal_status = build_goal_status(
                st.session_state.get("top_performers", []), targets, trend_data)
            ps  = _PS()
            log = ErrorLog(tempfile.gettempdir())
            dept_report = build_department_report(
                st.session_state.get("top_performers", []), ps, log)
            st.session_state.goal_status = goal_status
            st.session_state.dept_report = dept_report
            st.session_state.trend_data  = trend_data
        except Exception as _e:
            pass   # silently skip — archived data may lack history

    # Always refresh from archived DB data on Productivity page so it includes
    # prior imports and newly imported rows in one combined view.
    _last_arch_refresh = float(st.session_state.get("_archived_last_refresh_ts", 0.0) or 0.0)
    _arch_refresh_due = (time.time() - _last_arch_refresh) > 120
    if _arch_refresh_due:
        _show_loading = not st.session_state.get("_archived_loaded")
        try:
            if _show_loading:
                with st.spinner("Loading productivity data…"):
                    _build_archived_productivity()
            else:
                _build_archived_productivity()
            st.session_state["_archived_last_refresh_ts"] = time.time()
        except BaseException as _ae:
            _log_app_error("productivity", f"Archive load error: {repr(_ae)[:500]}", detail=traceback.format_exc())

    # Reapply goals only when targets changed (not every click)
    _fresh_targets = _cached_targets()
    _prev_targets  = st.session_state.get("_last_applied_targets")
    if st.session_state.pipeline_done and st.session_state.get("top_performers") and _fresh_targets != _prev_targets:
        try:
            _fresh_td      = st.session_state.get("trend_data", {})
            _fresh_gs      = build_goal_status(
                st.session_state.top_performers, _fresh_targets, _fresh_td)
            st.session_state.goal_status = _fresh_gs
            st.session_state["_last_applied_targets"] = dict(_fresh_targets)
            # Only rebuild dept_report for live data — archived version is already correct
            if not st.session_state.get("_archived_loaded"):
                _ps2  = _PS()
                _log2 = ErrorLog(tempfile.gettempdir())
                st.session_state.dept_report = build_department_report(
                    st.session_state.top_performers, _ps2, _log2)
        except Exception:
            pass

    # ── DEPT GOALS ────────────────────────────────────────────────────────────
    if chosen_prod == "🎯 Dept Goals":
        if not _plan_gate("pro", "Department Goals"):
            return
        st.subheader("Department UPH targets")
        st.caption("Set a UPH target for each department. Type a value and press Enter — it saves immediately and updates all charts.")
        st.caption("Trend window and Top/Bottom % below only affect highlighting/ranking views (Goal Status, Priority, Coaching). They do not change raw UPH history.")

        # Trend window slider lives here
        tw = st.slider("Trend window used for trend scoring (weeks)", 2, 12,
                       st.session_state.get("trend_weeks", 4), key="prod_tw")
        if tw != st.session_state.get("trend_weeks", 4):
            st.session_state.trend_weeks = tw
            _reapply_goals()
            st.rerun()

        st.divider()

        # Top / bottom highlight thresholds
        hc1, hc2 = st.columns(2)
        st.session_state.top_pct = hc1.slider(
            "Top % bucket (green highlight)", 0, 50,
            st.session_state.get("top_pct", 10), key="goals_top_pct"
        )
        st.session_state.bot_pct = hc2.slider(
            "Bottom % bucket (red highlight)", 0, 50,
            st.session_state.get("bot_pct", 10), key="goals_bot_pct"
        )

        st.divider()

        targets   = _cached_targets()
        goals_obj = load_goals()

        # Auto-populate departments from all available sources
        if st.session_state.pipeline_done:
            all_depts = set()
            # From goal_status
            for r in st.session_state.get("goal_status", []):
                if r.get("Department"): all_depts.add(r["Department"])
            # From top_performers
            for r in st.session_state.get("top_performers", []):
                if r.get("Department"): all_depts.add(r["Department"])
            # From employees table (catches archived data with blank uph_history dept)
            for e in (_cached_employees() or []):
                if e.get("department"): all_depts.add(e["department"])
            for d in all_depts:
                if d and d not in targets:
                    set_dept_target(d, 0.0)
                    _raw_cached_targets.clear()
            targets = _cached_targets()

        dept_list  = sorted(targets.keys())
        goal_changed = False

        if not dept_list:
            st.info("No departments yet. Run Import Data first — departments are detected automatically from your CSV.")
        else:
            # Header row
            hc1, hc2, hc3 = st.columns([3, 2, 1])
            hc1.markdown("**Department**")
            hc2.markdown("**UPH Target** *(press Enter to save)*")
            hc3.markdown("")
            st.divider()

            for dept in dept_list:
                cur = float(targets.get(dept, 0) or 0)
                c1, c2, c3 = st.columns([3, 2, 1])
                c1.markdown(f"**{dept}**")

                # Seed text input once — don't overwrite after user changes it
                seed_key = f"goal_seed_{dept}"
                txt_key  = f"goal_txt_{dept}"
                if seed_key not in st.session_state:
                    st.session_state[seed_key] = True
                    # Seed from persisted JSON value — this runs on every fresh session
                    st.session_state[txt_key] = str(int(cur)) if cur else ""
                elif txt_key not in st.session_state:
                    # Key was cleared somehow — re-seed from JSON
                    st.session_state[txt_key] = str(int(cur)) if cur else ""

                c2.text_input("UPH", key=txt_key,
                              label_visibility="collapsed",
                              placeholder="e.g. 18")

                # Parse and save on every render — if value changed, store it
                raw = st.session_state.get(txt_key, "")
                try:
                    new_val = float(raw.strip()) if raw.strip() else 0.0
                except (ValueError, TypeError):
                    new_val = cur

                if abs(new_val - cur) > 0.001:
                    set_dept_target(dept, new_val)
                    _audit("GOAL_TARGET", f"{dept} | {cur} → {new_val}")
                    _raw_cached_targets.clear()
                    goal_changed = True

                if c3.button("✕", key=f"rm_dept_{dept}", help="Remove department"):
                    goals_obj["dept_targets"].pop(dept, None)
                    save_goals(goals_obj)
                    _raw_cached_targets.clear()
                    st.session_state.pop(seed_key, None)
                    st.session_state.pop(txt_key,  None)
                    goal_changed = True

        # Departments are added automatically from the pipeline — no manual add form

        # Reapply goals to all charts after any change
        if goal_changed and st.session_state.pipeline_done:
            _reapply_goals()
            st.toast("✓ Goals updated", icon="🎯")
            st.rerun()

    # ── GOAL STATUS ───────────────────────────────────────────────────────────
    elif chosen_prod == "📊 Goal Status":
        if not _plan_gate("pro", "Goal Status"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        gs = st.session_state.get("goal_status", [])
        # If goal_status is empty but we have top_performers, rebuild it
        if not gs and st.session_state.get("top_performers"):
            try:
                gs = build_goal_status(
                    st.session_state.top_performers, _cached_targets(),
                    st.session_state.get("trend_data", {}))
                st.session_state.goal_status = gs
            except Exception:
                pass
        if not gs:
            st.info("No data yet — run Import Data or check that UPH history exists.")
            return

        total    = len(gs)
        on_goal  = sum(1 for r in gs if r.get("goal_status") == "on_goal")
        below    = sum(1 for r in gs if r.get("goal_status") == "below_goal")
        trending = sum(1 for r in gs if r.get("trend") == "down")
        flagged  = sum(1 for r in gs if r.get("flagged"))
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Employees", total)
        m2.metric("On goal 🟢", on_goal)
        m3.metric("Below goal 🔴", below)
        m4.metric("Trending ↓", trending)
        m5.metric("Flagged 🚩", flagged)
        st.divider()

        depts  = sorted({r.get("Department","") for r in gs if r.get("Department")})
        fc1,fc2 = st.columns(2)
        dept_f  = fc1.selectbox("Filter by department", ["All departments"] + depts, key="gs_dept")
        stat_f  = fc2.multiselect("Filter by status",
                    ["On goal","Below goal","No target set"],
                    default=["On goal","Below goal","No target set"], key="gs_stat")
        stat_map = {"On goal":"on_goal","Below goal":"below_goal","No target set":"no_goal"}
        allowed  = {stat_map[s] for s in stat_f}
        filtered = [r for r in gs
                    if (dept_f == "All departments" or r.get("Department") == dept_f)
                    and r.get("goal_status") in allowed]

        TREND  = {"up":"↑ Improving","down":"↓ Declining","flat":"→ Stable","insufficient_data":"—"}
        STATUS = {"on_goal":"🟢","below_goal":"🔴","no_goal":"⚪"}
        rows   = [{
            "🚩":        "🚩" if r.get("flagged") else "",
            "Dept":      r.get("Department",""),
            "Shift":     r.get("Shift",""),
            "Employee":  r.get("Employee Name",""),
            "Avg UPH":   round(float(r.get("Average UPH",0) or 0),2),
            "Target":    r.get("Target UPH") if r.get("Target UPH") not in ("—", None, "", 0) else None,
            "vs Target": r.get("vs Target") if r.get("vs Target") not in ("—", None, "") else None,
            "Status":    STATUS.get(r.get("goal_status",""),""),
            "Trend":     TREND.get(r.get("trend",""),"—"),
            "Change":    f"{r.get('change_pct',0):+.1f}%",
        } for r in filtered]

        df = pd.DataFrame(rows)
        hl_map = {"🟢":"#1a6b3a","🔴":"#8b1a1a","⚪":"#2a3a4a"}
        def hl(s):
            row = rows[s.name]
            clr = hl_map.get(row.get("Status",""), "#2a3a4a")
            if row.get("Trend","").startswith("↓") and clr == "#1a6b3a": clr = "#7a5c00"
            return [f"background-color:{clr}; color:#ffffff" for _ in s]
        st.dataframe(
            df.style.apply(hl, axis=1),
            use_container_width=True, hide_index=True)
        st.caption("Employees are compared against the UPH target set in Dept Goals. 🟢 Avg UPH ≥ target · 🔴 Avg UPH < target · 🟡 On goal but UPH is declining week-over-week · ⚪ No target set for this department")
        _gs_buf = io.BytesIO()
        df.to_excel(_gs_buf, index=False, engine="openpyxl")
        _gs_buf.seek(0)
        st.download_button("⬇️ Download Goal Status.xlsx", _gs_buf.read(),
                           f"Goal_Status_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_goal_status")

        st.divider()
        st.subheader("🚩 Flag employees for performance tracking")
        emp_labels      = [
            f"{_normalize_label_text(r.get('Employee Name',''))} — "
            f"{_normalize_label_text(r.get('Department',''), max_len=24)} "
            f"({_normalize_label_text(r.get('Shift',''), max_len=18)})"
            for r in filtered
        ]
        active_flag_ids = set(_cached_active_flags().keys())
        if emp_labels:
            _flag_tab1, _flag_tab2 = st.tabs(["Flag individual", "Bulk flag / unflag"])
            with _flag_tab1:
                if "flag_reason_val" not in st.session_state:
                    st.session_state.flag_reason_val = ""
                fc3, fc4, fc5 = st.columns([3,2,1])
                _below_labels = [l for l, r in zip(emp_labels, filtered) if r.get("goal_status") == "below_goal"]
                _flag_default  = emp_labels.index(_below_labels[0]) if _below_labels else 0
                sel_emp = fc3.selectbox("Employee", emp_labels, index=_flag_default, key="flag_sel")
                reason  = fc4.text_input("Reason", value=st.session_state.flag_reason_val,
                                          key="flag_reason", placeholder="Optional reason")
                if fc5.button("Flag", type="primary", use_container_width=True):
                    idx    = emp_labels.index(sel_emp)
                    emp    = filtered[idx]
                    emp_id = str(emp.get("EmployeeID", emp.get("Employee Name","")))
                    flag_employee(emp_id, emp.get("Employee Name",""), emp.get("Department",""), reason.strip())
                    _audit("FLAG", f"{emp.get('Employee Name','')} | dept={emp.get('Department','')} | reason={reason.strip()}")
                    _raw_cached_active_flags.clear()
                    _raw_cached_all_coaching_notes.clear()
                    _raw_cached_coaching_notes_for.clear()
                    st.session_state.flag_reason_val = ""
                    st.session_state.pop("flag_reason", None)
                    st.toast(f"✓ {emp.get('Employee Name','')} flagged", icon="🚩")
                    st.rerun()
            with _flag_tab2:
                # ── Currently flagged summary ──────────────────────────────
                _currently_flagged = [
                    (_lbl, _r) for _lbl, _r in zip(emp_labels, filtered)
                    if str(_r.get("EmployeeID", _r.get("Employee Name",""))) in active_flag_ids
                ]
                if _currently_flagged:
                    st.markdown(
                        f"<div style='background:#FFF3E0;border-left:4px solid #E65100;"
                        f"border-radius:0 6px 6px 0;padding:10px 14px;margin-bottom:12px;'>"
                        f"<b style='color:#BF360C;font-size:12px;text-transform:uppercase;"
                        f"letter-spacing:.05em;'>🚩 {len(_currently_flagged)} Active Flag(s)</b><br>"
                        + "".join(
                            f"<span style='color:#BF360C;font-size:13px;'>• {_html_mod.escape(_lbl)}</span><br>"
                            for _lbl, _ in _currently_flagged
                        )
                        + "</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.info("No employees currently flagged.")

                st.caption("Select employees below — 🚩 marks those already flagged.")
                # Build display labels with 🚩 prefix for already-flagged employees
                _bulk_display = []
                for _lbl, _r in zip(emp_labels, filtered):
                    _eid = str(_r.get("EmployeeID", _r.get("Employee Name", "")))
                    _bulk_display.append(f"🚩 {_lbl}" if _eid in active_flag_ids else _lbl)

                bulk_sel_disp = st.multiselect("Select employees", _bulk_display, key="bulk_flag_sel")
                # Strip the 🚩 prefix to recover the plain label for lookup
                bulk_sel      = [l.removeprefix("🚩 ") for l in bulk_sel_disp]

                bulk_reason = st.text_input("Reason (applies to all)", key="bulk_flag_reason", placeholder="Optional")
                bc1, bc2    = st.columns(2)
                if bc1.button("🚩 Flag all selected", type="primary", use_container_width=True, key="bulk_flag_btn"):
                    if bulk_sel:
                        for lbl in bulk_sel:
                            _bi = emp_labels.index(lbl)
                            _be = filtered[_bi]
                            _beid = str(_be.get("EmployeeID", _be.get("Employee Name","")))
                            flag_employee(_beid, _be.get("Employee Name",""), _be.get("Department",""), bulk_reason.strip())
                        _raw_cached_active_flags.clear(); _raw_cached_all_coaching_notes.clear(); _raw_cached_coaching_notes_for.clear()
                        st.toast(f"✓ {len(bulk_sel)} employee(s) flagged", icon="🚩"); st.rerun()
                if bc2.button("✓ Unflag all selected", type="secondary", use_container_width=True, key="bulk_unflag_btn"):
                    if bulk_sel:
                        for lbl in bulk_sel:
                            _bi = emp_labels.index(lbl)
                            _be = filtered[_bi]
                            _beid = str(_be.get("EmployeeID", _be.get("Employee Name","")))
                            if _beid in active_flag_ids:
                                unflag_employee(_beid)
                        _raw_cached_active_flags.clear()
                        st.toast(f"✓ {len(bulk_sel)} employee(s) unflagged", icon="✅"); st.rerun()

    # ── TRENDS ────────────────────────────────────────────────────────────────
    elif chosen_prod == "📈 Trends":
        if not _plan_gate("pro", "Trend Analysis"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        trends = st.session_state.dept_trends
        if not trends: st.info("No trend data yet."); return
        df_t = pd.DataFrame(trends)
        # Date range filter
        months = sorted(df_t["Month"].unique()) if "Month" in df_t.columns else []
        if months:
            import re as _re
            min_m, max_m = months[0], months[-1]
            # Persist applied range in session state
            # Reset range if data was refreshed (new pipeline run)
            _data_key = f"trends_data_key_{min_m}_{max_m}"
            if st.session_state.get("_trends_data_key") != _data_key:
                st.session_state.trend_applied_from = min_m
                st.session_state.trend_applied_to   = max_m
                st.session_state["_trends_data_key"] = _data_key
            elif "trend_applied_from" not in st.session_state:
                st.session_state.trend_applied_from = min_m
            if "trend_applied_to" not in st.session_state:
                st.session_state.trend_applied_to   = max_m

            tr1, tr2, tr3 = st.columns([3, 3, 1])
            from_raw = tr1.text_input("From (YYYY-MM)", value=st.session_state.trend_applied_from,
                                       key="trend_from", placeholder="e.g. 2024-03")
            to_raw   = tr2.text_input("To (YYYY-MM)",   value=st.session_state.trend_applied_to,
                                       key="trend_to",   placeholder="e.g. 2025-03")
            if tr3.button("Apply", type="primary", use_container_width=True, key="trend_apply"):
                fm = from_raw.strip()
                tm = to_raw.strip()
                st.session_state.trend_applied_from = fm if _re.match(r"\d{4}-\d{2}", fm) else min_m
                st.session_state.trend_applied_to   = tm if _re.match(r"\d{4}-\d{2}", tm) else max_m
                st.rerun()

            from_month = st.session_state.trend_applied_from
            to_month   = st.session_state.trend_applied_to
            mask = (df_t["Month"] >= from_month) & (df_t["Month"] <= to_month)
            df_t = df_t[mask]
        st.subheader("Average UPH over time by department")
        st.caption("Each line shows the average UPH for that department across all employees for each month. Use From/To to zoom into a date range.")
        try:
            import math as _math2
            df_t_clean = df_t[df_t["Average UPH"].apply(
                lambda x: _math2.isfinite(float(x)) if x is not None else False)]
            if not df_t_clean.empty:
                st.line_chart(df_t_clean.pivot(index="Month", columns="Department", values="Average UPH"),
                              use_container_width=True)
        except Exception: pass
        st.dataframe(df_t, use_container_width=True, hide_index=True)
        _tr_buf = io.BytesIO()
        df_t.to_excel(_tr_buf, index=False, engine="openpyxl")
        _tr_buf.seek(0)
        st.download_button("⬇️ Download Trends.xlsx", _tr_buf.read(),
                           f"Trends_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_trends_prod")

    # ── ROLLING AVG ───────────────────────────────────────────────────────────
    elif chosen_prod == "📉 Rolling Avg":
        if not _plan_gate("pro", "Rolling Averages"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        rolling = st.session_state.get("employee_rolling_avg", [])
        if not rolling: st.info("No rolling average data yet."); return
        df_r = pd.DataFrame(rolling)

        # Employee filter
        employees = sorted(df_r["Employee"].unique()) if "Employee" in df_r.columns else []
        if employees:
            emp_sel = st.multiselect("Filter employees", employees, key="roll_emp_sel")
            if emp_sel:
                df_r = df_r[df_r["Employee"].isin(emp_sel)]

        st.subheader("7-Day and 14-Day Rolling Average UPH per Employee")
        st.caption("Each point shows the rolling average UPH for that employee on that date.")

        # Select rolling period
        roll_period = st.selectbox("Rolling Period", ["7-Day", "14-Day"], key="roll_period")

        # Chart
        try:
            col_name = "7DayRollingAvg" if roll_period == "7-Day" else "14DayRollingAvg"
            df_r_clean = df_r.dropna(subset=[col_name])
            if not df_r_clean.empty:
                st.line_chart(df_r_clean.pivot(index="Date", columns="Employee", values=col_name),
                              use_container_width=True)
        except Exception: pass
        st.dataframe(df_r, use_container_width=True, hide_index=True)
        _r_buf = io.BytesIO()
        df_r.to_excel(_r_buf, index=False, engine="openpyxl")
        _r_buf.seek(0)
        st.download_button("⬇️ Download RollingAvg.xlsx", _r_buf.read(),
                           f"RollingAvg_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_rolling_prod")

    # ── RISK ASSESSMENT ───────────────────────────────────────────────────────
    elif chosen_prod == "⚠️ Risk Assessment":
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        risk = st.session_state.get("employee_risk", [])
        if not risk: st.info("No risk assessment data yet."); return
        df_risk = pd.DataFrame(risk)

        # Filter by risk level
        risk_levels = ["high", "medium", "low"]
        sel_levels = st.multiselect("Filter by Risk Level", risk_levels, default=risk_levels, key="risk_level_sel")
        if sel_levels:
            df_risk = df_risk[df_risk["Risk Level"].isin(sel_levels)]

        st.subheader("Employee Risk Assessment")
        st.caption("Risk levels based on under goal streak, downward trend, and below department average.")

        # Summary
        level_counts = df_risk["Risk Level"].value_counts()
        col1, col2, col3 = st.columns(3)
        col1.metric("High Risk", level_counts.get("high", 0))
        col2.metric("Medium Risk", level_counts.get("medium", 0))
        col3.metric("Low Risk", level_counts.get("low", 0))

        st.dataframe(df_risk, use_container_width=True, hide_index=True)
        _risk_buf = io.BytesIO()
        df_risk.to_excel(_risk_buf, index=False, engine="openpyxl")
        _risk_buf.seek(0)
        st.download_button("⬇️ Download RiskAssessment.xlsx", _risk_buf.read(),
                           f"RiskAssessment_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_risk_prod")

    # ── WEEKLY ────────────────────────────────────────────────────────────────
    elif chosen_prod == "📅 Weekly":
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        weekly = st.session_state.weekly_summary
        if not weekly: st.info("No weekly data yet."); return
        df_w = pd.DataFrame(weekly)

        # Date range filter
        all_from = sorted(df_w["From"].dropna().unique()) if "From" in df_w.columns else []
        all_to   = sorted(df_w["To"].dropna().unique())   if "To"   in df_w.columns else []
        all_dates = sorted(set(all_from + all_to))

        wk1, wk2 = st.columns(2)
        if all_dates:
            d_from = wk1.date_input("From", value=datetime.strptime(all_dates[0], "%Y-%m-%d").date(),
                                    key="wk_from")
            d_to   = wk2.date_input("To",   value=datetime.strptime(all_dates[-1], "%Y-%m-%d").date(),
                                    key="wk_to")
            mask = pd.Series([True] * len(df_w))
            if "From" in df_w.columns:
                mask &= df_w["From"] >= d_from.isoformat()
            if "To" in df_w.columns:
                mask &= df_w["To"] <= d_to.isoformat()
            df_w = df_w[mask]

        st.subheader("Total units per week by department")
        st.caption("Each row represents one department's output for a given week. Use the date range to filter.")

        # Chart
        try:
            df_w_clean = df_w.dropna(subset=["Total Units"])
            if not df_w_clean.empty:
                chart_label = df_w_clean["Week Range"] if "Week Range" in df_w_clean.columns else df_w_clean["Week"]
                chart_df = df_w_clean.copy()
                chart_df["Period"] = chart_label
                st.line_chart(chart_df.pivot(index="Period", columns="Department", values="Total Units"),
                              use_container_width=True)
        except Exception: pass

        # Table grouped by week
        display_cols = [c for c in ["Week Range", "Department", "Avg UPH", "Total Units", "Record Count"]
                        if c in df_w.columns]
        if not display_cols:
            display_cols = [c for c in df_w.columns if c not in ("From", "To")]

        weeks_sorted = sorted(df_w["Week"].unique()) if "Week" in df_w.columns else []
        for wk in weeks_sorted:
            wk_slice = df_w[df_w["Week"] == wk]
            if wk_slice.empty:
                continue
            wk_range = wk_slice["Week Range"].iloc[0] if "Week Range" in wk_slice.columns else wk
            st.markdown(f"**{wk}** &nbsp;&nbsp; {wk_range}")
            show_cols = [c for c in display_cols if c != "Week Range"]
            st.dataframe(wk_slice[show_cols], use_container_width=True, hide_index=True)

        _wk_buf = io.BytesIO()
        df_w[display_cols].to_excel(_wk_buf, index=False, engine="openpyxl")
        _wk_buf.seek(0)
        st.download_button("⬇️ Download Weekly.xlsx", _wk_buf.read(),
                           f"Weekly_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_weekly")

    elif chosen_prod == "💰 Labor Cost":
        if not _plan_gate("pro", "Labor Cost Impact"):
            return

        st.subheader("Labor Cost Impact Analysis")
        st.caption("See the dollar impact of employee performance vs targets. Enter your average hourly wage to calculate.")

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        _hourly_wage = st.number_input("Average hourly wage ($)", min_value=0.0, value=18.0, step=0.50, key="labor_wage")
        _targets_map = _cached_targets()

        # Build labor cost table
        _lc_rows = []
        for r in gs:
            name = r.get("Employee", r.get("Employee Name", ""))
            dept = r.get("Department", "")
            uph  = r.get("Avg UPH", r.get("Average UPH"))
            target = r.get("Target UPH")
            hours = r.get("Hours Worked") or r.get("HoursWorked")

            if target in ("—", None, "", 0):
                target = _targets_map.get(dept, 0)

            if not uph or target in ("—", None, "", 0):
                continue
            try:
                uph = float(uph)
                target = float(target)
            except (ValueError, TypeError):
                continue

            if target <= 0:
                continue

            # Estimate hours from data or default to 40
            try:
                hours = float(hours) if hours and hours not in ("—", None, "") else 40.0
            except (ValueError, TypeError):
                hours = 40.0

            expected_units = target * hours
            actual_units = uph * hours
            unit_diff = actual_units - expected_units
            # Cost per unit = wage / target_uph
            cost_per_unit = _hourly_wage / target if target > 0 else 0
            dollar_impact = unit_diff * cost_per_unit

            _lc_rows.append({
                "Employee": name,
                "Department": dept,
                "Avg UPH": round(uph, 1),
                "Target": round(target, 1),
                "UPH Diff": round(uph - target, 1),
                "Est. Hours": round(hours, 1),
                "Unit Diff": round(unit_diff, 0),
                "$ Impact": round(dollar_impact, 2),
            })

        if not _lc_rows:
            st.info("No employees with both UPH and targets set.")
            return

        df_lc = pd.DataFrame(_lc_rows).sort_values("$ Impact")

        # Summary metrics
        total_loss = sum(r["$ Impact"] for r in _lc_rows if r["$ Impact"] < 0)
        total_gain = sum(r["$ Impact"] for r in _lc_rows if r["$ Impact"] > 0)
        net_impact = total_loss + total_gain

        # TOP 3 UNDERPERFORMERS — Personal & Actionable
        st.subheader("⚠️ Top Cost Impact")
        underperformers = [r for r in _lc_rows if r["$ Impact"] < 0]
        underperformers.sort(key=lambda x: x["$ Impact"])  # Most negative first
        top_3 = underperformers[:3]
        
        if top_3:
            top_3_cost = sum(emp["$ Impact"] for emp in top_3)
            st.error(f"🔴 Top 3 underperformers costing **${abs(top_3_cost):,.0f}** this week")
            
            for i, emp in enumerate(top_3, 1):
                col1, col2, col3 = st.columns([2, 1.5, 1.5])
                col1.write(f"**{i}. {emp['Employee']}** ({emp['Department']})")
                col2.metric("UPH Gap", f"{emp['UPH Diff']:.1f}", "Below target")
                col3.metric("Weekly Cost", f"${abs(emp['$ Impact']):,.0f}", "Lost productivity")

        st.divider()
        
        # Summary metrics
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("Lost from Underperformance", f"-${abs(total_loss):,.0f}", delta_color="inverse")
        mc2.metric("Gained from Overperformance", f"+${total_gain:,.0f}")
        mc3.metric("Net Impact", f"${net_impact:,.0f}",
                    delta=f"${net_impact:,.0f}", delta_color="normal")

        st.markdown("---")

        # Color code: red for negative, green for positive
        st.subheader("Employee Breakdown")
        st.caption("Negative = costing you money vs target. Positive = saving you money.")

        def _color_impact(val):
            try:
                v = float(val)
                if v < 0: return "color: #FF4444"
                if v > 0: return "color: #44AA44"
            except: pass
            return ""

        styled = df_lc.style.applymap(_color_impact, subset=["$ Impact", "UPH Diff"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Department summary
        st.subheader("Department Summary")
        dept_summary = {}
        for r in _lc_rows:
            d = r["Department"]
            if d not in dept_summary:
                dept_summary[d] = {"Department": d, "Employees": 0, "Total $ Impact": 0}
            dept_summary[d]["Employees"] += 1
            dept_summary[d]["Total $ Impact"] += r["$ Impact"]
        df_dept = pd.DataFrame(dept_summary.values())
        df_dept["Total $ Impact"] = df_dept["Total $ Impact"].round(2)
        st.dataframe(df_dept, use_container_width=True, hide_index=True)

    # ── PRIORITY LIST ──────────────────────────────────────────────────────────────
    elif chosen_prod == "📋 Priority List":
        st.subheader("📋 Priority List")
        st.caption("Employees below goal ranked by risk (combines trend + streak + variance).")

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        history = st.session_state.get("history", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        # Filter for below_goal employees
        below_goal = [r for r in gs if r.get("goal_status") == "below_goal"]
        if not below_goal:
            st.success("✅ All employees are meeting their goals!")
            return

        # Calculate risk for all below-goal employees
        def _calc_priority_risk_level(emp, history):
            """Calculate performance risk level."""
            risk_score = 0.0
            details = {"trend_score": 0, "streak_score": 0, "variance_score": 0}
            
            trend = emp.get("trend", "insufficient_data")
            if trend == "down":
                risk_score += 4
                details["trend_score"] = 4
            elif trend == "flat":
                risk_score += 1
                details["trend_score"] = 1
            elif trend == "up":
                risk_score -= 2
                details["trend_score"] = -2
            
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            target_uph = emp.get("Target UPH", "—")
            under_goal_streak = 0
            
            if history and target_uph != "—":
                try:
                    target = float(target_uph)
                    emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                    if emp_history:
                        sorted_hist = sorted(emp_history, key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                        for r in reversed(sorted_hist):
                            try:
                                uph_val = float(r.get("UPH", 0) or 0)
                                if uph_val < target:
                                    under_goal_streak += 1
                                else:
                                    break
                            except (ValueError, TypeError):
                                break
                except (ValueError, TypeError):
                    pass
            
            if under_goal_streak >= 7:
                risk_score += 5
                details["streak_score"] = 5
            elif under_goal_streak >= 5:
                risk_score += 4
                details["streak_score"] = 4
            elif under_goal_streak >= 3:
                risk_score += 2.5
                details["streak_score"] = 2.5
            elif under_goal_streak >= 1:
                risk_score += 0.5
                details["streak_score"] = 0.5
            
            uph_values = []
            if history:
                emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                for r in emp_history:
                    try:
                        uph_val = float(r.get("UPH", 0) or 0)
                        if uph_val > 0:
                            uph_values.append(uph_val)
                    except (ValueError, TypeError):
                        pass
            
            if len(uph_values) >= 3:
                avg_uph = sum(uph_values) / len(uph_values)
                variance = sum((x - avg_uph) ** 2 for x in uph_values) / len(uph_values)
                std_dev = variance ** 0.5
                coeff_variation = (std_dev / avg_uph * 100) if avg_uph > 0 else 0
                
                if coeff_variation > 30:
                    risk_score += 3
                    details["variance_score"] = 3
                elif coeff_variation > 20:
                    risk_score += 1.5
                    details["variance_score"] = 1.5
                elif coeff_variation > 10:
                    risk_score += 0.5
                    details["variance_score"] = 0.5
                
                details["variance_pct"] = round(coeff_variation, 1)
            
            details["under_goal_streak"] = under_goal_streak
            details["total_score"] = round(risk_score, 1)
            
            if risk_score >= 7:
                return "🔴 High", risk_score, details
            elif risk_score >= 4:
                return "🟡 Medium", risk_score, details
            else:
                return "🟢 Low", risk_score, details

        # Score all below-goal employees
        priority_list = []
        for emp in below_goal:
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            risk_level, risk_score, risk_details = _calc_priority_risk_level(emp, history)
            priority_list.append({
                "name": emp.get("Employee", emp.get("Employee Name", "Unknown")),
                "department": emp.get("Department", ""),
                "emp_id": emp_id,
                "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
                "target_uph": emp.get("Target UPH", "—"),
                "trend": emp.get("trend", "—"),
                "change_pct": emp.get("change_pct", 0.0),
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_details": risk_details,
            })

        # Sort by risk score (highest first)
        priority_list.sort(key=lambda x: x["risk_score"], reverse=True)

        # Summary
        high_risk = len([r for r in priority_list if r["risk_level"] == "🔴 High"])
        med_risk = len([r for r in priority_list if r["risk_level"] == "🟡 Medium"])
        low_risk = len([r for r in priority_list if r["risk_level"] == "🟢 Low"])
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Below Goal", len(priority_list))
        col2.metric("🔴 High Risk", high_risk)
        col3.metric("🟡 Medium Risk", med_risk)
        col4.metric("🟢 Low Risk", low_risk)

        st.divider()

        # Display table
        table_data = []
        for r in priority_list:
            table_data.append({
                "Risk": r["risk_level"],
                "Name": r["name"],
                "Department": r["department"],
                "Current": r["avg_uph"],
                "Target": r["target_uph"],
                "Trend": r["trend"],
                "Streak": f"{r['risk_details']['under_goal_streak']}d",
                "Score": r["risk_score"],
            })

        df_prio = pd.DataFrame(table_data)
        
        def _color_risk(val):
            if "🔴" in str(val):
                return "background-color: #ffcccc; color: #8b0000"
            elif "🟡" in str(val):
                return "background-color: #fff9e6; color: #ff6600"
            elif "🟢" in str(val):
                return "background-color: #e6ffe6; color: #008000"
            return ""

        styled = df_prio.style.applymap(_color_risk, subset=["Risk"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Download
        csv_buf = df_prio.to_csv(index=False)
        st.download_button("⬇️ Download priority list", csv_buf, f"priority_list_{date.today()}.csv", "text/csv")

    # ── COACHING CORNER ───────────────────────────────────────────────────────────
    elif chosen_prod == "🧑‍🏫 Coaching":
        if not _plan_gate("pro", "Coaching"):
            return
        st.subheader("🧑‍🏫 Who Needs Coaching?")
        st.caption("Top 3 employees who need performance coaching based on trend + goal status + context.")

        # Define available context tags
        CONTEXT_TAGS = [
            "New employee",
            "Cross-training",
            "Equipment issues",
            "Shift change",
            "Short staffed",
        ]

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        history = st.session_state.get("history", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        # Filter for below_goal employees
        coaching_candidates = [r for r in gs if r.get("goal_status") == "below_goal"]
        if not coaching_candidates:
            st.success("✅ All employees are meeting their goals! Great work!")
            return

        # Load active flags and their context tags
        active_flags_data = get_active_flags()

        # Calculate coaching score for each candidate


        def _calc_coaching_score(emp, context_tags):
            """Score an employee for coaching need. Higher = more urgent.
            Context tags reduce urgency (e.g., new employees need coaching but lower priority).
            Also returns whether they meet strict auto-flag criteria."""
            score = 0.0
            reasons = []

            # Factor 1: How far below goal
            target = emp.get("Target UPH") or 0
            avg_uph = emp.get("Average UPH") or 0
            under_goal = False
            if target and target != "—":
                try:
                    target = float(target)
                    avg_uph = float(avg_uph)
                    pct_below = ((target - avg_uph) / target * 100) if target > 0 else 0
                    score += abs(pct_below) * 0.5  # 0.5 points per percent below
                    reasons.append(f"{pct_below:.1f}% below target")
                    under_goal = True
                except (ValueError, TypeError):
                    pass

            # Factor 2: Negative trend (direction + magnitude)
            trend = emp.get("trend", "insufficient_data")
            change_pct = emp.get("change_pct", 0.0)
            trending_down = False
            try:
                change_pct = float(change_pct)
            except (ValueError, TypeError):
                change_pct = 0.0

            if trend == "down":
                score += 5  # Declining is a red flag
                reasons.append(f"Declining trend ({change_pct:+.1f}%)")
                trending_down = True
            elif trend == "flat":
                score += 2  # Flat is concerning when below goal
                if change_pct <= -3 or change_pct >= 3:
                    reasons.append(f"Flat trend ({change_pct:+.1f}%)")
                else:
                    reasons.append("Flat trend")
            elif change_pct < 0:
                score += 1  # Slight decline
                reasons.append(f"Slight decline ({change_pct:+.1f}%)")

            # Factor 3: Streak (consecutive entries below their average)
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            streak = 0
            emp_avg = avg_uph
            if history and emp_avg:
                emp_history = [r for r in history
                               if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                if emp_history:
                    # Sort by date/week if available
                    sorted_hist = sorted(emp_history,
                                        key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                    # Count consecutive entries below the employee's average
                    for r in reversed(sorted_hist):
                        try:
                            uph_val = float(r.get("UPH", 0) or 0)
                            if uph_val < emp_avg:
                                streak += 1
                            else:
                                break
                        except (ValueError, TypeError):
                            break
                    if streak >= 3:
                        score += streak * 0.3
                        reasons.append(f"{streak} consecutive entries below avg")

            # Strict auto-flag criteria: under goal AND trending down AND multi-day streak
            meets_auto_flag_criteria = (under_goal and trending_down and streak >= 3)

            # Factor 4: Context modifier — reduce urgency based on situational factors
            context_impact = []
            if "New employee" in context_tags:
                score *= 0.5  # New employees are less urgent
                context_impact.append("(New — expect ramp-up period)")
            if "Cross-training" in context_tags:
                score *= 0.6  # Cross-training reduces urgency
                context_impact.append("(In cross-training)")
            if "Equipment issues" in context_tags:
                score *= 0.4  # Equipment issues are not coaching-related
                context_impact.append("(Address equipment first)")
            if "Shift change" in context_tags:
                score *= 0.5  # Shift changes take time to adjust
                context_impact.append("(Recent shift change)")
            if "Short staffed" in context_tags:
                score *= 0.7  # Staffing is a workload not coaching issue
                context_impact.append("(Team capacity issue)")

            return score, reasons, context_impact, meets_auto_flag_criteria

        # Score and rank all candidates
        scored = []
        for emp in coaching_candidates:
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            # Get context tags for this employee from their flag
            emp_context = []
            if emp_id in active_flags_data:
                flag_info = active_flags_data[emp_id]
                emp_context = flag_info.get("context_tags", [])

            score, reasons, context_impact, meets_auto_flag_criteria = _calc_coaching_score(emp, emp_context)
            risk_level, risk_score, risk_details = _calc_risk_level(emp, history)
            
            emp_name = emp.get("Employee", emp.get("Employee Name", "Unknown"))
            scored.append({
                "employee": emp_name,
                "department": emp.get("Department", ""),
                "emp_id": emp_id,
                "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
                "target_uph": emp.get("Target UPH", "—"),
                "trend": emp.get("trend", "—"),
                "change_pct": emp.get("change_pct", 0.0),
                "score": score,
                "reasons": reasons,
                "context_tags": emp_context,
                "context_impact": context_impact,
                "meets_auto_flag_criteria": meets_auto_flag_criteria,
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_details": risk_details,
            })

        # Sort by score (descending) and take top 3
        scored.sort(key=lambda x: x["score"], reverse=True)
        top_3 = scored[:3]

        if not top_3:
            st.success("✅ All employees are meeting their goals!")
            return

        # Display each in an expandable card
        for idx, emp in enumerate(top_3, 1):
            with st.container(border=True):
                col1, col2, col3 = st.columns([2, 1, 1])
                col1.markdown(f"### #{idx} — {emp['employee']}")
                col2.markdown(f"**{emp['department']}**")

                # Status badge
                if emp['context_tags']:
                    status_text = "Under goal BUT " + ", ".join(emp['context_tags']).lower()
                    col3.markdown(f"*{status_text}*")
                else:
                    col3.markdown("⚠️ **Needs coaching**")

                st.divider()

                # Current performance
                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("Current UPH", f"{emp['avg_uph']}", delta=f"{emp['change_pct']:+.1f}%")
                mc2.metric("Target UPH", emp['target_uph'] if emp['target_uph'] != "—" else "—")
                mc3.metric("Trend", emp['trend'].replace("_", " ").title())
                mc4.markdown(f"<div style='text-align: center; padding: 10px;'><div style='font-size: 2em;'>{emp['risk_level'].split()[0]}</div><div style='font-size: 0.8em;'>Risk</div></div>", unsafe_allow_html=True)

                # Risk breakdown (collapsed)
                with st.expander("📊 Risk breakdown", expanded=False):
                    rd = emp['risk_details']
                    rc1, rc2, rc3 = st.columns(3)
                    rc1.metric("Trend score", rd['trend_score'])
                    rc2.metric("Under-goal streak", f"{rd['under_goal_streak']} days")
                    rc3.metric("Variance", f"{rd.get('variance_pct', 0):.1f}%")
                    st.caption(f"Total risk score: {rd['total_score']}")

                # Context tags display
                if emp['context_tags']:
                    st.markdown("#### Context")
                    tag_cols = st.columns(len(emp['context_tags']))
                    for tag_col, tag in zip(tag_cols, emp['context_tags']):
                        tag_col.markdown(f"🏷️ **{tag}**", help="This context reduces coaching priority")

                # Why they need coaching
                st.markdown("#### Performance Issues")
                for reason in emp["reasons"]:
                    st.markdown(f"• {reason}")

                # Context impact notes
                if emp['context_impact']:
                    st.markdown("#### Context Notes")
                    for note in emp['context_impact']:
                        st.info(note, icon="📌")

                # Suggested actions based on trend and gap
                st.markdown("#### Suggested Actions")
                actions = []

                # Equipment-specific action
                if "Equipment issues" in emp['context_tags']:
                    actions.append("**Fix the equipment first** — Resolve tool/system issues before coaching on performance.")

                # New employee specific action
                if "New employee" in emp['context_tags']:
                    actions.append("**Structured onboarding check** — Ensure they have proper training and mentoring.")

                # Cross-training specific action
                if "Cross-training" in emp['context_tags']:
                    actions.append("**Support the transition** — Provide extra mentoring during skill-building phase.")

                # Shift change specific action
                if "Shift change" in emp['context_tags']:
                    actions.append("**Allow adjustment time** — Follow up in 2 weeks to see if new rhythm improves performance.")

                # Staffing specific action
                if "Short staffed" in emp['context_tags']:
                    actions.append("**Increase team capacity** — Hiring or task redistribution may solve this faster than coaching.")

                # Trend-based actions (only if not covered by context)
                if not any(tag in emp['context_tags'] for tag in ["Equipment issues", "Short staffed"]):
                    if emp["trend"] == "down":
                        actions.append("**Identify obstacles** — Ask what's changed. Look for workload spikes or personal issues.")
                    elif emp["trend"] == "flat":
                        actions.append("**Break the plateau** — Try new training, task rotation, or peer mentoring.")

                # Gap-based actions
                try:
                    target = float(emp["target_uph"]) if emp["target_uph"] != "—" else 0
                    if target > 0:
                        gap = target - emp["avg_uph"]
                        if gap > 5 and "New employee" not in emp['context_tags']:
                            actions.append("**Major gap** — Structured improvement plan with weekly check-ins.")
                        elif gap > 2 and "New employee" not in emp['context_tags']:
                            actions.append("**1-on-1 coaching** — Discuss goals and what support they need.")
                except (ValueError, TypeError):
                    pass

                # Default actions
                if not actions:
                    actions.append("**1-on-1 conversation** — Discuss performance, barriers, and support.")

                for action in actions:
                    st.markdown(f"→ {action}")

                st.divider()

                # Context tagging section
                st.markdown("#### Add Context")
                st.caption("Select tags that explain underperformance (coaching urgency will adjust):")
                selected_context = st.multiselect(
                    "Context tags:",
                    CONTEXT_TAGS,
                    default=emp['context_tags'],
                    key=f"context_{emp['emp_id']}",
                    label_visibility="collapsed",
                )

                # Save context tags if changed
                if selected_context != emp['context_tags']:
                    try:
                        flags = get_active_flags()
                        if emp['emp_id'] in flags:
                            flags[emp['emp_id']]["context_tags"] = selected_context
                            # Save back to goals
                            goals_data = load_goals()
                            if emp['emp_id'] in goals_data["flagged_employees"]:
                                goals_data["flagged_employees"][emp['emp_id']]["context_tags"] = selected_context
                                save_goals(goals_data)
                            st.success("✓ Context tags updated")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error saving context: {e}")

                # Add quick coaching note button
                note_text = st.text_input(
                    f"Quick note for {emp['employee']}:",
                    key=f"coach_note_{emp['emp_id']}",
                    placeholder="e.g. 'Discussed new tools, plan to follow up Friday'",
                )
                if note_text:
                    if st.button(f"Log note", key=f"coach_save_{emp['emp_id']}", use_container_width=True):
                        try:
                            add_note(emp['emp_id'], note_text)
                            st.success(f"✅ Note saved for {emp['employee']}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error saving note: {e}")

                # Flag employee
                is_flagged = emp['emp_id'] in active_flags_data
                
                # Show auto-flag criteria status
                if not is_flagged:
                    if emp['meets_auto_flag_criteria']:
                        st.success("✅ Meets auto-flag criteria: under goal + trending down + multi-day streak", icon="✅")
                    else:
                        # Explain what's missing
                        missing = []
                        if emp['trend'] != "down":
                            missing.append("not trending down")
                        if not any(r for r in emp['reasons'] if "consecutive" in r):
                            missing.append("no multi-day streak")
                        missing_text = " + ".join(missing) if missing else "one or more criteria"
                        st.warning(f"⚠️ Doesn't meet auto-flag criteria ({missing_text}). Coaching may still help.", icon="⚠️")

                flag_col1, flag_col2 = st.columns([1, 1])
                if is_flagged:
                    if flag_col1.button(f"🚩 Unflag", key=f"unflag_{emp['emp_id']}", use_container_width=True):
                        try:
                            unflag_employee(emp['emp_id'])
                            st.success(f"✓ {emp['employee']} unflagged")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error unflagging: {e}")
                else:
                    if flag_col1.button(f"🚩 Flag for tracking", key=f"flag_{emp['emp_id']}", use_container_width=True):
                        try:
                            flag_employee(emp['emp_id'], emp['employee'], emp['department'],
                                         reason="Coaching priority: " + ", ".join(emp["reasons"][:2]))
                            # Save context tags with the flag
                            goals_data = load_goals()
                            if emp['emp_id'] in goals_data["flagged_employees"]:
                                goals_data["flagged_employees"][emp['emp_id']]["context_tags"] = selected_context
                                save_goals(goals_data)
                            st.success(f"✓ {emp['employee']} flagged for tracking")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error flagging: {e}")



# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EMAIL SETUP (preserved from previous version)
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_period_dates(period: str):
    """Convert a named period string to (start_date, end_date)."""
    from datetime import timedelta
    today = date.today()
    if period == "Prior day":
        return today - timedelta(days=1), today - timedelta(days=1)
    elif period == "Current week":
        return today - timedelta(days=today.weekday()), today
    elif period == "Prior week":
        end = today - timedelta(days=today.weekday() + 1)
        return end - timedelta(days=6), end
    elif period == "Prior month":
        first_of_this = today.replace(day=1)
        end = first_of_this - timedelta(days=1)
        return end.replace(day=1), end
    return today - timedelta(days=1), today - timedelta(days=1)


def _build_period_report(d_start, d_end, dept_choice: str, depts: list,
                          gs: list, targets: dict, tenant_id: str = "",
                          plan_name: str = "starter"):
    """
    Build Excel report bytes, subject, and HTML body for a date range.
    d_start / d_end are date objects. Returns (xl_bytes, subj, body).
    """
    from collections import defaultdict as _dc
    today = date.today()

    from_iso     = d_start.isoformat()
    to_iso       = d_end.isoformat()
    period_label = from_iso if from_iso == to_iso else f"{from_iso} – {to_iso}"
    dept_label   = dept_choice if dept_choice != "All departments" else "All Departments"
    subj         = f"Performance Report — {dept_label} — {period_label}"

    try:
        _sb = _get_db_client()
        _uph_q = _sb.table("uph_history").select(
            "emp_id, units, hours_worked, uph, work_date, department"
        ).gte("work_date", from_iso).lte("work_date", to_iso)
        _emp_q = _sb.table("employees").select("emp_id, name, department")
        if tenant_id:
            _uph_q = _uph_q.eq("tenant_id", tenant_id)
            _emp_q = _emp_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq3
            _uph_q = _tq3(_uph_q)
            _emp_q = _tq3(_emp_q)
        subs = (_uph_q.execute().data or [])
        emps_lookup = {e["emp_id"]: e for e in (_emp_q.execute().data or [])}
    except Exception:
        subs = []
        emps_lookup = {e["emp_id"]: e for e in (_cached_employees() or [])}
    if dept_choice != "All departments":
        subs = [s for s in subs
                if (s.get("department","") == dept_choice or
                    emps_lookup.get(s.get("emp_id",""),{}).get("department","") == dept_choice)]

    if not subs:
        body = (f"<h2>{dept_label} — {period_label}</h2>"
                f"<p>No work data was found for <strong>{dept_label}</strong> "
                f"between <strong>{from_iso}</strong> and <strong>{to_iso}</strong>.</p>"
                f"<p>Employees may not have had submissions during this period, "
                f"or the date range falls outside your imported data.</p>")
        return None, subj, body

    emp_agg = _dc(lambda: {"units": 0.0, "hours": 0.0, "dept": ""})
    for s in subs:
        eid = s.get("emp_id","")
        emp_agg[eid]["units"] += float(s.get("units") or 0)
        emp_agg[eid]["hours"] += float(s.get("hours_worked") or 0)
        emp_agg[eid]["dept"]   = emps_lookup.get(eid,{}).get("department","")

    try:
        _sb = _get_db_client()
        from datetime import timedelta
        _recent = (d_end - timedelta(days=30)).isoformat()
        _hist_q = _sb.table("uph_history").select("emp_id, uph, work_date").gte("work_date", _recent)
        if tenant_id:
            _hist_q = _hist_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq_hist
            _hist_q = _tq_hist(_hist_q)
        _email_hist = _hist_q.execute().data or []
    except Exception:
        _email_hist = []

    gs_by_emp = {str(r.get("EmployeeID", "")): r for r in (gs or []) if r.get("EmployeeID")}

    def _trend_snapshot(emp_id: str) -> tuple[str, float]:
        if emp_id in gs_by_emp:
            _row = gs_by_emp[emp_id]
            return _row.get("trend", "insufficient_data"), float(_row.get("change_pct", 0) or 0)
        emp_hist = []
        for row in _email_hist:
            if row.get("emp_id") == emp_id:
                try:
                    emp_hist.append((row.get("work_date", ""), float(row.get("uph") or 0)))
                except Exception:
                    pass
        if len(emp_hist) < 4:
            return "insufficient_data", 0.0
        emp_hist.sort(key=lambda item: item[0])
        recent = [v for _, v in emp_hist[-3:]]
        previous = [v for _, v in emp_hist[-6:-3]]
        if not previous:
            return "insufficient_data", 0.0
        prev_avg = sum(previous) / len(previous)
        recent_avg = sum(recent) / len(recent)
        if prev_avg <= 0:
            return "insufficient_data", 0.0
        pct = round(((recent_avg - prev_avg) / prev_avg) * 100, 1)
        if pct >= 5:
            return "up", pct
        if pct <= -5:
            return "down", pct
        return "flat", pct

    scope_gs = []
    for eid, agg in emp_agg.items():
        emp_info = emps_lookup.get(eid, {})
        dept     = agg["dept"]
        uph      = round(agg["units"] / agg["hours"], 2) if agg["hours"] > 0 else 0
        tgt      = float(targets.get(dept, 0) or 0)
        trend, change_pct = _trend_snapshot(eid)
        scope_gs.append({
            "Employee Name": emp_info.get("name", eid),
            "EmployeeID":    eid,
            "Department":    dept,
            "Average UPH":   uph,
            "Total Units":   round(agg["units"]),
            "Hours Worked":  round(agg["hours"], 2),
            "Target UPH":    tgt if tgt else "—",
            "goal_status":   ("on_goal" if tgt and uph >= tgt
                              else "below_goal" if tgt else "no_goal"),
            "trend":         trend,
            "change_pct":    change_pct,
            "flagged":       False,
        })
    scope_gs.sort(key=lambda r: float(r.get("Average UPH", 0) or 0), reverse=True)

    on_g  = sum(1 for r in scope_gs if r["goal_status"] == "on_goal")
    below = sum(1 for r in scope_gs if r["goal_status"] == "below_goal")

    # Calculate risk levels for email (quick version)
    def _email_risk_level(emp, history_all):
        """Quick risk calculation for email."""
        risk_score = 0.0
        trend = emp.get("trend", "insufficient_data")
        if trend == "down":
            risk_score += 4
        elif trend == "flat":
            risk_score += 1
        elif trend == "up":
            risk_score -= 2
        
        # Approximation: use goal_status to estimate streak
        emp_id = emp.get("EmployeeID", "")
        if history_all and emp.get("Target UPH") != "—":
            try:
                target = float(emp.get("Target UPH"))
                emp_hist = [r for r in history_all if r.get("emp_id") == emp_id]
                streak = 0
                for r in reversed(emp_hist[-10:]):  # Check last 10 entries only
                    try:
                        uph = float(r.get("uph") or 0)
                        if uph < target:
                            streak += 1
                        else:
                            break
                    except:
                        break
                if streak >= 3:
                    risk_score += 3
            except:
                pass
        
        if risk_score >= 7:
            return "🔴 High"
        elif risk_score >= 4:
            return "🟡 Medium"
        else:
            return "🟢 Low"

    # Top 3 performers
    _top3 = scope_gs[:3]
    _top3_html = ""
    if _top3:
        _top3_html = "<h3>🏆 Top Performers</h3><ol>"
        for t in _top3:
            _top3_html += f"<li><strong>{t['Employee Name']}</strong> ({t['Department']}) — {t['Average UPH']} UPH</li>"
        _top3_html += "</ol>"

    # Bottom 3 performers (with targets & risk)
    _bottom = [r for r in reversed(scope_gs) if r["goal_status"] == "below_goal"][:3]
    _bottom_html = ""
    if _bottom:
        _bottom_html = "<h3>⚠️ Needs Coaching</h3><table style='width:100%; border-collapse: collapse;'>"
        _bottom_html += "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Employee</th><th style='text-align: center; border: 1px solid #ddd;'>Risk</th><th style='text-align: right; border: 1px solid #ddd;'>UPH</th></tr>"
        for b in _bottom:
            _risk = _email_risk_level(b, _email_hist)
            _tgt = b.get('Target UPH', '—')
            _diff = ""
            try:
                _diff = f" (vs {_tgt})"
            except (ValueError, TypeError):
                pass
            _bottom_html += f"<tr><td style='padding: 8px; border: 1px solid #ddd;'><strong>{b['Employee Name']}</strong><br/><span style='font-size: 0.9em; color: #666;'>{b['Department']}</span></td><td style='text-align: center; border: 1px solid #ddd;'>{_risk}</td><td style='text-align: right; border: 1px solid #ddd;'>{b['Average UPH']}{_diff}</td></tr>"
        _bottom_html += "</table>"

    # ACTION SECTION 1: TODAY's critical attention (🔴 High risk)
    _critical_html = ""
    _critical_list = []
    for r in scope_gs:
        if r["goal_status"] == "below_goal":
            _risk = _email_risk_level(r, _email_hist)
            if "🔴" in _risk:
                _critical_list.append(r)
    
    if _critical_list:
        _critical_html = "<h3 style='color: #8b0000;'>🔴 PRIORITY: Needs Attention TODAY</h3>"
        _critical_html += "<div style='background: #ffe6e6; border-left: 4px solid #8b0000; padding: 12px; margin-bottom: 16px;'>"
        for c in _critical_list[:3]:  # Top 3 critical
            _crit_risk = _email_risk_level(c, _email_hist)
            _action = ""
            try:
                _cur = float(c.get("Average UPH") or 0)
                _tgt = float(c.get("Target UPH") or 0)
                _gap = _tgt - _cur
                if _gap > 5:
                    _action = "👉 <strong>Action:</strong> Schedule 1-on-1. Discuss specific goals & support needed."
                else:
                    _action = "👉 <strong>Action:</strong> Check in on any blockers or issues."
            except:
                _action = "👉 <strong>Action:</strong> Check in immediately."
            
            _critical_html += f"<div style='margin-bottom: 12px;'><strong>{c['Employee Name']}</strong> ({c['Department']}) — {c['Average UPH']} UPH<br/>{_action}</div>"
        _critical_html += "</div>"

    # ACTION SECTION 2: Who improved this week (trend = up)
    _improved_html = ""
    _improved_list = [r for r in scope_gs if r.get("trend") == "up" and r.get("change_pct", 0) > 0]
    if _improved_list:
        _improved_html = "<h3 style='color: #008000;'>🟢 Recognition: Who Improved This Week</h3>"
        _improved_html += "<div style='background: #e6ffe6; border-left: 4px solid #008000; padding: 12px; margin-bottom: 16px;'>"
        for imp in _improved_list[:3]:  # Top 3 improvers
            _pct = imp.get("change_pct", 0)
            _improved_html += f"<div style='margin-bottom: 8px;'><strong>✓ {imp['Employee Name']}</strong> ({imp['Department']}) — <span style='color: green;'>+{_pct:.1f}%</span> trending up<br/>👉 <strong>Action:</strong> Recognize this improvement. Ask what's working & how to sustain it.</div>"
        _improved_html += "</div>"

    # ACTION SECTION 3: Who is trending down (early warning)
    _trending_down_html = ""
    _trending_down = [r for r in scope_gs if r.get("trend") == "down" and r.get("goal_status") != "no_goal"]
    if _trending_down:
        _trending_down_html = "<h3 style='color: #ff6600;'>⚠️ Early Warning: Trending Down</h3>"
        _trending_down_html += "<div style='background: #fff9e6; border-left: 4px solid #ff6600; padding: 12px; margin-bottom: 16px;'>"
        _trending_down_html += f"<p><strong>{len(_trending_down)} employee(s) showing declining performance:</strong></p>"
        for td in _trending_down[:5]:  # Show up to 5
            _pct = td.get("change_pct", 0)
            _status = "at risk" if td.get("goal_status") == "below_goal" else "still on track"
            _trending_down_html += f"<div style='margin-bottom: 8px;'><strong>• {td['Employee Name']}</strong> ({td['Department']}) — {_pct:.1f}% down ({_status})<br/>👉 <strong>Action:</strong> Proactive check-in. Address trend early before it worsens.</div>"
        _trending_down_html += "</div>"

    # Average UPH across all employees
    _avg_all = round(sum(r["Average UPH"] for r in scope_gs) / len(scope_gs), 1) if scope_gs else 0

    _dept_health_rows = []
    for _dept in sorted({r.get("Department", "") for r in scope_gs if r.get("Department")}):
        _dept_rows = [r for r in scope_gs if r.get("Department") == _dept]
        _dept_on = sum(1 for r in _dept_rows if r.get("goal_status") == "on_goal")
        _dept_total = len(_dept_rows)
        _dept_pct = round((_dept_on / _dept_total) * 100) if _dept_total else 0
        _dept_health_rows.append(
            f"<tr><td style='padding:8px;border:1px solid #ddd;'><strong>{_dept}</strong></td>"
            f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_dept_on}/{_dept_total}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_dept_pct}%</td></tr>"
        )
    _dept_health_html = ""
    if _dept_health_rows:
        _dept_health_html = (
            "<h3>📊 Department Health</h3>"
            "<table style='width:100%; border-collapse: collapse;'>"
            "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Department</th>"
            "<th style='text-align: center; border: 1px solid #ddd;'>On Goal</th>"
            "<th style='text-align: center; border: 1px solid #ddd;'>Health</th></tr>"
            + "".join(_dept_health_rows)
            + "</table>"
        )

    _top_risks_html = ""
    if plan_name in ("pro", "business"):
        _risk_rows = []
        for _row in [r for r in scope_gs if r.get("goal_status") == "below_goal"][:5]:
            _risk_rows.append(
                f"<tr><td style='padding:8px;border:1px solid #ddd;'><strong>{_row['Employee Name']}</strong></td>"
                f"<td style='padding:8px;border:1px solid #ddd;'>{_row['Department']}</td>"
                f"<td style='padding:8px;border:1px solid #ddd;text-align:right;'>{_row['Average UPH']}</td>"
                f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_email_risk_level(_row, _email_hist)}</td></tr>"
            )
        if _risk_rows:
            _top_risks_html = (
                "<h3>🔴 Top Risks — Action Required Today</h3>"
                "<table style='width:100%; border-collapse: collapse;'>"
                "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Employee</th>"
                "<th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Dept</th>"
                "<th style='padding: 8px; text-align: right; border: 1px solid #ddd;'>UPH</th>"
                "<th style='padding: 8px; text-align: center; border: 1px solid #ddd;'>Risk</th></tr>"
                + "".join(_risk_rows)
                + "</table>"
            )

    _cost_html = ""
    if plan_name in ("pro", "business"):
        try:
            from settings import Settings as _ImpactS
            _impact_settings = _ImpactS(tenant_id=tenant_id) if tenant_id else _ImpactS()
            _avg_wage = float(_impact_settings.get("avg_hourly_wage", 18.0))
        except Exception:
            _avg_wage = 18.0
        _cost_total = 0.0
        for _row in scope_gs:
            try:
                _target = float(_row.get("Target UPH") or 0)
                _avg = float(_row.get("Average UPH") or 0)
                _hours = float(_row.get("Hours Worked") or 0)
                if _target > 0 and 0 < _avg < _target:
                    _cost_total += max(((_target - _avg) / _target) * _hours * _avg_wage, 0)
            except Exception:
                pass
        _cost_html = f"<h3>💰 Cost Impact</h3><p>Estimated labor cost impact for below-goal performance in this period: <strong>${_cost_total:,.0f}</strong></p>"

    body  = (f"<h2>{dept_label} — {period_label}</h2>"
             f"<p><strong>{len(scope_gs)}</strong> employees · "
             f"Avg UPH: <strong>{_avg_all}</strong> · "
             f"<strong style='color:green'>{on_g} on goal</strong> · "
             f"<strong style='color:red'>{below} below goal</strong></p>"
             f"<hr/>"
             f"{_dept_health_html}"
             f"{_top3_html}"
             f"{_top_risks_html}"
             f"{_critical_html}"
             f"{_improved_html}"
             f"{_trending_down_html}"
             f"{_bottom_html}"
             f"{_cost_html}"
             f"<hr/>"
             f"<p style='font-size: 0.9em; color: #666;'>See the attached Excel report for full details and department breakdown.</p>")

    xl_data = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        HDR_FILL  = PatternFill("solid", fgColor="FF0F2D52")
        HDR_FONT  = Font(bold=True, color="FFFFFFFF", size=10, name="Arial")
        GRN_FILL  = PatternFill("solid", fgColor="FFD4EDDA")
        RED_FILL  = PatternFill("solid", fgColor="FFF8D7DA")
        DARK_FONT = Font(color="FF1A2D42", size=10, name="Arial")
        STATUS_LABELS = {"on_goal":"On Goal","below_goal":"Below Goal","no_goal":"No Target"}

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Summary"
        ws["A1"] = f"{dept_label} — {period_label}"
        ws["A1"].font = Font(bold=True, size=13, name="Arial", color="FF0F2D52")
        ws["A2"] = f"Generated: {today.isoformat()}   |   {on_g} on goal · {below} below goal"
        ws["A2"].font = Font(size=10, name="Arial", color="FF5A7A9C")

        hdrs = ["Employee","Department","Total Units","Hours Worked","Avg UPH","Target UPH","Status"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(4, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(scope_gs, 5):
            fill = (GRN_FILL if r["goal_status"] == "on_goal"
                    else RED_FILL if r["goal_status"] == "below_goal" else None)
            for ci, v in enumerate([r["Employee Name"], r["Department"],
                                     r["Total Units"], r["Hours Worked"],
                                     r["Average UPH"], r["Target UPH"],
                                     STATUS_LABELS.get(r["goal_status"],"—")], 1):
                c = ws.cell(ri, ci, v); c.font = DARK_FONT
                c.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
                if fill: c.fill = fill

        if scope_gs:
            chart = BarChart(); chart.type = "bar"; chart.style = 10
            chart.title = f"Avg UPH — {period_label}"
            chart.width = 20; chart.height = max(8, len(scope_gs) * 0.55)
            last_r = 4 + len(scope_gs)
            chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=last_r), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last_r))
            chart.series[0].graphicalProperties.solidFill = "0F2D52"
            ws.add_chart(chart, "I5")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or "")) for c in col), default=8) + 3, 40)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); xl_data = buf.read()
    except Exception:
        pass

    return xl_data, subj, body


def page_email():
    st.title("📧 Email Setup")
    st.caption("Configure who receives reports, when they are sent, and send manual reports.")
    try:
        from email_engine import (save_smtp_config, get_smtp_config, add_recipient,
                                   remove_recipient, get_recipients, add_schedule,
                                   remove_schedule, get_schedules, get_schedules_due_now, send_report_email,
                                   build_dept_email_body, DAY_NAMES,
                                   import_recipients_from_csv)
    except ImportError:
        st.error("Email module not found."); return

    tab_smtp, tab_recip, tab_sched, tab_send = st.tabs([
        "1️⃣ SMTP Setup", "2️⃣ Recipients", "3️⃣ Schedules", "📤 Send Now"
    ])

    # ── SMTP ─────────────────────────────────────────────────────────────────
    with tab_smtp:
        st.subheader("Email server settings")
        st.caption("Pick your email provider below and we'll fill in the server details automatically.")

        cfg = get_smtp_config()

        # Provider quick-select
        providers = {
            "Gmail":            ("smtp.gmail.com",      587),
            "Outlook / Office 365": ("smtp.office365.com", 587),
            "Yahoo":            ("smtp.mail.yahoo.com", 587),
            "Custom":           ("", 587),
        }
        # Detect current provider from saved server
        cur_server = cfg.get("server","")
        detected = next((k for k,v in providers.items() if v[0] == cur_server), "Custom")
        pcols = st.columns(len(providers))
        for i,(pname,pvals) in enumerate(providers.items()):
            active = (detected == pname)
            if pcols[i].button(pname, key=f"prov_{i}",
                               type="primary" if active else "secondary",
                               use_container_width=True):
                st.session_state["smtp_server_override"] = pvals[0]
                st.session_state["smtp_port_override"]   = pvals[1]
                st.rerun()

        # App password help
        with st.expander("❓ How to get an App Password"):
            st.markdown("""
#### Gmail

**Before you start:** You must have 2-Step Verification turned on — App Passwords won't appear without it.

1. Go to **myaccount.google.com**
2. Click **Security** in the left sidebar
3. Under *"How you sign in to Google"*, click **2-Step Verification** and turn it on if it's off
4. Go back to the Security page and type **"App passwords"** in the search bar at the top — it won't appear in the menu, search is the only way to find it
5. Click **App passwords** in the results
6. Type a name like **Productivity Planner** and click **Create**
7. Google shows you a **16-character password** — copy it immediately, you won't see it again
8. Paste it into the **App password** field below

> ⚠️ If App Passwords doesn't appear even after enabling 2-Step Verification, your account may be managed by a Google Workspace admin who needs to enable it.

---

#### Outlook / Office 365

**Before you start:** You must have 2-Step Verification turned on for your Microsoft account.

1. Go to **account.microsoft.com** and sign in
2. Click **Security** at the top of the page
3. Click **Advanced security options** (under "Security basics")
4. Scroll down to the *"App passwords"* section
5. Click **Create a new app password**
6. Microsoft shows you a password — copy it immediately, you won't see it again
7. Paste it into the **App password** field below

> ⚠️ If you don't see the App passwords section, your organization's admin may need to enable it. Work and school accounts managed by an IT department may require admin approval first.

---

#### Yahoo

**Before you start:** You must have 2-Step Verification turned on for your Yahoo account.

1. Go to **login.yahoo.com** and sign in
2. Click your **profile icon** (top-right) and select **Account Info**
3. Click **Account Security** in the left sidebar
4. Make sure **Two-step verification** is turned on — if it's off, click it and follow the prompts to enable it
5. Scroll down and click **Generate app password** (or "Manage app passwords")
6. Select **Other app** from the dropdown and type **Productivity Planner**
7. Click **Generate**
8. Yahoo shows you a **16-character password** — copy it immediately, you won't see it again
9. Paste it into the **App password** field below

> ⚠️ If Generate app password doesn't appear, make sure Two-step verification is fully enabled and you've refreshed the page.

---

#### Custom SMTP Server

If your email provider isn't listed above, select **Custom** and fill in the Advanced server settings:

1. Ask your email provider or IT team for the **SMTP server address** (e.g. `mail.yourcompany.com`)
2. Ask for the **SMTP port** — usually **587** (TLS) or **465** (SSL)
3. Enter your **full email address** in the "Your email address" field
4. Enter your **email password** or app-specific password in the "App password" field
5. Open **Advanced server settings**, enter the server address and port, and make sure **Use TLS encryption** is checked (recommended)
6. Click **Save settings** and then **Send test email to myself** to verify

> ⚠️ Some corporate servers require a VPN connection or only allow sending from within the company network.

---

You **cannot** use your regular login password for Gmail, Outlook, or Yahoo — they block it for third-party apps.
App passwords are typically 16 characters and look like: **abcd efgh ijkl mnop**
            """)

        # Use override if provider button was just pressed
        server_val = st.session_state.get("smtp_server_override", cfg.get("server",""))
        port_val   = st.session_state.get("smtp_port_override",   int(cfg.get("port",587)))

        with st.form("smtp_form"):
            username = st.text_input("Your email address", value=cfg.get("username",""),
                                      placeholder="you@gmail.com")
            password = st.text_input("App password", value=cfg.get("password",""),
                                      type="password", placeholder="16-character app password")
            # Advanced — collapsed by default
            with st.expander("Advanced server settings"):
                c1, c2 = st.columns(2)
                server  = c1.text_input("SMTP server", value=server_val, placeholder="smtp.gmail.com")
                port    = c2.number_input("Port", value=port_val, min_value=1, max_value=65535)
                use_tls = st.checkbox("Use TLS encryption (recommended)", value=bool(cfg.get("use_tls",True)))

            if st.form_submit_button("Save settings", type="primary", use_container_width=True):
                save_smtp_config(server, port, username, password, username, use_tls)
                # Clear overrides now that they are saved
                st.session_state.pop("smtp_server_override", None)
                st.session_state.pop("smtp_port_override", None)
                st.success("✓ Settings saved.")
                st.rerun()

        if st.button("Send test email to myself", use_container_width=True):
            cfg2 = get_smtp_config()
            if not cfg2.get("username"):
                st.warning("Save your email address first.")
            else:
                with st.spinner("Sending test email to yourself…"):
                    ok, err = send_report_email(
                        [cfg2["username"]],
                        "Productivity Planner — Test Email",
                        "<p>Your email configuration is working correctly! 🎉</p>",
                    )
                if ok:
                    st.success(f"✓ Test email sent to {cfg2['username']}")
                    st.caption("💡 If you don't see it in 1-2 minutes, check your spam folder.")
                else:
                    st.error(f"❌ Send failed")
                    # Parse error to provide actionable help
                    err_lower = str(err).lower()
                    if "authentication" in err_lower or "535" in err_lower:
                        st.warning("**Incorrect email or app password.** Review your SMTP settings and make sure you're using an app-specific password (not your email password).")
                    elif "timeout" in err_lower or "connection" in err_lower or "refused" in err_lower:
                        st.warning("**Connection failed.** Check that the server and port are correct. Try port 465 (SSL) in Advanced settings instead.")
                    elif "certificate" in err_lower or "tls" in err_lower:
                        st.warning("**Encryption error.** Try unchecking 'Use TLS encryption' in Advanced settings, or switch to port 465 (SSL).")
                    else:
                        st.caption(f"Technical details: {err}")

    # ── Recipients ────────────────────────────────────────────────────────────
    with tab_recip:
        st.subheader("Who receives reports")

        recipients = get_recipients()
        if recipients:
            # Show each recipient with an inline remove button
            for r in recipients:
                rc1, rc2, rc3 = st.columns([3, 4, 1])
                rc1.markdown(f"**{r['name']}**")
                rc2.caption(r["email"])
                if rc3.button("✕", key=f"rm_recip_{r['email']}", help="Remove"):
                    remove_recipient(r["email"])
                    _success_then_rerun(f"✓ {r['email']} removed.")
            st.divider()
        else:
            st.info("No recipients yet.")

        st.subheader("Add recipients")
        st.caption("Add one or more at a time. Press **+ Add row** then **Save all** when done.")

        # Dynamic multi-row add form
        if "new_recips" not in st.session_state:
            st.session_state.new_recips = [{"name": "", "email": ""}]

        for ni, nr in enumerate(st.session_state.new_recips):
            nr1, nr2, nr3 = st.columns([3, 4, 1])
            nr["name"]  = nr1.text_input("Name",  value=nr["name"],  key=f"nrn_{ni}",
                                          placeholder="Jane Smith",    label_visibility="visible" if ni == 0 else "collapsed")
            nr["email"] = nr2.text_input("Email", value=nr["email"], key=f"nre_{ni}",
                                          placeholder="jane@acme.com", label_visibility="visible" if ni == 0 else "collapsed")
            if nr3.button("✕", key=f"nrr_{ni}") and len(st.session_state.new_recips) > 1:
                st.session_state.new_recips.pop(ni); st.rerun()

        ba1, ba2 = st.columns(2)
        if ba1.button("+ Add row"):
            st.session_state.new_recips.append({"name": "", "email": ""}); st.rerun()

        if ba2.button("💾 Save all", type="primary", use_container_width=True):
            saved = 0
            for nr in st.session_state.new_recips:
                if nr["name"].strip() and nr["email"].strip():
                    add_recipient(nr["name"].strip(), nr["email"].strip(), [])
                    saved += 1
            if saved:
                st.session_state.new_recips = [{"name": "", "email": ""}]
                _success_then_rerun(f"✓ {saved} recipient(s) added.")
            else:
                st.warning("Enter at least one name and email before saving.")

    # ── Schedules ─────────────────────────────────────────────────────────────
    with tab_sched:
        if not _plan_gate("pro", "Automated Schedules"):
            return
        st.subheader("Automated send schedules")
        from settings import Settings as _SchedTzS
        _sched_settings = _SchedTzS()
        _sched_tz = _sched_settings.get("timezone", "")
        _tz_options = _email_timezone_options()
        _tz_display = ["(Select a timezone)"] + _tz_options
        _tz_idx = (_tz_options.index(_sched_tz) + 1) if _sched_tz in _tz_options else 0

        st.markdown("##### Timezone")
        _tz_col1, _tz_col2 = st.columns([3, 1])
        _sched_tz_choice = _tz_col1.selectbox(
            "Timezone used for schedules",
            _tz_display,
            index=_tz_idx,
            key="sched_timezone_select",
        )
        if _tz_col2.button("Save timezone", key="save_sched_tz", type="primary", use_container_width=True):
            _new_tz = "" if _sched_tz_choice.startswith("(") else _sched_tz_choice
            _sched_settings.set("timezone", _new_tz)
            st.success(f"✓ Timezone set to '{_new_tz or 'server local time'}'.")
            st.rerun()
        if _sched_tz:
            st.caption(f"📍 Current timezone: {_sched_tz}")
        else:
            st.warning("Select and save a timezone before relying on automated schedules.")

        _dbg_now = datetime.now()
        if _sched_tz:
            try:
                from zoneinfo import ZoneInfo
                _dbg_now = datetime.now(ZoneInfo(_sched_tz))
            except Exception:
                pass
        _dbg_day = _dbg_now.strftime("%A")
        _dbg_time = _dbg_now.strftime("%H:%M")
        st.info(f"Scheduler local time: {_dbg_now.strftime('%Y-%m-%d %H:%M:%S %Z')} ({_dbg_day})")

        _debug_cols = st.columns(2)
        if _debug_cols[0].button("Run due schedules now", key="run_due_sched_now", use_container_width=True):
            _results = _run_scheduled_reports_for_tenant(force_now=False)
            if _results:
                _ok_count = sum(1 for r in _results if r.get("ok"))
                st.success(f"Ran {_ok_count}/{len(_results)} due schedule(s).")
                for _res in _results:
                    if _res.get("ok"):
                        st.write(f"✓ {_res.get('name')} → {', '.join(_res.get('recipients', []))}")
                    else:
                        st.write(f"✕ {_res.get('name')} → {_res.get('error')}")
            else:
                st.info("No schedules are due right now.")
        if _debug_cols[1].button("Send all active schedules now", key="force_all_sched_now", use_container_width=True):
            _results = _run_scheduled_reports_for_tenant(force_now=True)
            if _results:
                _ok_count = sum(1 for r in _results if r.get("ok"))
                st.success(f"Sent {_ok_count}/{len(_results)} active schedule(s).")
                for _res in _results:
                    if _res.get("ok"):
                        st.write(f"✓ {_res.get('name')} → {', '.join(_res.get('recipients', []))}")
                    else:
                        st.write(f"✕ {_res.get('name')} → {_res.get('error')}")
            else:
                st.info("No active schedules found to send.")

        _all_recips = get_recipients()
        _recip_labels = [f"{r['name']} <{r['email']}>" for r in _all_recips]
        _recip_emails = {f"{r['name']} <{r['email']}>": r["email"] for r in _all_recips}

        schedules = get_schedules()
        if schedules:
            for s in schedules:
                _last = s.get("last_sent","Never")
                _sper = s.get("report_period", "Prior day")
                if _sper == "Custom":
                    _ds = s.get("date_start", "?")
                    _de = s.get("date_end", "?")
                    _period_lbl = _ds if _ds == _de else f"{_ds} → {_de}"
                else:
                    _period_lbl = _sper
                with st.expander(f"**{s['name']}** · {', '.join(s.get('days',[]))} at {s.get('send_time','')} · {_period_lbl} · Last sent: {_last}"):
                    # Show assigned recipients
                    assigned = s.get("recipients", [])
                    if assigned:
                        st.caption(f"Recipients: {', '.join(assigned)}")
                    else:
                        if not _all_recips:
                            st.warning("⚠️ No recipients configured. Add them in the Recipients tab before this schedule will send.")
                        else:
                            st.caption("No recipients assigned — all recipients will receive this.")
                    # Edit recipients inline
                    cur_labels = [lbl for lbl, em in _recip_emails.items() if em in assigned]
                    new_labels = st.multiselect("Recipients for this schedule",
                                                _recip_labels, default=cur_labels,
                                                key=f"sched_recips_{s['name']}")
                    sc1, sc2, sc3 = st.columns(3)
                    if sc1.button("💾 Save recipients", key=f"sched_save_{s['name']}", use_container_width=True):
                        new_emails = [_recip_emails[lbl] for lbl in new_labels if lbl in _recip_emails]
                        from email_engine import update_schedule_recipients
                        update_schedule_recipients(s["name"], new_emails)
                        st.toast("✓ Recipients updated", icon="✅")
                        st.rerun()
                    if sc2.button("▶ Test now", key=f"sched_test_{s['name']}", use_container_width=True):
                        with st.spinner("Building and sending test report…"):
                            _results = _run_scheduled_reports_for_tenant(force_now=True, schedule_names=[s["name"]])
                        if _results and _results[0].get("ok"):
                            st.success(f"Test sent to {', '.join(_results[0].get('recipients', []))}")
                        elif _results:
                            st.error(f"Send failed: {_results[0].get('error')}")
                        else:
                            st.warning("No recipients to send to.")
                    if sc3.button("🗑 Remove", key=f"sched_rm_{s['name']}", type="secondary", use_container_width=True):
                        remove_schedule(s["name"])
                        st.rerun()
            st.divider()

        st.subheader("Create new schedule")
        with st.form("add_sched_form", clear_on_submit=True):
            sname   = st.text_input("Schedule name", placeholder="e.g. Monday Morning Summary")

            # Date range: pick custom dates or a rolling period
            from datetime import timedelta as _sch_td
            _sch_today = date.today()
            speriod_mode = st.radio("Report covers",
                                     ["Custom date range", "Rolling period"],
                                     horizontal=True, key="sched_period_mode")
            if speriod_mode == "Custom date range":
                _sd1, _sd2 = st.columns(2)
                _sch_start = _sd1.date_input("Start date", value=_sch_today - _sch_td(days=1),
                                              key="sched_date_start")
                _sch_end   = _sd2.date_input("End date (same for single day)",
                                              value=_sch_start, key="sched_date_end")
                speriod = "Custom"
            else:
                speriod = st.selectbox("Report period",
                                        ["Prior day", "Prior week", "Prior month", "Current week"])
                _sch_start = _sch_end = None

            srecips = st.multiselect("Recipients", _recip_labels)
            if not _recip_labels:
                st.warning("⚠️ **No recipients available.** Add recipients in the Recipients tab before creating this schedule.")
            else:
                st.caption("If none are selected, this schedule will use all global recipients.")
            
            st.caption("**Send on these days**")
            daily   = st.checkbox("Every day")
            if not daily:
                day_cols = st.columns(7)
                sel_days = [d for i,d in enumerate(DAY_NAMES)
                            if day_cols[i].checkbox(d[:3], key=f"sd_{d}")]
            else:
                sel_days = ["Daily"]
            send_time = st.time_input("Send time",
                                      value=datetime.strptime("08:00","%H:%M").time())
            subj = st.text_input("Email subject (optional)",
                                 placeholder="Weekly Performance Report")
            if st.form_submit_button("Create schedule", type="primary", disabled=not bool(_sched_tz)):
                if not sname.strip():
                    st.warning("Give the schedule a name.")
                elif not _sched_tz:
                    st.warning("Select and save a timezone first.")
                elif not sel_days:
                    st.warning("Select at least one day.")
                elif speriod == "Custom" and _sch_end < _sch_start:
                    st.warning("End date cannot be before start date.")
                else:
                    sel_emails = [_recip_emails[lbl] for lbl in srecips if lbl in _recip_emails]
                    _ds_str = _sch_start.isoformat() if _sch_start else ""
                    _de_str = _sch_end.isoformat() if _sch_end else ""
                    add_schedule(sname.strip(), [], sel_days,
                                 send_time.strftime("%H:%M"), subj, speriod,
                                 sel_emails, _ds_str, _de_str)
                    _success_then_rerun(f"✓ Schedule '{sname}' created.")

        st.markdown("##### Schedule timing check")
        _dbg_due = get_schedules_due_now(timezone=_sched_tz)
        if _dbg_due:
            st.success(f"{len(_dbg_due)} schedule(s) are due right now.")
            for _dbg_s in _dbg_due:
                st.write(f"✓ {_dbg_s.get('name')} at {_dbg_s.get('send_time')} on {_dbg_s.get('days')}")
        else:
            _all_scheds = get_schedules()
            if _all_scheds:
                for _dbg_s in _all_scheds:
                    _dbg_st = _dbg_s.get('send_time', '?')
                    _dbg_da = _dbg_s.get('days', [])
                    _dbg_match_day = _dbg_day in _dbg_da or 'Daily' in _dbg_da
                    _dbg_status = 'Due now' if (_dbg_match_day and _dbg_time >= _dbg_st) else (f'Wait until {_dbg_st}' if _dbg_match_day else 'Wrong day')
                    st.write(f"• {_dbg_s.get('name')} → {_dbg_status}")
            else:
                st.info("No schedules configured yet. Create one above.")

        # ── Scheduler log viewer ─────────────────────────────────────────────
        st.divider()
        with st.expander("📄 Scheduler log"):
            st.caption("Background email scheduler activity. Useful for debugging missed sends.")
            try:
                _log_path = _tenant_log_path("dpd_email_scheduler")
                import os as _slos
                if _slos.path.exists(_log_path):
                    _lines = open(_log_path, encoding="utf-8").readlines()
                    _recent = _lines[-30:] if len(_lines) > 30 else _lines
                    st.code("".join(reversed(_recent)), language=None)
                    if st.button("Clear log", key="clear_sched_log", type="secondary"):
                        open(_log_path, "w").close()
                        st.rerun()
                else:
                    st.info("No scheduler log yet — it will appear after the email background thread runs.")
            except Exception as _sle:
                st.warning(f"Could not read log: {_sle}")

    # ── Send Now ──────────────────────────────────────────────────────────────
    with tab_send:
        st.subheader("Send a report now")
        st.caption("Manually send a department performance report without waiting for a schedule.")

        # Load archived data if pipeline hasn't run this session
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            st.info("No productivity data available yet.")
        else:
            gs         = st.session_state.goal_status
            depts      = sorted({r.get("Department","") for r in gs if r.get("Department")})
            recipients = get_recipients()

            if not recipients:
                st.warning("No recipients set up yet. Add them in the Recipients tab.")
            else:
                recip_labels = [f"{r['name']} <{r['email']}>" for r in recipients]
                chosen = st.multiselect("Send to", recip_labels,
                                        default=recip_labels[:1] if recip_labels else [])

                from datetime import timedelta as _sn_td
                _sn_today = date.today()
                dc1, dc2 = st.columns(2)
                _sn_start = dc1.date_input("Start date", value=_sn_today - _sn_td(days=1),
                                            key="send_now_start")
                # Reset end date when start changes so they stay in sync
                _prev_start = st.session_state.get("_sn_prev_start")
                if _prev_start and _sn_start != _prev_start:
                    st.session_state.pop("send_now_end", None)
                st.session_state["_sn_prev_start"] = _sn_start
                _sn_end   = dc2.date_input("End date (same as start for a single day)",
                                            value=_sn_start, key="send_now_end")
                if _sn_end < _sn_start:
                    st.warning("End date cannot be before start date.")

                dept_choice = st.selectbox("Department", ["All departments"] + depts)

                if st.button("Send report now", type="primary", use_container_width=True):
                    if not chosen:
                        st.warning("Select at least one recipient.")
                    elif _sn_end < _sn_start:
                        st.warning("Fix the date range first.")
                    else:
                        to_addrs = [r.split("<")[1].rstrip(">") for r in chosen]
                        with st.spinner("Building report…"):
                            xl_data, subj, body = _build_period_report(
                                _sn_start, _sn_end, dept_choice, depts, gs, _cached_targets())
                        with st.spinner("Sending…"):
                            ok, err = send_report_email(to_addrs, subj, body, xl_data)
                        if ok:
                            if xl_data is None:
                                st.info(f"ℹ️ No data for period — notification sent to {len(to_addrs)} recipient(s).")
                            else:
                                st.success(f"✓ Report sent to {len(to_addrs)} recipient(s).")
                        else:
                            st.error(f"Send failed: {err}")






# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

def page_settings():
    st.title("⚙️ Settings")

    tab_app, tab_sub, tab_access, tab_errors = st.tabs(["App Settings", "💳 Subscription", "🔐 Access Control", "🔴 Error Log"])

    # ── Subscription tab ──────────────────────────────────────────────────
    with tab_sub:
        st.subheader("Your Subscription")
        try:
            from database import (
                get_subscription, get_employee_count, get_employee_limit,
                create_billing_portal_url, get_live_stripe_subscription_status,
                refund_latest_subscription_payment,
            )
            _tid_local = st.session_state.get("tenant_id", "")
            sub = get_subscription(_tid_local)
            if sub:
                _plan_raw   = sub.get("plan", "unknown").lower()
                _plan_label = _plan_raw.capitalize()
                _status     = sub.get("status", "unknown")
                _limit      = sub.get("employee_limit", 0)
                _limit_str  = "Unlimited" if _limit == -1 else str(_limit)
                _emp_count  = get_employee_count(_tid_local)
                _period_end = sub.get("current_period_end", "")

                # ── Current plan banner ───────────────────────────────
                _pc = {"starter": "#6b7280", "pro": "#2563eb", "business": "#7c3aed"}.get(_plan_raw, "#6b7280")
                _renew_str = ""
                if _period_end:
                    try:
                        _pe = datetime.fromisoformat(_period_end.replace("Z", "+00:00"))
                        _renew_str = _pe.strftime("%b %d, %Y")
                    except Exception:
                        pass
                st.markdown(f"""
                <div style="background:{_pc}12;border:2px solid {_pc};border-radius:10px;
                            padding:14px 20px;margin-bottom:4px;display:flex;
                            align-items:center;gap:16px;">
                  <div>
                    <div style="font-size:11px;font-weight:700;text-transform:uppercase;
                                letter-spacing:.07em;color:{_pc};">Your Current Plan</div>
                    <div style="font-size:24px;font-weight:800;color:#111;
                                line-height:1.15;">{_plan_label}</div>
                  </div>
                  <div style="margin-left:auto;text-align:right;line-height:1.6;">
                    <div style="font-size:12px;color:#444;">{_emp_count} / {_limit_str} employees used</div>
                    <div style="font-size:12px;color:#444;">{_status.replace("_"," ").title()}</div>
                    {"<div style='font-size:12px;color:#888;'>Renews " + _renew_str + "</div>" if _renew_str else ""}
                  </div>
                </div>
                """, unsafe_allow_html=True)

                if sub.get("cancel_at_period_end"):
                    st.warning("Your subscription will cancel at the end of the current period.")

                _app_url    = st.context.headers.get("Origin", "http://localhost:8501")
                _return_url = _app_url + "/?portal=return"
                _portal_url = create_billing_portal_url(return_url=_return_url)

                # Plan-targeted deep links so Stripe opens the intended update flow.
                _price_map = {
                    "starter": st.secrets.get("STRIPE_PRICE_STARTER", ""),
                    "pro": st.secrets.get("STRIPE_PRICE_PRO", ""),
                    "business": st.secrets.get("STRIPE_PRICE_BUSINESS", ""),
                }
                _plan_portal_urls = {}
                if _portal_url:
                    for _plan_key in ("starter", "pro", "business"):
                        _target_price = _price_map.get(_plan_key, "")
                        _target_url = create_billing_portal_url(
                            return_url=_return_url,
                            target_price_id=_target_price,
                            flow="subscription_update",
                        )
                        _plan_portal_urls[_plan_key] = _target_url or _portal_url

                if _portal_url:
                    st.link_button("Manage Subscription (billing, cancel, update card)",
                                   _portal_url, use_container_width=True, type="primary")
                else:
                    st.info("Billing portal not available. Contact support.")

                with st.expander("Live Stripe verification", expanded=True):
                    _live = get_live_stripe_subscription_status(_tid_local)
                    if not _live:
                        st.info("No live Stripe subscription found yet.")
                    elif _live.get("error"):
                        st.error(_live.get("error"))
                    else:
                        _stripe_period = ""
                        try:
                            if _live.get("current_period_end"):
                                _stripe_period = datetime.fromtimestamp(int(_live.get("current_period_end"))).strftime("%b %d, %Y")
                        except Exception:
                            _stripe_period = ""
                        st.write(f"Effective access now: {_live.get('db_plan', _plan_raw).capitalize()} ({_live.get('db_status', _status)})")
                        st.write(f"Stripe live plan: {_live.get('current_plan', '').capitalize()} | Stripe status: {_live.get('status', '')}")
                        if _stripe_period:
                            st.write(f"Current billing period ends: {_stripe_period}")
                        if _live.get("has_pending_update") and _live.get("pending_plan"):
                            st.success(
                                f"Pending change detected: the app should keep {_live.get('db_plan', _plan_raw).capitalize()} access until {_stripe_period or 'period end'}, then switch to {_live.get('pending_plan', '').capitalize()}."
                            )
                        else:
                            st.info("No pending Stripe plan change detected right now.")
                        if st.button("Refresh live Stripe status", key="refresh_live_stripe", use_container_width=True):
                            st.rerun()

                with st.expander("Refund most recent subscription payment"):
                    st.caption("Issues a full refund for the latest successful subscription charge in Stripe.")
                    _refund_ok = st.checkbox(
                        "I understand this refunds the latest subscription payment.",
                        key="refund_confirm_latest",
                    )
                    if st.button(
                        "Issue refund",
                        key="issue_latest_refund",
                        type="secondary",
                        use_container_width=True,
                        disabled=not _refund_ok,
                    ):
                        _ok, _msg = refund_latest_subscription_payment(_tid_local)
                        if _ok:
                            st.success(f"Refund issued successfully ({_msg}).")
                        else:
                            st.error(_msg)

                # ── Plan comparison ───────────────────────────────────
                st.markdown("---")
                st.markdown("##### Compare Plans")

                _PORD  = ["starter", "pro", "business"]
                _PINFO = {
                    "starter":  {"label": "Starter",  "price": "$49/mo",  "emp": "Up to 25",  "clr": "#6b7280"},
                    "pro":      {"label": "Pro",       "price": "$149/mo", "emp": "Up to 100", "clr": "#2563eb"},
                    "business": {"label": "Business",  "price": "$299/mo", "emp": "Unlimited", "clr": "#7c3aed"},
                }
                _FEATS = {
                    "starter":  ["CSV upload & auto-detection", "Dashboard & rankings",
                                 "Dept-level UPH tracking", "Weekly email reports", "Excel & PDF exports"],
                    "pro":      ["Everything in Starter", "Goal setting & UPH targets",
                                 "Employee trend analysis", "Underperformer flagging & alerts",
                                 "Scheduled automated reports", "Custom date range reports",
                                 "Coaching notes per employee"],
                    "business": ["Everything in Pro", "Order & client tracking",
                                 "Submission plans & progress", "Client trend recording",
                                 "Multi-department management", "Priority email support"],
                }
                # What you GAIN when moving from key[0] → key[1]
                _GAINS = {
                    ("starter", "pro"):      ["75 more employee slots (25 → 100)", "Goal setting & UPH targets",
                                              "Employee trend analysis", "Underperformer alerts",
                                              "Scheduled automated reports", "Custom date ranges",
                                              "Coaching notes per employee"],
                    ("pro", "business"):     ["Unlimited employees (100 → ∞)", "Order & client tracking",
                                              "Submission plans & progress", "Client trend recording",
                                              "Multi-department management", "Priority email support"],
                    ("starter", "business"): ["Unlimited employees", "Goal setting & UPH targets",
                                              "Employee trend analysis", "Underperformer alerts",
                                              "Scheduled reports & custom date ranges",
                                              "Coaching notes", "Order & client tracking",
                                              "Submission plans", "Priority email support"],
                    ("pro", "starter"):      ["75 employee slots (100 → 25)", "Goal setting & UPH targets",
                                              "Employee trend analysis", "Underperformer alerts",
                                              "Scheduled reports", "Custom date ranges",
                                              "Coaching notes per employee"],
                    ("business", "pro"):     ["Unlimited employees (capped at 100)", "Order & client tracking",
                                              "Submission plans & progress", "Client trend recording",
                                              "Multi-department management", "Priority email support"],
                    ("business", "starter"): ["Unlimited employees", "Goal setting & UPH targets",
                                              "Employee trend analysis", "Underperformer alerts",
                                              "Scheduled reports", "Coaching notes",
                                              "Order & client tracking", "Submission plans",
                                              "Priority email support"],
                }

                _cur_idx = _PORD.index(_plan_raw) if _plan_raw in _PORD else -1
                _pcols   = st.columns(3)
                for _ci, _pk in enumerate(_PORD):
                    _pi    = _PINFO[_pk]
                    _is_cur = _pk == _plan_raw
                    _is_up  = _ci > _cur_idx >= 0
                    _is_dn  = _ci < _cur_idx
                    with _pcols[_ci]:
                        if _is_cur:
                            _badge_html = f"<div style='color:{_pc};font-size:11px;font-weight:700;text-transform:uppercase;'>✓ Your Plan</div>"
                        elif _is_up:
                            _badge_html = "<div style='color:#16a34a;font-size:11px;font-weight:700;text-transform:uppercase;'>↑ Upgrade</div>"
                        else:
                            _badge_html = "<div style='color:#dc2626;font-size:11px;font-weight:700;text-transform:uppercase;'>↓ Downgrade</div>"
                        st.markdown(_badge_html, unsafe_allow_html=True)
                        st.markdown(f"**{_pi['label']}** &nbsp; {_pi['price']}")
                        st.caption(_pi['emp'] + " employees")

                        if _is_cur:
                            for _f in _FEATS.get(_pk, []):
                                st.markdown(f"<div style='font-size:12px;color:#555;line-height:1.8;'>✓ {_f}</div>",
                                            unsafe_allow_html=True)
                            st.markdown("")
                            st.button("Current Plan", disabled=True,
                                      use_container_width=True, key=f"sub_btn_{_pk}")
                        elif _is_up:
                            _delta = _GAINS.get((_plan_raw, _pk), [])
                            if _delta:
                                st.markdown("<div style='font-size:12px;font-weight:600;margin-top:4px;'>You'd gain:</div>",
                                            unsafe_allow_html=True)
                                for _g in _delta:
                                    st.markdown(f"<div style='font-size:12px;color:#16a34a;line-height:1.8;'>+ {_g}</div>",
                                                unsafe_allow_html=True)
                            st.markdown("")
                            _target_url = _plan_portal_urls.get(_pk)
                            if _target_url:
                                st.link_button(f"Upgrade to {_pi['label']} →", _target_url,
                                               use_container_width=True, type="primary")
                        else:  # downgrade
                            _delta = _GAINS.get((_pk, _plan_raw), [])
                            if _delta:
                                st.markdown("<div style='font-size:12px;font-weight:600;margin-top:4px;'>You'd lose:</div>",
                                            unsafe_allow_html=True)
                                for _l in _delta:
                                    st.markdown(f"<div style='font-size:12px;color:#dc2626;line-height:1.8;'>− {_l}</div>",
                                                unsafe_allow_html=True)
                            st.markdown("")
                            _target_url = _plan_portal_urls.get(_pk)
                            if _target_url:
                                st.link_button(f"Downgrade to {_pi['label']}", _target_url,
                                               use_container_width=True)
            else:
                st.info("No active subscription found.")
                _app_url = st.context.headers.get("Origin", "http://localhost:8501")
                try:
                    _portal_url = create_billing_portal_url(return_url=_app_url + "/?portal=return")
                    if _portal_url:
                        st.link_button("Manage Subscription", _portal_url,
                                       use_container_width=True, type="primary")
                except Exception:
                    pass
        except Exception as _sub_err:
            st.error(f"Could not load subscription info: {_sub_err}")

    st.caption("Productivity Planner · Powered by Supply Chain Automation Co")

    with tab_app:
        st.session_state.chart_months = st.slider("History window used across charts (months)", 0, 60, st.session_state.chart_months)
        st.caption("This limits how many months of historical data are included in dashboard/productivity trend charts.")
        st.session_state.smart_merge  = True   # always on
        from settings import Settings as _AppSettings
        _tzs = _AppSettings()

        st.divider()
        st.subheader(" Labor Cost Settings")
        st.caption("Used to calculate the financial impact of performance gaps.")
        _cur_wage = _tzs.get("avg_hourly_wage", 18.0)
        _wage_input = st.number_input(
            "Average hourly wage ($)",
            min_value=0.0,
            value=_cur_wage,
            step=1.0,
            key="settings_hourly_wage",
            help="Used to calculate labor cost impact across all reports and dashboards"
        )
        if st.button("Save wage settings", key="save_wage"):
            _tzs.set("avg_hourly_wage", float(_wage_input))
            st.success(f"✓ Average hourly wage set to ${_wage_input:.2f}")

        st.divider()
        st.info("Shift comparison and attendance tracking — coming in a future release.")

        st.divider()
        st.subheader("📋 Audit log")
        st.caption("Recent changes to goals, flags, and journal entries.")
        if st.button("View recent activity", key="view_audit"):
            try:
                import os as _os
                log_path = _tenant_log_path("dpd_audit")
                if _os.path.exists(log_path):
                    lines = open(log_path, encoding="utf-8").readlines()
                    recent = lines[-50:] if len(lines) > 50 else lines
                    st.code("".join(reversed(recent)), language=None)
                else:
                    st.info("No audit log yet — changes will appear here after goals or flags are modified.")
            except Exception as _ae:
                st.error(f"Could not read audit log: {_ae}")

        st.divider()
        st.subheader("🧹 Data cleanup")
        st.caption("Remove duplicate UPH history rows created by running the pipeline multiple times on the same data.")
        if "confirm_cleanup" not in st.session_state:
            st.session_state.confirm_cleanup = False
        if not st.session_state.confirm_cleanup:
            if st.button("Remove duplicate UPH history rows", type="secondary"):
                st.session_state.confirm_cleanup = True
                st.rerun()
        else:
            st.warning("⚠️ This will permanently delete duplicate rows. Are you sure?")
            cc1, cc2 = st.columns(2)
            if cc1.button("Yes, delete duplicates", type="primary", use_container_width=True):
                with st.spinner("Scanning and removing duplicates…"):
                    try:
                        from database import delete_duplicate_uph_history
                        deleted = delete_duplicate_uph_history()
                        _raw_cached_uph_history.clear()
                        st.session_state.confirm_cleanup = False
                        if deleted:
                            st.success(f"✓ Removed {deleted:,} duplicate row(s).")
                        else:
                            st.success("✓ No duplicates found — data is clean.")
                    except BaseException as _ce:
                        st.session_state.confirm_cleanup = False
                        try:    st.error(f"Cleanup error: {repr(_ce)[:200]}")
                        except Exception: st.error("Cleanup failed.")
                        _log_app_error("cleanup", f"Duplicate cleanup failed: {repr(_ce)[:500]}", detail=traceback.format_exc())
            if cc2.button("Cancel", use_container_width=True):
                st.session_state.confirm_cleanup = False
                st.rerun()

    with tab_access:
        st.subheader("Account")
        _uname = st.session_state.get("user_name", "")
        _urole = st.session_state.get("user_role", "member")
        if _uname:
            st.caption(f"Signed in as **{_uname}** · {_urole}")

        # ── Change password ──────────────────────────────────────────────
        st.markdown("**Change password**")
        with st.form("change_pw_form", clear_on_submit=True):
            cur_pw   = st.text_input("Current password", type="password")
            new_pw   = st.text_input("New password", type="password",
                                      placeholder="Min 6 characters")
            conf_pw  = st.text_input("Confirm new password", type="password")
            if st.form_submit_button("Update password", type="primary"):
                if not cur_pw:
                    st.warning("Enter your current password.")
                elif len(new_pw) < 6:
                    st.warning("New password must be at least 6 characters.")
                elif new_pw != conf_pw:
                    st.warning("Passwords don't match.")
                else:
                    try:
                        from database import SUPABASE_URL, SUPABASE_KEY
                        from supabase import create_client as _sc
                        _sb = _sc(SUPABASE_URL, SUPABASE_KEY)
                        sess = st.session_state.get("supabase_session", {})
                        # Re-authenticate with current password to verify identity
                        _sb.auth.sign_in_with_password({
                            "email": st.session_state.get("user_email", _uname),
                            "password": cur_pw,
                        })
                        # Set session and update password
                        _sb.auth.set_session(sess["access_token"], sess["refresh_token"])
                        _sb.auth.update_user({"password": new_pw})
                        st.success("✓ Password updated.")
                    except Exception as _cpe:
                        st.error(f"Failed: {_cpe}")
                        _log_app_error("auth", f"Password change failed: {_cpe}")

        # ── Data export (GDPR) ────────────────────────────────────────────
        st.divider()
        st.markdown("**Your data**")
        st.caption("Download a copy of all your data, or permanently delete your account.")

        if st.button("📥 Export all my data", key="gdpr_export"):
            with st.spinner("Collecting your data…"):
                try:
                    from database import export_all_tenant_data
                    _export = export_all_tenant_data()
                    if _export:
                        _json_bytes = json.dumps(_export, indent=2, default=str).encode("utf-8")
                        st.download_button(
                            "⬇️ Download JSON",
                            data=_json_bytes,
                            file_name="my_data_export.json",
                            mime="application/json",
                            key="gdpr_download",
                        )
                    else:
                        st.info("No data found for your account.")
                except Exception as _gdpr_e:
                    st.error(f"Export failed: {_gdpr_e}")
                    _log_app_error("gdpr", f"Data export failed: {_gdpr_e}", detail=traceback.format_exc())

        # ── Account deletion ─────────────────────────────────────────────
        if "confirm_delete_account" not in st.session_state:
            st.session_state.confirm_delete_account = False

        if not st.session_state.confirm_delete_account:
            if st.button("🗑️ Delete my account and all data", type="secondary", key="gdpr_delete_start"):
                st.session_state.confirm_delete_account = True
                st.rerun()
        else:
            st.error("⚠️ This will **permanently delete** your account, all employees, UPH history, goals, settings, and email config. This cannot be undone.")
            _del_confirm = st.text_input(
                "Type DELETE to confirm", key="gdpr_delete_confirm",
                placeholder="DELETE",
            )
            dc1, dc2 = st.columns(2)
            if dc1.button("Permanently delete everything", type="primary", use_container_width=True, key="gdpr_delete_go"):
                if _del_confirm == "DELETE":
                    with st.spinner("Deleting all data…"):
                        try:
                            from database import delete_all_tenant_data
                            delete_all_tenant_data()
                            _full_sign_out()
                            st.rerun()
                        except Exception as _del_e:
                            st.error(f"Deletion failed: {_del_e}")
                            _log_app_error("gdpr", f"Account deletion failed: {_del_e}", detail=traceback.format_exc())
                else:
                    st.warning("Type DELETE exactly to confirm.")
            if dc2.button("Cancel", use_container_width=True, key="gdpr_delete_cancel"):
                st.session_state.confirm_delete_account = False
                st.rerun()

        # ── Sign out ─────────────────────────────────────────────────────
        st.divider()
        if st.button("Sign out", type="secondary"):
            _full_sign_out()
            st.rerun()

    # ── Error Log tab ─────────────────────────────────────────────────────
    with tab_errors:
        st.subheader("Error Log")
        st.caption("Recent errors and warnings logged by the application. Use this to diagnose issues.")

        # Filters
        fc1, fc2, fc3 = st.columns(3)
        _err_cat = fc1.selectbox("Category", ["All", "login", "pipeline", "email", "employees",
                                               "clients", "productivity", "gdpr", "auth",
                                               "cleanup", "import", "database", "password_reset"],
                                 key="err_cat_filter")
        _err_sev = fc2.selectbox("Severity", ["All", "error", "warning", "info"], key="err_sev_filter")
        _err_limit = fc3.number_input("Show last N", min_value=10, max_value=500, value=50, step=10, key="err_limit")

        try:
            from database import get_error_reports, clear_error_reports
            _cat_arg = "" if _err_cat == "All" else _err_cat
            _sev_arg = "" if _err_sev == "All" else _err_sev
            errors = get_error_reports(limit=_err_limit, category=_cat_arg, severity=_sev_arg)

            if errors:
                st.markdown(f"**{len(errors)}** error(s) found")

                # Summary badges
                _err_count = sum(1 for e in errors if e.get("severity") == "error")
                _warn_count = sum(1 for e in errors if e.get("severity") == "warning")
                _info_count = sum(1 for e in errors if e.get("severity") == "info")
                bc1, bc2, bc3 = st.columns(3)
                bc1.metric("Errors", _err_count)
                bc2.metric("Warnings", _warn_count)
                bc3.metric("Info", _info_count)

                st.divider()

                _SEV_ICON = {"error": "🔴", "warning": "🟡", "info": "🔵"}
                for err in errors:
                    sev = err.get("severity", "error")
                    icon = _SEV_ICON.get(sev, "⚪")
                    cat = err.get("category", "unknown")
                    msg = err.get("message", "")
                    ts = err.get("created_at", "")[:19].replace("T", " ")
                    user = err.get("user_email", "")

                    with st.expander(f"{icon} **[{cat}]** {msg[:120]}{'…' if len(msg) > 120 else ''} — {ts}"):
                        st.markdown(f"**Category:** {cat}")
                        st.markdown(f"**Severity:** {sev}")
                        st.markdown(f"**Time:** {ts}")
                        if user:
                            st.markdown(f"**User:** {user}")
                        st.markdown(f"**Message:** {msg}")
                        detail = err.get("detail", "")
                        if detail:
                            st.markdown("**Detail / Stack trace:**")
                            st.code(detail, language=None)

                st.divider()
                if st.button("🗑️ Clear all error logs", type="secondary", key="clear_errors"):
                    clear_error_reports()
                    st.success("✓ Error log cleared.")
                    st.rerun()
            else:
                st.success("No errors logged. Everything is running smoothly.")
        except Exception as _err_ui_err:
            st.warning(f"Could not load error reports: {_err_ui_err}")
            st.caption("The error_reports table may not exist yet. Run the migration in migrations/001_setup.sql.")


# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════════════

_SUSPICIOUS_NAME_RE = re.compile(r"(<\s*/?\s*script\b|drop\s+table|;--|javascript:)", re.IGNORECASE)


def _normalize_label_text(value, max_len: int = 64) -> str:
    """Normalize UI labels so imported text stays readable and safe-looking."""
    s = str(value or "")
    s = s.replace("\x00", " ")
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("|", " ").replace("<", " ").replace(">", " ")
    s = s.strip(" '\"")
    if len(s) > max_len:
        s = s[: max_len - 3].rstrip() + "..."
    return s or "Unknown"


def _sanitize_employee_name(raw_name, emp_id: str = "") -> tuple[str, bool]:
    """Return a cleaned display name and a flag indicating suspicious input."""
    raw = str(raw_name or "")
    suspicious = bool(_SUSPICIOUS_NAME_RE.search(raw))
    cleaned = _normalize_label_text(raw, max_len=64)
    if suspicious:
        fallback = f"Employee {emp_id}".strip() if emp_id else "Employee"
        return _normalize_label_text(fallback, max_len=64), True
    return cleaned, False

def _parse_csv(raw: bytes) -> tuple[list[dict], list[str]]:
    """Parse CSV bytes → (rows, headers). Delegates to data_loader."""
    hdrs, rows = _dl_parse_csv(raw)
    return rows, hdrs


def _auto_detect(headers: list) -> dict:
    """Auto-detect CSV column mapping. Delegates to data_loader."""
    return _dl_auto_detect(headers)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

# ── Access control ────────────────────────────────────────────────────────────
# Set APP_PASSWORD in Settings page. Leave blank to disable the gate.

def _check_access() -> bool:
    """Return True if the user is authenticated (or no password is set)."""
    pw = st.session_state.get("app_password_set", "")
    if not pw:
        return True   # no password configured — open access
    if st.session_state.get("authenticated"):
        return True   # already logged in this session

    st.markdown("""
    <div style="max-width:400px;margin:80px auto;background:#fff;
                border:1px solid #E2EBF4;border-radius:12px;padding:36px;">
      <div style="font-size:22px;font-weight:700;color:#0F2D52;
                  margin-bottom:24px;">📦 Productivity Planner</div>
    </div>""", unsafe_allow_html=True)

    entered = st.text_input("Password", type="password",
                             placeholder="Enter your access password",
                             key="login_input")
    if st.button("Sign in", type="primary"):
        if entered == pw:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False


_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_LOCKOUT_SECONDS = 900  # 15 minutes


def _check_login_lockout() -> bool:
    """Return True if the user is locked out from too many failed login attempts."""
    attempts = st.session_state.get("_login_attempts", 0)
    lockout_until = st.session_state.get("_login_lockout_until", 0)
    if lockout_until and time.time() < lockout_until:
        remaining = int(lockout_until - time.time())
        mins = remaining // 60
        secs = remaining % 60
        st.error(f"Too many failed attempts. Try again in {mins}m {secs}s.")
        return True
    # Reset lockout if time has passed
    if lockout_until and time.time() >= lockout_until:
        st.session_state["_login_attempts"] = 0
        st.session_state["_login_lockout_until"] = 0
    return False


def _record_failed_login():
    """Increment failed login counter and lock out if threshold reached."""
    attempts = st.session_state.get("_login_attempts", 0) + 1
    st.session_state["_login_attempts"] = attempts
    if attempts >= _LOGIN_MAX_ATTEMPTS:
        st.session_state["_login_lockout_until"] = time.time() + _LOGIN_LOCKOUT_SECONDS
        st.error(f"Account locked for 15 minutes after {_LOGIN_MAX_ATTEMPTS} failed attempts.")
    else:
        remaining = _LOGIN_MAX_ATTEMPTS - attempts
        st.error(f"Invalid email or password. {remaining} attempt(s) remaining.")


def _login_page():
    """Supabase Auth login screen — shown when no valid session exists."""

    st.markdown("""
    <div style="max-width:400px;margin:80px auto 0;text-align:center;">
      <div style="background:#0F2D52;border-radius:12px;padding:32px 36px;">
        <div style="font-size:28px;margin-bottom:4px;">📦</div>
        <div style="font-size:20px;font-weight:700;color:#fff;letter-spacing:-.02em;">
          Productivity Planner
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Password reset mode ───────────────────────────────────────────
    if st.session_state.get("_show_reset_pw"):
        st.markdown("<div style='max-width:400px;margin:24px auto 0;'>", unsafe_allow_html=True)

        # Sub-mode: paste reset link ────────────────────────────────────
        if st.session_state.get("_show_paste_link"):
            # If we already verified, show password form directly
            if st.session_state.get("_recovery_access_token"):
                st.subheader("Set new password")
                new_pw  = st.text_input("New password", type="password", key="paste_new_pw")
                conf_pw = st.text_input("Confirm password", type="password", key="paste_conf_pw")
                if st.button("Update password", type="primary", use_container_width=True, key="paste_update_pw"):
                    if not new_pw or len(new_pw) < 6:
                        st.warning("Password must be at least 6 characters.")
                    elif new_pw != conf_pw:
                        st.warning("Passwords do not match.")
                    else:
                        try:
                            import requests as _req
                            from database import SUPABASE_URL, SUPABASE_KEY
                            _upd_resp = _req.put(
                                f"{SUPABASE_URL}/auth/v1/user",
                                json={"password": new_pw},
                                headers={
                                    "apikey": SUPABASE_KEY,
                                    "Authorization": f"Bearer {st.session_state['_recovery_access_token']}",
                                    "Content-Type": "application/json",
                                },
                                timeout=10,
                            )
                            if _upd_resp.status_code == 200:
                                st.session_state.pop("_recovery_access_token", None)
                                st.session_state.pop("_show_paste_link", None)
                                st.session_state.pop("_show_reset_pw", None)
                                st.success("Password updated! Redirecting to sign in...")
                                time.sleep(2)
                                st.rerun()
                            else:
                                st.error(f"Failed to update password: {_upd_resp.text}")
                        except Exception as _upe:
                            st.error(f"Failed to update password: {_upe}")
                st.markdown("</div>", unsafe_allow_html=True)
                st.stop()

            # Step 1: paste the link
            st.subheader("Paste your reset link")
            st.caption("Open the reset email, copy the full link, and paste it below.")
            _paste_url = st.text_input("Reset link from email", placeholder="Paste the full URL from your email", key="paste_reset_url")
            pc1, pc2 = st.columns(2)
            if pc1.button("Continue", type="primary", use_container_width=True):
                if not _paste_url.strip():
                    st.warning("Paste the link from your reset email.")
                else:
                    from urllib.parse import parse_qs, urlparse
                    _parsed = urlparse(_paste_url.strip())
                    _at, _rt = "", ""
                    _paste_err = ""

                    # Case 1: URL fragment has access_token (redirected URL)
                    if _parsed.fragment:
                        _fp = parse_qs(_parsed.fragment)
                        _at = (_fp.get("access_token") or [""])[0]
                        _rt = (_fp.get("refresh_token") or [""])[0]

                    # Case 2: Supabase verify URL with ?token=...&type=recovery
                    if not _at and _parsed.query:
                        _qp2 = parse_qs(_parsed.query)
                        _verify_token = (_qp2.get("token") or [""])[0]
                        _verify_type  = (_qp2.get("type") or [""])[0]
                        if _verify_token and _verify_type == "recovery":
                            try:
                                import requests as _req
                                from database import SUPABASE_URL, SUPABASE_KEY
                                _vresp = _req.post(
                                    f"{SUPABASE_URL}/auth/v1/verify",
                                    json={"token_hash": _verify_token, "type": "recovery"},
                                    headers={"apikey": SUPABASE_KEY, "Content-Type": "application/json"},
                                    timeout=10,
                                )
                                if _vresp.status_code == 200:
                                    _vdata = _vresp.json()
                                    _at = _vdata.get("access_token", "")
                                    _rt = _vdata.get("refresh_token", "")
                                else:
                                    _paste_err = f"Supabase returned {_vresp.status_code}: {_vresp.text[:150]}"
                            except Exception as _ve:
                                _paste_err = f"Verify failed: {_ve}"
                                _log_app_error("password_reset", f"Verify API error: {_ve}")
                        elif not _verify_token:
                            _paste_err = "No token found in the URL."
                        elif _verify_type != "recovery":
                            _paste_err = f"Wrong link type: '{_verify_type}' (expected 'recovery')."

                    if _at:
                        # Store token and rerun to show password form
                        st.session_state["_recovery_access_token"] = _at
                        st.rerun()
                    else:
                        st.error(_paste_err or "Could not find a reset token in that link.")
            if pc2.button("Back", use_container_width=True):
                st.session_state["_show_paste_link"] = False
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
            return

        # Sub-mode: request reset email ─────────────────────────────────
        st.subheader("Reset your password")
        st.caption("Enter your email address and we'll send you a password reset link.")
        reset_email = st.text_input("Email", placeholder="you@company.com", key="reset_email")
        rc1, rc2 = st.columns(2)
        if rc1.button("Send reset link", type="primary", use_container_width=True):
            if not reset_email.strip():
                st.warning("Enter your email address.")
            else:
                try:
                    from database import SUPABASE_URL, SUPABASE_KEY
                    from supabase import create_client as _sc
                    _sb = _sc(SUPABASE_URL, SUPABASE_KEY)
                    # Pass redirect URL so Supabase sends user back here with PKCE code
                    _redirect = st.context.headers.get("Origin", "http://localhost:8501")
                    _sb.auth.reset_password_email(reset_email.strip(), {
                        "redirect_to": _redirect,
                    })
                    st.success("Reset link sent! Check your inbox (and spam folder).")
                    st.info("After you get the email, click **\"I have a reset link\"** below and paste the link.")
                except Exception as _re:
                    # Don't reveal whether email exists
                    st.success("If that email exists, a reset link has been sent. Check your inbox (and spam folder).")
                    _log_app_error("password_reset", f"Reset request for {reset_email.strip()}: {_re}")
        st.markdown("---")
        rc3, rc4 = st.columns(2)
        if rc3.button("I have a reset link", use_container_width=True):
            st.session_state["_show_paste_link"] = True
            st.rerun()
        if rc4.button("Back to sign in", use_container_width=True):
            st.session_state["_show_reset_pw"] = False
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
        return

    # ── Sign in mode ──────────────────────────────────────────────────
    st.markdown("<div style='max-width:400px;margin:24px auto 0;'>", unsafe_allow_html=True)

    if _check_login_lockout():
        st.markdown("</div>", unsafe_allow_html=True)
        return

    email    = st.text_input("Email",    placeholder="you@company.com",  key="login_email")
    password = st.text_input("Password", placeholder="••••••••••••",      key="login_password", type="password")

    if st.button("Sign in", type="primary", use_container_width=True):
        if not email.strip() or not password.strip():
            st.error("Enter your email and password.")
        else:
            try:
                from database import SUPABASE_URL, SUPABASE_KEY
                from supabase import create_client as _sc
                _sb   = _sc(SUPABASE_URL, SUPABASE_KEY)
                resp  = _sb.auth.sign_in_with_password({"email": email.strip(), "password": password})

                # ── Email verification check ──────────────────────────
                _user = resp.user
                _email_confirmed = getattr(_user, "email_confirmed_at", None)
                if not _email_confirmed:
                    st.warning("Your email has not been verified. Check your inbox for a confirmation link.")
                    _log_app_error("login", f"Unverified email login attempt: {email.strip()}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    return

                _at = resp.session.access_token
                _rt = resp.session.refresh_token
                # Store session tokens + expiry for auto-refresh
                st.session_state["supabase_session"] = {
                    "access_token": _at, "refresh_token": _rt,
                }
                st.session_state["_sb_token_expires_at"] = (
                    resp.session.expires_at
                    if hasattr(resp.session, "expires_at") and resp.session.expires_at
                    else time.time() + 3600
                )
                # Create a fresh client with the auth session baked in
                _sb2 = _sc(SUPABASE_URL, SUPABASE_KEY)
                _sb2.auth.set_session(_at, _rt)

                _uid = resp.user.id
                prof_resp = _sb2.table("user_profiles").select("tenant_id, role, name") \
                               .eq("id", _uid).execute()
                if prof_resp.data:
                    _prof = prof_resp.data[0]
                else:
                    # First login — provision tenant + profile
                    import uuid as _uuid
                    _new_tid = str(_uuid.uuid4())
                    _display = email.strip().split("@")[0]
                    # Use RPC to bypass RLS — falls back to direct insert
                    try:
                        _rpc_result = _sb2.rpc("provision_tenant", {
                            "p_user_id":     _uid,
                            "p_tenant_name": _display,
                            "p_user_name":   _display,
                        }).execute()
                        # RPC returns the generated tenant_id
                        if _rpc_result.data:
                            _new_tid = _rpc_result.data
                    except Exception as _rpc_err:
                        # RPC doesn't exist yet — try direct insert
                        _log_app_error("provision_tenant", f"RPC failed for {_uid}: {_rpc_err}, trying direct insert")
                        _sb2.table("tenants").insert({
                            "id": _new_tid, "name": _display,
                        }).execute()
                        _sb2.table("user_profiles").insert({
                            "id": _uid, "tenant_id": _new_tid,
                            "role": "admin", "name": _display,
                        }).execute()
                    _prof = {"tenant_id": _new_tid, "role": "admin",
                             "name": _display}
                st.session_state["tenant_id"] = _prof["tenant_id"]
                st.session_state["user_role"] = _prof.get("role", "member")
                st.session_state["user_name"]  = _prof.get("name") or email.strip()
                st.session_state["user_email"] = email.strip()
                # Clear login attempt counter on success
                st.session_state["_login_attempts"] = 0
                st.session_state["_login_lockout_until"] = 0
                _bust_cache()
                st.rerun()
            except Exception as _le:
                _record_failed_login()
                _log_app_error("login", f"Failed login for {email.strip()}: {_le}")

    # Forgot password link
    if st.button("Forgot password?", type="secondary", use_container_width=True, key="forgot_pw_btn"):
        st.session_state["_show_reset_pw"] = True
        st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)


_SESSION_TIMEOUT_SECONDS = 3600  # 1 hour idle timeout


def _check_session_timeout():
    """Auto-logout if user has been idle for more than 1 hour."""
    now = time.time()
    last_activity = st.session_state.get("_last_activity", now)
    if now - last_activity > _SESSION_TIMEOUT_SECONDS:
        # Clear session
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        return True  # timed out
    st.session_state["_last_activity"] = now
    return False


def _verify_checkout_and_activate():
    """After Stripe checkout, verify payment and create subscription in DB."""
    import requests as _req
    import database as _db
    _debug = []
    get_client = _db.get_client
    _get_config = _db._get_config
    PLAN_LIMITS = _db.PLAN_LIMITS
    _get_tid = getattr(_db, "get_tenant_id", lambda: st.session_state.get("tenant_id", ""))
    _get_uid = getattr(_db, "get_user_id", lambda: st.session_state.get("user_id", ""))

    stripe_key = _get_config("STRIPE_SECRET_KEY")
    tid = _get_tid()
    uid = _get_uid()
    if not stripe_key:
        _debug.append("no STRIPE_SECRET_KEY")
        st.session_state["_verify_debug"] = _debug
        return False
    if not tid:
        _debug.append("no tenant_id")
        st.session_state["_verify_debug"] = _debug
        return False
    if not uid:
        _debug.append("no user_id in session; continuing tenant-level sync")

    _debug.append(f"tenant_id={tid[:8]}...")

    # Get the tenant's Stripe customer ID
    sb = get_client()
    try:
        t_resp = sb.table("tenants").select("stripe_customer_id").eq("id", tid).execute()
        cust_id = t_resp.data[0].get("stripe_customer_id") if t_resp.data else None
    except Exception as e:
        _debug.append(f"tenant lookup err: {e}")
        st.session_state["_verify_debug"] = _debug
        return False
    if not cust_id:
        _debug.append("no stripe_customer_id on tenant")
        st.session_state["_verify_debug"] = _debug
        return False

    _debug.append(f"stripe_customer={cust_id[:12]}...")

    # List recent subscriptions for this customer and pick the best candidate
    try:
        sub_resp = _req.get(
            "https://api.stripe.com/v1/subscriptions",
            auth=(stripe_key, ""),
            params={"customer": cust_id, "status": "all", "limit": 10},
            timeout=10,
        )
        if sub_resp.status_code != 200:
            _debug.append(f"stripe list subs: {sub_resp.status_code} {sub_resp.text[:100]}")
            st.session_state["_verify_debug"] = _debug
            return False
        subs = sub_resp.json().get("data", [])
        if not subs:
            _debug.append("no subscriptions found in Stripe")
            st.session_state["_verify_debug"] = _debug
            return False

        # Prefer active/trialing/past_due, newest first.
        _preferred = ["active", "trialing", "past_due", "unpaid", "incomplete"]
        subs_sorted = sorted(
            subs,
            key=lambda s: (_preferred.index(s.get("status")) if s.get("status") in _preferred else 999,
                           -(s.get("created") or 0)),
        )
        stripe_sub = subs_sorted[0]
        _debug.append(f"stripe status picked: {stripe_sub.get('status')}")
    except Exception as e:
        _debug.append(f"stripe API err: {e}")
        st.session_state["_verify_debug"] = _debug
        return False

    _debug.append(f"found stripe sub {stripe_sub['id'][:16]}...")

    # Determine plan: try price metadata first, then match price ID against secrets
    plan = ""
    try:
        price_obj  = stripe_sub["items"]["data"][0]["price"]
        price_meta = price_obj.get("metadata", {})
        plan       = price_meta.get("plan", "").lower().strip()
        _debug.append(f"plan from metadata: '{plan}'")
    except Exception as _pe:
        _debug.append(f"price metadata err: {_pe}")

    # Fallback: match against price IDs stored in secrets
    if not plan or plan not in PLAN_LIMITS:
        try:
            _price_id = stripe_sub["items"]["data"][0]["price"]["id"]
            _sec_starter  = _get_config("STRIPE_PRICE_STARTER")  or ""
            _sec_pro      = _get_config("STRIPE_PRICE_PRO")      or ""
            _sec_business = _get_config("STRIPE_PRICE_BUSINESS") or ""
            if _price_id and _sec_business and _price_id == _sec_business:
                plan = "business"
            elif _price_id and _sec_pro and _price_id == _sec_pro:
                plan = "pro"
            elif _price_id and _sec_starter and _price_id == _sec_starter:
                plan = "starter"
            else:
                plan = plan or "starter"  # last resort default
            _debug.append(f"plan from price ID match: '{plan}'")
        except Exception as _pie:
            _debug.append(f"price ID fallback err: {_pie}")
            plan = plan or "starter"

    if not plan:
        plan = "starter"
    limit = PLAN_LIMITS.get(plan, 25)
    _debug.append(f"plan={plan} limit={limit}")

    # Upsert subscription in our DB
    try:
        _existing_user_id = ""
        if not uid:
            try:
                _prev = sb.table("subscriptions").select("user_id").eq("tenant_id", tid).execute()
                if _prev.data:
                    _existing_user_id = _prev.data[0].get("user_id") or ""
            except Exception:
                _existing_user_id = ""

        _sub_row = {
            "tenant_id": tid,
            "stripe_customer_id": cust_id,
            "stripe_subscription_id": stripe_sub["id"],
            "plan": plan,
            "status": stripe_sub.get("status", "active"),
            "employee_limit": limit,
            "cancel_at_period_end": stripe_sub.get("cancel_at_period_end", False),
        }
        _row_user_id = uid or _existing_user_id
        if _row_user_id:
            _sub_row["user_id"] = _row_user_id
        _cpe = stripe_sub.get("current_period_end")
        if _cpe:
            from datetime import datetime, timezone
            _sub_row["current_period_end"] = datetime.fromtimestamp(_cpe, tz=timezone.utc).isoformat()
        sb.table("subscriptions").upsert(_sub_row, on_conflict="tenant_id").execute()
        _debug.append("DB upsert SUCCESS")
        st.session_state["_current_plan"] = plan
        st.session_state["_verify_debug"] = _debug
        return True
    except Exception as e:
        _debug.append(f"DB upsert FAILED: {e}")
        st.session_state["_verify_debug"] = _debug
        return False


def _subscription_page():
    """Subscription gate page — shown when no active subscription is found."""
    from database import (create_stripe_checkout_url, get_subscription,
                          get_employee_count, create_billing_portal_url,
                          get_live_stripe_subscription_status,
                          refund_latest_subscription_payment)

    # ── Shared plan data (mirrors Settings tab) ───────────────────────────
    _PORD  = ["starter", "pro", "business"]
    _PINFO = {
        "starter":  {"label": "Starter",  "price": "$49/mo",  "emp": "Up to 25",  "clr": "#6b7280",
                     "desc": "For small teams getting started."},
        "pro":      {"label": "Pro",       "price": "$149/mo", "emp": "Up to 100", "clr": "#2563eb",
                     "desc": "For growing operations that need deeper insights."},
        "business": {"label": "Business",  "price": "$299/mo", "emp": "Unlimited", "clr": "#7c3aed",
                     "desc": "For large warehouses and multi-site operations."},
    }
    _FEATS = {
        "starter":  ["CSV upload & auto-detection", "Dashboard & rankings",
                     "Dept-level UPH tracking", "Weekly email reports", "Excel & PDF exports"],
        "pro":      ["Everything in Starter", "Goal setting & UPH targets",
                     "Employee trend analysis", "Underperformer flagging & alerts",
                     "Scheduled automated reports", "Custom date range reports",
                     "Coaching notes per employee"],
        "business": ["Everything in Pro", "Order & client tracking",
                     "Submission plans & progress", "Client trend recording",
                     "Multi-department management", "Priority email support"],
    }

    # ── Handle checkout query params ──────────────────────────────────────
    _qp = st.query_params
    if _qp.get("checkout") == "success":
        with st.spinner("Activating your subscription..."):
            activated = _verify_checkout_and_activate()
        if activated:
            st.balloons()
            st.success("Welcome! Your subscription is now active.")
            st.query_params.clear()
            st.session_state.pop("_sub_active", None)
            st.session_state.pop("_checkout_url", None)
            st.session_state.pop("_checkout_plan", None)
            time.sleep(2)
            st.rerun()
        else:
            st.warning("Payment received! It may take a moment to activate. Please refresh in a few seconds.")
            st.query_params.clear()
    if _qp.get("checkout") == "canceled":
        st.info("Checkout canceled — no charge was made. Choose a plan below.")
        st.query_params.clear()

    # ── Check existing subscription ───────────────────────────────────────
    try:
        existing_sub = get_subscription()
    except Exception:
        existing_sub = None

    _app_url    = st.context.headers.get("Origin", "http://localhost:8501")
    _portal_url = None
    _price_map = {}
    try:
        _portal_url = create_billing_portal_url(return_url=_app_url + "/?portal=return")
        _price_map = {
            "starter": st.secrets.get("STRIPE_PRICE_STARTER", ""),
            "pro": st.secrets.get("STRIPE_PRICE_PRO", ""),
            "business": st.secrets.get("STRIPE_PRICE_BUSINESS", ""),
        }
    except Exception:
        pass

    _sub_status  = (existing_sub or {}).get("status", "")
    _sub_plan    = (existing_sub or {}).get("plan", "").lower()
    _period_end  = (existing_sub or {}).get("current_period_end", "")
    _renew_str   = ""
    if _period_end:
        try:
            _pe = datetime.fromisoformat(_period_end.replace("Z", "+00:00"))
            _renew_str = _pe.strftime("%b %d, %Y")
        except Exception:
            pass

    # ── Page header ───────────────────────────────────────────────────────
    st.markdown("""
    <div style="max-width:860px;margin:32px auto 0;text-align:center;">
      <div style="background:#0F2D52;border-radius:12px;padding:28px 36px;margin-bottom:8px;">
        <div style="font-size:26px;margin-bottom:4px;">📦</div>
        <div style="font-size:20px;font-weight:700;color:#fff;letter-spacing:-.02em;">
          Productivity Planner
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Branch on sub status ──────────────────────────────────────────────

    # PAST DUE — fix payment to reactivate
    if _sub_status == "past_due":
        _pi = _PINFO.get(_sub_plan, {})
        st.markdown(f"""
        <div style="background:#fef2f2;border:2px solid #dc2626;border-radius:10px;
                    padding:16px 20px;margin:16px 0 8px;">
          <div style="font-weight:700;color:#dc2626;font-size:15px;">⚠ Payment Past Due</div>
          <div style="color:#555;margin-top:4px;font-size:13px;">
            Your <strong>{_pi.get("label", _sub_plan.capitalize()) if _pi else _sub_plan.capitalize()}</strong> plan
            is on hold because your last payment failed.
            Update your card to restore access immediately.
          </div>
        </div>
        """, unsafe_allow_html=True)
        if _portal_url:
            st.link_button("Update Payment Method →", _portal_url,
                           use_container_width=True, type="primary")
        else:
            st.info("Contact support to update your payment method.")
        st.markdown("---")
        if st.button("Sign out"):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()
        st.stop()

    # CANCELED — show what they had, let them reactivate or pick a new plan
    if _sub_status in ("canceled", "incomplete_expired") and _sub_plan:
        _pi = _PINFO.get(_sub_plan, {})
        _ended = f"  Ended {_renew_str}." if _renew_str else ""
        st.markdown(f"""
        <div style="background:#fffbeb;border:2px solid #d97706;border-radius:10px;
                    padding:16px 20px;margin:16px 0 8px;">
          <div style="font-weight:700;color:#d97706;font-size:15px;">Your plan has ended</div>
          <div style="color:#555;margin-top:4px;font-size:13px;">
            You previously had the <strong>{_pi.get("label", _sub_plan.capitalize()) if _pi else _sub_plan.capitalize()}</strong> plan.{_ended}
            Re-subscribe below to restore access.
          </div>
        </div>
        """, unsafe_allow_html=True)

    # NO SUB / TRIALING / OTHER — generic header
    if not existing_sub or _sub_status not in ("past_due", "canceled", "incomplete_expired"):
        st.markdown("<h2 style='text-align:center;margin:16px 0 4px;'>Choose Your Plan</h2>",
                    unsafe_allow_html=True)
        st.markdown("<p style='text-align:center;color:#555;margin-bottom:20px;'>"
                    "Pick the plan that fits your team. Cancel any time.</p>",
                    unsafe_allow_html=True)

    # ── If checkout URL already generated, show single CTA ───────────────
    _success = _app_url + "/?checkout=success"
    _cancel  = _app_url + "/?checkout=canceled"

    if st.session_state.get("_checkout_url"):
        _plan_name = st.session_state.get("_checkout_plan", "your plan")
        st.subheader(f"Ready to checkout: {_plan_name}")
        st.link_button("Complete checkout on Stripe →", st.session_state["_checkout_url"],
                       use_container_width=True, type="primary")
        if st.button("← Choose a different plan"):
            st.session_state.pop("_checkout_url", None)
            st.session_state.pop("_checkout_plan", None)
            st.rerun()
        st.markdown("---")
        if st.button("Sign out"):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()
        st.stop()

    # ── Get price IDs ─────────────────────────────────────────────────────
    try:
        _price_starter  = st.secrets.get("STRIPE_PRICE_STARTER", "")
        _price_pro      = st.secrets.get("STRIPE_PRICE_PRO", "")
        _price_business = st.secrets.get("STRIPE_PRICE_BUSINESS", "")
    except Exception:
        _price_starter = _price_pro = _price_business = ""
    _prices = {"starter": _price_starter, "pro": _price_pro, "business": _price_business}

    # ── Plan cards ────────────────────────────────────────────────────────
    _cols = st.columns(3)
    for _ci, _pk in enumerate(_PORD):
        _pi      = _PINFO[_pk]
        _is_prev = (_pk == _sub_plan and _sub_status in ("canceled", "incomplete_expired"))
        _is_pop  = (_pk == "pro")
        with _cols[_ci]:
            # Badge row
            if _is_prev:
                st.markdown(f"<div style='color:#d97706;font-size:11px;font-weight:700;"
                            f"text-transform:uppercase;'>↺ Your Previous Plan</div>",
                            unsafe_allow_html=True)
            elif _is_pop:
                st.markdown("<div style='color:#2563eb;font-size:11px;font-weight:700;"
                            "text-transform:uppercase;'>★ Most Popular</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown("<div style='font-size:11px;'>&nbsp;</div>", unsafe_allow_html=True)

            st.markdown(f"**{_pi['label']}**")
            st.markdown(f"### {_pi['price']}")
            st.caption(f"{_pi['emp']} employees · {_pi['desc']}")
            st.markdown("---")
            for _f in _FEATS[_pk]:
                st.markdown(f"<div style='font-size:12px;color:#444;line-height:1.9;'>✓ {_f}</div>",
                            unsafe_allow_html=True)
            st.markdown("")

            # CTA
            _price_id = _prices.get(_pk, "")
            if _is_prev and _portal_url:
                # Reactivate via plan-targeted portal flow (safe fallback to generic portal).
                _reactivate_url = create_billing_portal_url(
                    return_url=_app_url + "/?portal=return",
                    target_price_id=_price_map.get(_pk, ""),
                    flow="subscription_update",
                ) or _portal_url
                st.link_button(f"Reactivate {_pi['label']} →", _reactivate_url,
                               use_container_width=True, type="primary")
            elif _price_id:
                _btn_label  = f"Get {_pi['label']}"
                _btn_type   = "primary" if _is_pop else "secondary"
                if st.button(_btn_label, use_container_width=True, type=_btn_type,
                             key=f"btn_{_pk}"):
                    with st.spinner("Connecting to Stripe..."):
                        url, err = create_stripe_checkout_url(_price_id, _success, _cancel)
                    if url:
                        st.session_state["_checkout_url"] = url
                        st.session_state["_checkout_plan"] = _pi["label"]
                        st.rerun()
                    else:
                        st.error(f"Checkout failed: {err}")

    if not (_price_starter or _price_pro or _price_business):
        st.info("Payment system is being configured. Check back soon.")

    st.markdown("---")
    if st.button("Sign out"):
        for k in list(st.session_state.keys()): del st.session_state[k]
        st.rerun()


def main():
    # ── Auth gate ──────────────────────────────────────────────────────────
    if "supabase_session" not in st.session_state:
        _login_page()
        st.stop()

    # ── Idle timeout ──────────────────────────────────────────────────────
    if _check_session_timeout():
        st.info("Session expired due to inactivity. Please sign in again.")
        _login_page()
        st.stop()

    # ── Billing portal return — force subscription re-sync ────────────────
    if st.query_params.get("portal") == "return":
        st.query_params.clear()
        st.session_state.pop("_sub_active", None)
        st.session_state.pop("_portal_synced_plan", None)
        _bust_cache()
        with st.spinner("Refreshing your subscription…"):
            try:
                _synced_ok = _verify_checkout_and_activate()
            except Exception:
                _synced_ok = False
        if _synced_ok:
            # Read back what was written so we can show a confirmation
            try:
                from database import get_subscription as _gs
                _new_sub = _gs()
                if _new_sub:
                    st.session_state["_portal_synced_plan"] = _new_sub.get("plan", "").capitalize()
            except Exception:
                pass
        st.rerun()

    # ── Periodic Stripe sync (heals stale DB plan/status) ─────────────────
    _now_sync = time.time()
    _last_sub_sync = float(st.session_state.get("_last_sub_sync_ts", 0) or 0)
    if _now_sync - _last_sub_sync > 300:
        try:
            _sync_ok = _verify_checkout_and_activate()
            if _sync_ok:
                _bust_cache()
        except Exception:
            pass
        st.session_state["_last_sub_sync_ts"] = _now_sync

    # ── Subscription gate ─────────────────────────────────────────────────
    # Admin bypass: add ADMIN_EMAILS in secrets to skip subscription check
    _admin_emails = []
    try:
        _admin_emails = [e.strip().lower() for e in st.secrets.get("ADMIN_EMAILS", "").split(",") if e.strip()]
    except Exception:
        pass
    _user_email = st.session_state.get("user_email", "").lower()
    if _user_email and _user_email in _admin_emails:
        st.session_state["_sub_active"] = True

    if not st.session_state.get("_sub_active"):
        try:
            from database import has_active_subscription
            if has_active_subscription():
                st.session_state["_sub_active"] = True
            else:
                # No local subscription — try syncing from Stripe in case
                # payment was completed but webhook/verification missed it
                _synced = False
                try:
                    _synced = _verify_checkout_and_activate()
                except Exception:
                    pass
                if _synced and has_active_subscription():
                    st.session_state["_sub_active"] = True
                else:
                    _subscription_page()
                    st.stop()
        except Exception as _sub_err:
            # If subscription check fails (table missing, etc.), let user through
            st.session_state["_sub_active"] = True

    # Start background email thread (idempotent — only creates once per process)
    _start_email_thread()

    # Show confirmation if we just synced a plan from the billing portal
    if st.session_state.get("_portal_synced_plan"):
        _synced_label = st.session_state.pop("_portal_synced_plan")
        st.toast(f"Subscription updated — you're now on the {_synced_label} plan.", icon="✅")

    # Check email schedules at most once per 60 seconds (not every click)
    _now = time.time()
    if _now - st.session_state.get("_last_email_check", 0) > 60:
        st.session_state["_last_email_check"] = _now
        try:
            _run_scheduled_reports_for_tenant(force_now=False)
        except Exception as _eml_err:
            _email_log(f"Page-render email check error: {_eml_err}")
            _log_app_error("email", f"Schedule check error: {_eml_err}", detail=traceback.format_exc())

    page = render_sidebar()
    handlers = {
        "supervisor": page_supervisor,
        "dashboard": page_dashboard,
        "import": page_import,
        "employees": page_employees,
        "productivity": page_productivity,
        "email": page_email,
        "settings": page_settings,
    }
    handler = handlers.get(page, page_import)
    try:
        handler()
    except Exception as _page_err:
        _tb = traceback.format_exc()
        _log_app_error(
            "page",
            f"Page render failed ({page}): {_page_err}",
            detail=_tb,
            severity="error",
        )
        st.error("This page encountered an unexpected error.")
        with st.expander("Technical details"):
            st.code(_tb)


if __name__ == "__main__":
    try:
        main()
    except Exception as _fatal_err:
        _fatal_tb = traceback.format_exc()
        try:
            _log_app_error("fatal", f"Unhandled app error: {_fatal_err}", detail=_fatal_tb)
        except Exception:
            pass
        st.error("A fatal app error occurred. Please refresh or contact support.")
        with st.expander("Technical details"):
            st.code(_fatal_tb)
