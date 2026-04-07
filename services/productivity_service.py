"""Productivity domain helpers extracted from pages/productivity.py.

Pure calculation / DB / export functions — no Streamlit calls.
"""
import io
from collections import defaultdict
from datetime import date, timedelta

from core.dependencies import _cached_employees, _get_db_client


# ── Priority list risk scoring ─────────────────────────────────────────────

def _calc_priority_risk_level(emp: dict, history: list) -> tuple:
    """Calculate performance risk level."""
    risk_score = 0.0
    details = {"trend_score": 0, "streak_score": 0, "variance_score": 0}

    trend = emp.get("trend", "insufficient_data")
    if trend == "down":
        risk_score += 4
        details["trend_score"] = 4
    elif trend == "flat":
        risk_score += 1
        details["trend_score"] = 1
    elif trend == "up":
        risk_score -= 2
        details["trend_score"] = -2

    emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
    target_uph = emp.get("Target UPH", "—")
    under_goal_streak = 0

    if history and target_uph != "—":
        try:
            target = float(target_uph)
            emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
            if emp_history:
                sorted_hist = sorted(emp_history, key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                for r in reversed(sorted_hist):
                    try:
                        uph_val = float(r.get("UPH", 0) or 0)
                        if uph_val < target:
                            under_goal_streak += 1
                        else:
                            break
                    except (ValueError, TypeError):
                        break
        except (ValueError, TypeError):
            pass

    if under_goal_streak >= 7:
        risk_score += 5
        details["streak_score"] = 5
    elif under_goal_streak >= 5:
        risk_score += 4
        details["streak_score"] = 4
    elif under_goal_streak >= 3:
        risk_score += 2.5
        details["streak_score"] = 2.5
    elif under_goal_streak >= 1:
        risk_score += 0.5
        details["streak_score"] = 0.5

    uph_values = []
    if history:
        emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
        for r in emp_history:
            try:
                uph_val = float(r.get("UPH", 0) or 0)
                if uph_val > 0:
                    uph_values.append(uph_val)
            except (ValueError, TypeError):
                pass

    if len(uph_values) >= 3:
        avg_uph = sum(uph_values) / len(uph_values)
        variance = sum((x - avg_uph) ** 2 for x in uph_values) / len(uph_values)
        std_dev = variance ** 0.5
        coeff_variation = (std_dev / avg_uph * 100) if avg_uph > 0 else 0

        if coeff_variation > 30:
            risk_score += 3
            details["variance_score"] = 3
        elif coeff_variation > 20:
            risk_score += 1.5
            details["variance_score"] = 1.5
        elif coeff_variation > 10:
            risk_score += 0.5
            details["variance_score"] = 0.5

        details["variance_pct"] = round(coeff_variation, 1)

    details["under_goal_streak"] = under_goal_streak
    details["total_score"] = round(risk_score, 1)

    if risk_score >= 7:
        return "🔴 High", risk_score, details
    elif risk_score >= 4:
        return "🟡 Medium", risk_score, details
    else:
        return "🟢 Low", risk_score, details


# ── Coaching score calculation ─────────────────────────────────────────────

def _calc_coaching_score(emp: dict, context_tags: list, history: list) -> tuple:
    """Score an employee for coaching need. Higher = more urgent.
    Context tags reduce urgency (e.g., new employees need coaching but lower priority).
    Also returns whether they meet strict auto-flag criteria.
    Returns (score, reasons, context_impact, meets_auto_flag_criteria).
    """
    score = 0.0
    reasons = []

    # Factor 1: How far below goal
    target = emp.get("Target UPH") or 0
    avg_uph = emp.get("Average UPH") or 0
    under_goal = False
    if target and target != "—":
        try:
            target = float(target)
            avg_uph = float(avg_uph)
            pct_below = ((target - avg_uph) / target * 100) if target > 0 else 0
            score += abs(pct_below) * 0.5  # 0.5 points per percent below
            reasons.append(f"{pct_below:.1f}% below target")
            under_goal = True
        except (ValueError, TypeError):
            pass

    # Factor 2: Negative trend (direction + magnitude)
    trend = emp.get("trend", "insufficient_data")
    change_pct = emp.get("change_pct", 0.0)
    trending_down = False
    try:
        change_pct = float(change_pct)
    except (ValueError, TypeError):
        change_pct = 0.0

    if trend == "down":
        score += 5  # Declining is a red flag
        reasons.append(f"Declining trend ({change_pct:+.1f}%)")
        trending_down = True
    elif trend == "flat":
        score += 2  # Flat is concerning when below goal
        if change_pct <= -3 or change_pct >= 3:
            reasons.append(f"Flat trend ({change_pct:+.1f}%)")
        else:
            reasons.append("Flat trend")
    elif change_pct < 0:
        score += 1  # Slight decline
        reasons.append(f"Slight decline ({change_pct:+.1f}%)")

    # Factor 3: Streak (consecutive entries below their average)
    emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
    streak = 0
    emp_avg = avg_uph
    if history and emp_avg:
        emp_history = [r for r in history
                       if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
        if emp_history:
            sorted_hist = sorted(emp_history,
                                 key=lambda r: (r.get("Date", "") or r.get("Week", "")))
            for r in reversed(sorted_hist):
                try:
                    uph_val = float(r.get("UPH", 0) or 0)
                    if uph_val < emp_avg:
                        streak += 1
                    else:
                        break
                except (ValueError, TypeError):
                    break
            if streak >= 3:
                score += streak * 0.3
                reasons.append(f"{streak} consecutive entries below avg")

    # Strict auto-flag criteria: under goal AND trending down AND multi-day streak
    meets_auto_flag_criteria = (under_goal and trending_down and streak >= 3)

    # Factor 4: Context modifier — reduce urgency based on situational factors
    context_impact = []
    if "New employee" in context_tags:
        score *= 0.5
        context_impact.append("(New — expect ramp-up period)")
    if "Cross-training" in context_tags:
        score *= 0.6
        context_impact.append("(In cross-training)")
    if "Equipment issues" in context_tags:
        score *= 0.4
        context_impact.append("(Address equipment first)")
    if "Shift change" in context_tags:
        score *= 0.5
        context_impact.append("(Recent shift change)")
    if "Short staffed" in context_tags:
        score *= 0.7
        context_impact.append("(Team capacity issue)")

    return score, reasons, context_impact, meets_auto_flag_criteria


# ── Period report helpers ───────────────────────────────────────────────────

def _resolve_period_dates(period: str) -> tuple:
    """Convert a named period string to (start_date, end_date)."""
    today = date.today()
    if period == "Prior day":
        return today - timedelta(days=1), today - timedelta(days=1)
    elif period == "Current week":
        return today - timedelta(days=today.weekday()), today
    elif period == "Prior week":
        end = today - timedelta(days=today.weekday() + 1)
        return end - timedelta(days=6), end
    elif period == "Prior month":
        first_of_this = today.replace(day=1)
        end = first_of_this - timedelta(days=1)
        return end.replace(day=1), end
    return today - timedelta(days=1), today - timedelta(days=1)


def _email_risk_level(emp: dict, history_all: list) -> str:
    """Quick risk calculation for email."""
    risk_score = 0.0
    trend = emp.get("trend", "insufficient_data")
    if trend == "down":
        risk_score += 4
    elif trend == "flat":
        risk_score += 1
    elif trend == "up":
        risk_score -= 2

    emp_id = emp.get("EmployeeID", "")
    if history_all and emp.get("Target UPH") != "—":
        try:
            target = float(emp.get("Target UPH"))
            emp_hist = [r for r in history_all if r.get("emp_id") == emp_id]
            streak = 0
            for r in reversed(emp_hist[-10:]):
                try:
                    uph = float(r.get("uph") or 0)
                    if uph < target:
                        streak += 1
                    else:
                        break
                except Exception:
                    break
            if streak >= 3:
                risk_score += 3
        except Exception:
            pass

    if risk_score >= 7:
        return "🔴 High"
    elif risk_score >= 4:
        return "🟡 Medium"
    else:
        return "🟢 Low"


def _build_period_report(d_start, d_end, dept_choice: str, depts: list,
                          gs: list, targets: dict, tenant_id: str = "",
                          plan_name: str = "starter") -> tuple:
    """
    Build Excel report bytes, subject, and HTML body for a date range.
    d_start / d_end are date objects. Returns (xl_bytes, subj, body).
    """
    today = date.today()

    from_iso     = d_start.isoformat()
    to_iso       = d_end.isoformat()
    period_label = from_iso if from_iso == to_iso else f"{from_iso} – {to_iso}"
    dept_label   = dept_choice if dept_choice != "All departments" else "All Departments"
    subj         = f"Performance Report — {dept_label} — {period_label}"

    try:
        _sb = _get_db_client()
        _uph_q = _sb.table("uph_history").select(
            "emp_id, units, hours_worked, uph, work_date, department"
        ).gte("work_date", from_iso).lte("work_date", to_iso)
        _emp_q = _sb.table("employees").select("id, emp_id, name, department")
        if tenant_id:
            _uph_q = _uph_q.eq("tenant_id", tenant_id)
            _emp_q = _emp_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq3
            _uph_q = _tq3(_uph_q)
            _emp_q = _tq3(_emp_q)
        subs = (_uph_q.execute().data or [])
        # uph_history.emp_id is a bigint FK to employees.id — key by numeric id
        emps_lookup = {e["id"]: e for e in (_emp_q.execute().data or [])}
    except Exception:
        subs = []
        emps_lookup = {e["id"]: e for e in (_cached_employees() or [])}
    # Fallback to unit_submissions if uph_history returned nothing for that range
    if not subs:
        try:
            _sb2 = _get_db_client()
            _us_q = _sb2.table("unit_submissions").select(
                "emp_id, units, hours_worked, work_date, department"
            ).gte("work_date", from_iso).lte("work_date", to_iso)
            if tenant_id:
                _us_q = _us_q.eq("tenant_id", tenant_id)
            else:
                from database import _tq as _tq_us
                _us_q = _tq_us(_us_q)
            _us_rows = _us_q.execute().data or []
            for row in _us_rows:
                h = float(row.get("hours_worked") or 0)
                u = float(row.get("units") or 0)
                row["uph"] = round(u / h, 2) if h > 0 else 0.0
            subs = _us_rows
        except Exception:
            pass

    if dept_choice != "All departments":
        subs = [s for s in subs
                if (s.get("department", "") == dept_choice or
                    emps_lookup.get(s.get("emp_id", ""), {}).get("department", "") == dept_choice)]

    if not subs:
        body = (
            f"{dept_label} - {period_label}\n"
            f"No work data was found for {dept_label} between {from_iso} and {to_iso}.\n"
            "Employees may not have had submissions during this period, or the date "
            "range falls outside your imported data."
        )
        return None, subj, body

    emp_agg = defaultdict(lambda: {"units": 0.0, "hours": 0.0, "dept": ""})
    for s in subs:
        eid = s.get("emp_id", "")
        emp_agg[eid]["units"] += float(s.get("units") or 0)
        emp_agg[eid]["hours"] += float(s.get("hours_worked") or 0)
        emp_agg[eid]["dept"]   = emps_lookup.get(eid, {}).get("department", "")

    try:
        _sb = _get_db_client()
        _recent = (d_end - timedelta(days=30)).isoformat()
        _hist_q = _sb.table("uph_history").select("emp_id, uph, work_date").gte("work_date", _recent)
        if tenant_id:
            _hist_q = _hist_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq_hist
            _hist_q = _tq_hist(_hist_q)
        _email_hist = _hist_q.execute().data or []
    except Exception:
        _email_hist = []

    gs_by_emp = {str(r.get("EmployeeID", "")): r for r in (gs or []) if r.get("EmployeeID")}

    def _trend_snapshot(emp_id: str) -> tuple:
        if emp_id in gs_by_emp:
            _row = gs_by_emp[emp_id]
            return _row.get("trend", "insufficient_data"), float(_row.get("change_pct", 0) or 0)
        emp_hist = []
        for row in _email_hist:
            if row.get("emp_id") == emp_id:
                try:
                    emp_hist.append((row.get("work_date", ""), float(row.get("uph") or 0)))
                except Exception:
                    pass
        if len(emp_hist) < 4:
            return "insufficient_data", 0.0
        emp_hist.sort(key=lambda item: item[0])
        recent = [v for _, v in emp_hist[-3:]]
        previous = [v for _, v in emp_hist[-6:-3]]
        if not previous:
            return "insufficient_data", 0.0
        prev_avg = sum(previous) / len(previous)
        recent_avg = sum(recent) / len(recent)
        if prev_avg <= 0:
            return "insufficient_data", 0.0
        pct = round(((recent_avg - prev_avg) / prev_avg) * 100, 1)
        if pct >= 5:
            return "up", pct
        if pct <= -5:
            return "down", pct
        return "flat", pct

    scope_gs = []
    for eid, agg in emp_agg.items():
        emp_info = emps_lookup.get(eid, {})
        dept     = agg["dept"]
        uph      = round(agg["units"] / agg["hours"], 2) if agg["hours"] > 0 else 0
        tgt      = float(targets.get(dept, 0) or 0)
        trend, change_pct = _trend_snapshot(eid)
        scope_gs.append({
            "Employee Name": emp_info.get("name", eid),
            "EmployeeID":    eid,
            "Department":    dept,
            "Average UPH":   uph,
            "Total Units":   round(agg["units"]),
            "Hours Worked":  round(agg["hours"], 2),
            "Target UPH":    tgt if tgt else "—",
            "goal_status":   ("on_goal" if tgt and uph >= tgt
                              else "below_goal" if tgt else "no_goal"),
            "trend":         trend,
            "change_pct":    change_pct,
            "flagged":       False,
        })
    scope_gs.sort(key=lambda r: float(r.get("Average UPH", 0) or 0), reverse=True)

    on_g  = sum(1 for r in scope_gs if r["goal_status"] == "on_goal")
    below = sum(1 for r in scope_gs if r["goal_status"] == "below_goal")

    # Build a plain-text report body (no HTML rendering in service layer).
    _avg_all = round(sum(r["Average UPH"] for r in scope_gs) / len(scope_gs), 1) if scope_gs else 0

    _lines = [
        f"{dept_label} - {period_label}",
        f"Employees: {len(scope_gs)} | Avg UPH: {_avg_all} | On goal: {on_g} | Below goal: {below}",
        "",
        "Department Health",
    ]
    for _dept in sorted({r.get("Department", "") for r in scope_gs if r.get("Department")}):
        _dept_rows = [r for r in scope_gs if r.get("Department") == _dept]
        _dept_on = sum(1 for r in _dept_rows if r.get("goal_status") == "on_goal")
        _dept_total = len(_dept_rows)
        _dept_pct = round((_dept_on / _dept_total) * 100) if _dept_total else 0
        _lines.append(f"- {_dept}: {_dept_on}/{_dept_total} on goal ({_dept_pct}%)")

    _top3 = scope_gs[:3]
    if _top3:
        _lines.extend(["", "Top Performers"])
        for t in _top3:
            _lines.append(f"- {t['Employee Name']} ({t['Department']}) - {t['Average UPH']} UPH")

    _bottom = [r for r in reversed(scope_gs) if r["goal_status"] == "below_goal"][:3]
    if _bottom:
        _lines.extend(["", "Needs Coaching"])
        for b in _bottom:
            _risk = _email_risk_level(b, _email_hist)
            _lines.append(
                f"- {b['Employee Name']} ({b['Department']}) - {b['Average UPH']} UPH | Risk: {_risk}"
            )

    _critical_list = []
    for r in scope_gs:
        if r["goal_status"] == "below_goal" and "🔴" in _email_risk_level(r, _email_hist):
            _critical_list.append(r)
    if _critical_list:
        _lines.extend(["", "Priority: Needs Attention Today"])
        for c in _critical_list[:3]:
            _action = "Check in immediately."
            try:
                _cur = float(c.get("Average UPH") or 0)
                _tgt = float(c.get("Target UPH") or 0)
                _gap = _tgt - _cur
                _action = (
                    "Schedule 1-on-1 and discuss specific goals/support."
                    if _gap > 5 else "Check in on blockers or process issues."
                )
            except Exception:
                pass
            _lines.append(f"- {c['Employee Name']} ({c['Department']}) - {c['Average UPH']} UPH | Action: {_action}")

    _improved_list = [r for r in scope_gs if r.get("trend") == "up" and r.get("change_pct", 0) > 0]
    if _improved_list:
        _lines.extend(["", "Recognition: Improved This Period"])
        for imp in _improved_list[:3]:
            _pct = imp.get("change_pct", 0)
            _lines.append(
                f"- {imp['Employee Name']} ({imp['Department']}) +{_pct:.1f}% trend"
            )

    _trending_down = [r for r in scope_gs if r.get("trend") == "down" and r.get("goal_status") != "no_goal"]
    if _trending_down:
        _lines.extend(["", f"Early Warning: {len(_trending_down)} trending down"])
        for td in _trending_down[:5]:
            _pct = td.get("change_pct", 0)
            _status = "at risk" if td.get("goal_status") == "below_goal" else "still on track"
            _lines.append(f"- {td['Employee Name']} ({td['Department']}) {_pct:.1f}% ({_status})")

    if plan_name in ("pro", "business"):
        _lines.extend(["", "Top Risks"])
        for _row in [r for r in scope_gs if r.get("goal_status") == "below_goal"][:5]:
            _lines.append(
                f"- {_row['Employee Name']} ({_row['Department']}) {_row['Average UPH']} UPH | {_email_risk_level(_row, _email_hist)}"
            )

    if plan_name in ("pro", "business"):
        try:
            from settings import Settings as _ImpactS
            _impact_settings = _ImpactS(tenant_id=tenant_id) if tenant_id else _ImpactS()
            _avg_wage = float(_impact_settings.get("avg_hourly_wage", 18.0))
        except Exception:
            _avg_wage = 18.0
        _cost_total = 0.0
        for _row in scope_gs:
            try:
                _target = float(_row.get("Target UPH") or 0)
                _avg = float(_row.get("Average UPH") or 0)
                _hours = float(_row.get("Hours Worked") or 0)
                if _target > 0 and 0 < _avg < _target:
                    _cost_total += max(((_target - _avg) / _target) * _hours * _avg_wage, 0)
            except Exception:
                pass
        _lines.extend(["", f"Estimated cost impact: ${_cost_total:,.0f}"])

    _lines.extend(["", "See attached Excel report for full details and department breakdown."])
    body = "\n".join(_lines)

    xl_data = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        HDR_FILL  = PatternFill("solid", fgColor="FF0F2D52")
        HDR_FONT  = Font(bold=True, color="FFFFFFFF", size=10, name="Arial")
        GRN_FILL  = PatternFill("solid", fgColor="FFD4EDDA")
        RED_FILL  = PatternFill("solid", fgColor="FFF8D7DA")
        DARK_FONT = Font(color="FF1A2D42", size=10, name="Arial")
        STATUS_LABELS = {"on_goal": "On Goal", "below_goal": "Below Goal", "no_goal": "No Target"}

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Summary"
        ws["A1"] = f"{dept_label} — {period_label}"
        ws["A1"].font = Font(bold=True, size=13, name="Arial", color="FF0F2D52")
        ws["A2"] = f"Generated: {today.isoformat()}   |   {on_g} on goal · {below} below goal"
        ws["A2"].font = Font(size=10, name="Arial", color="FF5A7A9C")

        hdrs = ["Employee", "Department", "Total Units", "Hours Worked", "Avg UPH", "Target UPH", "Status"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(4, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(scope_gs, 5):
            fill = (GRN_FILL if r["goal_status"] == "on_goal"
                    else RED_FILL if r["goal_status"] == "below_goal" else None)
            for ci, v in enumerate([r["Employee Name"], r["Department"],
                                     r["Total Units"], r["Hours Worked"],
                                     r["Average UPH"], r["Target UPH"],
                                     STATUS_LABELS.get(r["goal_status"], "—")], 1):
                c = ws.cell(ri, ci, v); c.font = DARK_FONT
                c.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
                if fill: c.fill = fill

        if scope_gs:
            chart = BarChart(); chart.type = "bar"; chart.style = 10
            chart.title = f"Avg UPH — {period_label}"
            chart.width = 20; chart.height = max(8, len(scope_gs) * 0.55)
            last_r = 4 + len(scope_gs)
            chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=last_r), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last_r))
            chart.series[0].graphicalProperties.solidFill = "0F2D52"
            ws.add_chart(chart, "I5")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or "")) for c in col), default=8) + 3, 40)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); xl_data = buf.read()
    except Exception:
        pass

    return xl_data, subj, body
