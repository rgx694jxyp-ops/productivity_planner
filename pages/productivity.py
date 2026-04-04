from app import (
    _apply_mode_styling,
    _audit,
    _cached_active_flags,
    _cached_employees,
    _cached_targets,
    _calc_risk_level,
    _get_current_plan,
    _html_mod,
    _log_app_error,
    _plan_gate,
    _render_breadcrumb,
    _render_confidence_ux,
    _render_primary_action_rail,
    _render_session_context_bar,
    _render_soft_action_buttons,
    date,
    datetime,
    io,
    pd,
    st,
    tempfile,
    time,
    traceback,
)
try:
    from pages.common import _normalize_label_text
except Exception:
    def _normalize_label_text(value, max_len: int = 64) -> str:
        s = str(value or "").replace("\x00", " ").strip()
        s = " ".join(s.split())
        s = s.replace("|", " ").replace("<", " ").replace(">", " ")
        s = s.strip(" '\"")
        if len(s) > max_len:
            s = s[: max_len - 3].rstrip() + "..."
        return s or "Unknown"
from pages.employees import _build_archived_productivity

def page_productivity():
    st.title("📈 Productivity")
    st.caption("UPH rankings, department goals, trend charts, and performance tracking.")

    # Context and breadcrumbs
    _prod_view = st.session_state.get("prod_view", "")
    _render_breadcrumb("productivity", _prod_view if _prod_view else None)
    _render_session_context_bar()

    try:
        from goals          import (analyse_trends, build_goal_status, get_all_targets,
                                    set_dept_target, flag_employee, unflag_employee,
                                    add_note, get_active_flags, load_goals, save_goals)
        from ranker         import build_department_report
        from error_log      import ErrorLog
        from exporter       import export_excel
    except ImportError as e:
        st.error(f"Productivity module error: {e}"); return

    # Secondary page navigation: first choose mode, then choose view.
    _plan_now = _get_current_plan()
    _all_opts = ["🎯 Dept Goals", "📊 Goal Status", "📈 Trends", "📉 Rolling Avg", "📅 Weekly", "💰 Labor Cost", "📋 Priority List", "🧑‍🏫 Coaching"]
    if _plan_now in ("pro", "business", "admin"):
        _monitor_opts = ["📊 Goal Status", "📈 Trends", "📉 Rolling Avg", "📅 Weekly", "📋 Priority List", "🧑‍🏫 Coaching"]
        _plan_opts = ["🎯 Dept Goals", "💰 Labor Cost"]
    else:
        # Starter tier: core weekly output + ranking-focused view.
        _monitor_opts = ["📅 Weekly", "📋 Priority List"]
        _plan_opts = []

    _mode_options = ["Monitor"] + (["Plan"] if _plan_opts else [])
    _mode_default = st.session_state.get("prod_mode", "Monitor")
    if _mode_default not in _mode_options:
        _mode_default = "Monitor"
    prod_mode = st.radio(
        "Productivity mode",
        _mode_options,
        index=_mode_options.index(_mode_default),
        horizontal=True,
        key="prod_mode",
    )
    # Apply mode-specific visual styling
    _apply_mode_styling(prod_mode)
    
    PROD_OPTS = _monitor_opts if prod_mode == "Monitor" else _plan_opts

    if "prod_view" not in st.session_state or st.session_state.prod_view not in PROD_OPTS:
        st.session_state.prod_view = PROD_OPTS[0]

    chosen_prod = st.radio(
        "Productivity views",
        PROD_OPTS,
        index=PROD_OPTS.index(st.session_state.prod_view),
        horizontal=True,
        label_visibility="collapsed",
        key="prod_view",
    )

    _render_primary_action_rail(
        st.session_state.get("goal_status", []),
        st.session_state.get("history", []),
        key_prefix="prod_primary",
    )
    _render_confidence_ux(st.session_state.get("history", []))

    st.divider()

    _targets_snapshot = _cached_targets()

    class _PS:
        def get(self, k, d=None): return st.session_state.get(k, d)
        def get_output_dir(self): return tempfile.gettempdir()
        def get_dept_target_uph(self, d):
            t = _targets_snapshot.get(d, 0)
            return float(t) if t else 0.0
        def all_mappings(self): return st.session_state.get("mapping", {})

    # ── Helper: re-apply goals to current pipeline data ──────────────────────
    def _reapply_goals():
        if not st.session_state.pipeline_done:
            return
        try:
            targets  = _cached_targets()
            tw       = st.session_state.get("trend_weeks", 4)
            history  = st.session_state.get("history", [])
            mapping  = st.session_state.get("mapping", {})
            trend_data = analyse_trends(history, mapping, weeks=tw) if history else {}
            goal_status = build_goal_status(
                st.session_state.get("top_performers", []), targets, trend_data)
            ps  = _PS()
            log = ErrorLog(tempfile.gettempdir())
            dept_report = build_department_report(
                st.session_state.get("top_performers", []), ps, log)
            st.session_state.goal_status = goal_status
            st.session_state.dept_report = dept_report
            st.session_state.trend_data  = trend_data
        except Exception as _e:
            pass   # silently skip — archived data may lack history

    # Always refresh from archived DB data on Productivity page so it includes
    # prior imports and newly imported rows in one combined view.
    _last_arch_refresh = float(st.session_state.get("_archived_last_refresh_ts", 0.0) or 0.0)
    _arch_refresh_due = (time.time() - _last_arch_refresh) > 600
    if _arch_refresh_due:
        _show_loading = not st.session_state.get("_archived_loaded")
        try:
            if _show_loading:
                with st.spinner("Loading productivity data…"):
                    _build_archived_productivity()
            else:
                _build_archived_productivity()
            st.session_state["_archived_last_refresh_ts"] = time.time()
        except BaseException as _ae:
            _log_app_error("productivity", f"Archive load error: {repr(_ae)[:500]}", detail=traceback.format_exc())

    # Reapply goals only when targets changed (not every click)
    _fresh_targets = _cached_targets()
    _prev_targets  = st.session_state.get("_last_applied_targets")
    if st.session_state.pipeline_done and st.session_state.get("top_performers") and _fresh_targets != _prev_targets:
        try:
            _fresh_td      = st.session_state.get("trend_data", {})
            _fresh_gs      = build_goal_status(
                st.session_state.top_performers, _fresh_targets, _fresh_td)
            st.session_state.goal_status = _fresh_gs
            st.session_state["_last_applied_targets"] = dict(_fresh_targets)
            # Only rebuild dept_report for live data — archived version is already correct
            if not st.session_state.get("_archived_loaded"):
                _ps2  = _PS()
                _log2 = ErrorLog(tempfile.gettempdir())
                st.session_state.dept_report = build_department_report(
                    st.session_state.top_performers, _ps2, _log2)
        except Exception:
            pass

    # ── DEPT GOALS ────────────────────────────────────────────────────────────
    if chosen_prod == "🎯 Dept Goals":
        if not _plan_gate("pro", "Department Goals"):
            return
        st.subheader("Department UPH targets")
        st.caption("Set a simple target for each department. Once targets are in place, the app can tell you who is on track and who needs help.")
        st.caption("Trend window and Top/Bottom % only affect ranking/highlighting views.")

        st.info("Quick start: if you do not have formal standards yet, enter your best current expectation for each department and refine later.")

        with st.expander("Advanced scoring controls"):
            tw = st.slider("Trend window used for trend scoring (weeks)", 2, 12,
                           st.session_state.get("trend_weeks", 4), key="prod_tw")
            if tw != st.session_state.get("trend_weeks", 4):
                st.session_state.trend_weeks = tw
                _reapply_goals()
                st.rerun()

            hc1, hc2 = st.columns(2)
            st.session_state.top_pct = hc1.slider(
                "Top % bucket (green highlight)", 0, 50,
                st.session_state.get("top_pct", 10), key="goals_top_pct"
            )
            st.session_state.bot_pct = hc2.slider(
                "Bottom % bucket (red highlight)", 0, 50,
                st.session_state.get("bot_pct", 10), key="goals_bot_pct"
            )

        st.divider()

        targets   = _cached_targets()
        goals_obj = load_goals()

        # Auto-populate departments from all available sources
        if st.session_state.pipeline_done:
            all_depts = set()
            # From goal_status
            for r in st.session_state.get("goal_status", []):
                if r.get("Department"): all_depts.add(r["Department"])
            # From top_performers
            for r in st.session_state.get("top_performers", []):
                if r.get("Department"): all_depts.add(r["Department"])
            # From employees table (catches archived data with blank uph_history dept)
            for e in (_cached_employees() or []):
                if e.get("department"): all_depts.add(e["department"])
            for d in all_depts:
                if d and d not in targets:
                    set_dept_target(d, 0.0)
                    _raw_cached_targets.clear()
            targets = _cached_targets()

        dept_list  = sorted(targets.keys())
        goal_changed = False
        pending_goal_changes: list[tuple[str, float, float]] = []

        if not dept_list:
            st.info("No departments yet. Run Import Data first — departments are detected automatically from your CSV.")
        else:
            # Header row
            hc1, hc2, hc3 = st.columns([3, 2, 1])
            hc1.markdown("**Department**")
            hc2.markdown("**UPH Target**")
            hc3.markdown("")
            st.divider()

            for dept in dept_list:
                cur = float(targets.get(dept, 0) or 0)
                c1, c2, c3 = st.columns([3, 2, 1])
                c1.markdown(f"**{dept}**")

                # Seed text input once — don't overwrite after user changes it
                seed_key = f"goal_seed_{dept}"
                txt_key  = f"goal_txt_{dept}"
                if seed_key not in st.session_state:
                    st.session_state[seed_key] = True
                    # Seed from persisted JSON value — this runs on every fresh session
                    st.session_state[txt_key] = str(int(cur)) if cur else ""
                elif txt_key not in st.session_state:
                    # Key was cleared somehow — re-seed from JSON
                    st.session_state[txt_key] = str(int(cur)) if cur else ""

                c2.text_input("UPH", key=txt_key,
                              label_visibility="collapsed",
                              placeholder="e.g. 18")

                # Parse draft value; defer DB writes until explicit save.
                raw = st.session_state.get(txt_key, "")
                try:
                    new_val = float(raw.strip()) if raw.strip() else 0.0
                except (ValueError, TypeError):
                    new_val = cur

                if abs(new_val - cur) > 0.001:
                    pending_goal_changes.append((dept, cur, new_val))

                if c3.button("✕", key=f"rm_dept_{dept}", help="Remove department"):
                    goals_obj["dept_targets"].pop(dept, None)
                    save_goals(goals_obj)
                    _raw_cached_targets.clear()
                    st.session_state.pop(seed_key, None)
                    st.session_state.pop(txt_key,  None)
                    goal_changed = True

            if pending_goal_changes:
                st.info(f"{len(pending_goal_changes)} unsaved goal change(s).")
                if st.button("Save goal changes", key="save_dept_goal_changes", type="primary", use_container_width=True):
                    for _dept, _cur, _new in pending_goal_changes:
                        set_dept_target(_dept, _new)
                        _audit("GOAL_TARGET", f"{_dept} | {_cur} → {_new}")
                    _raw_cached_targets.clear()
                    goal_changed = True

        # Departments are added automatically from the pipeline — no manual add form

        # Reapply goals to all charts after any change
        if goal_changed and st.session_state.pipeline_done:
            _reapply_goals()
            st.toast("✓ Goals updated", icon="🎯")
            st.rerun()

    # ── GOAL STATUS ───────────────────────────────────────────────────────────
    elif chosen_prod == "📊 Goal Status":
        if not _plan_gate("pro", "Goal Status"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        gs = st.session_state.get("goal_status", [])
        # If goal_status is empty but we have top_performers, rebuild it
        if not gs and st.session_state.get("top_performers"):
            try:
                gs = build_goal_status(
                    st.session_state.top_performers, _cached_targets(),
                    st.session_state.get("trend_data", {}))
                st.session_state.goal_status = gs
            except Exception:
                pass
        if not gs:
            st.info("No data yet — run Import Data or check that UPH history exists.")
            return

        total    = len(gs)
        on_goal  = sum(1 for r in gs if r.get("goal_status") == "on_goal")
        below    = sum(1 for r in gs if r.get("goal_status") == "below_goal")
        trending = sum(1 for r in gs if r.get("trend") == "down")
        flagged  = sum(1 for r in gs if r.get("flagged"))
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Employees", total)
        m2.metric("On goal 🟢", on_goal)
        m3.metric("Below goal 🔴", below)
        m4.metric("Trending ↓", trending)
        m5.metric("Flagged 🚩", flagged)
        st.divider()

        depts  = sorted({r.get("Department","") for r in gs if r.get("Department")})
        fc1,fc2 = st.columns(2)
        dept_f  = fc1.selectbox("Filter by department", ["All departments"] + depts, key="gs_dept")
        stat_f  = fc2.multiselect("Filter by status",
                    ["On goal","Below goal","No target set"],
                    default=["On goal","Below goal","No target set"], key="gs_stat")
        stat_map = {"On goal":"on_goal","Below goal":"below_goal","No target set":"no_goal"}
        allowed  = {stat_map[s] for s in stat_f}
        filtered = [r for r in gs
                    if (dept_f == "All departments" or r.get("Department") == dept_f)
                    and r.get("goal_status") in allowed]

        TREND  = {"up":"↑ Improving","down":"↓ Declining","flat":"→ Stable","insufficient_data":"—"}
        STATUS = {"on_goal":"🟢","below_goal":"🔴","no_goal":"⚪"}
        rows   = [{
            "🚩":        "🚩" if r.get("flagged") else "",
            "Dept":      r.get("Department",""),
            "Shift":     r.get("Shift",""),
            "Employee":  r.get("Employee Name",""),
            "Avg UPH":   round(float(r.get("Average UPH",0) or 0),2),
            "Target":    r.get("Target UPH") if r.get("Target UPH") not in ("—", None, "", 0) else None,
            "vs Target": r.get("vs Target") if r.get("vs Target") not in ("—", None, "") else None,
            "Status":    STATUS.get(r.get("goal_status",""),""),
            "Trend":     TREND.get(r.get("trend",""),"—"),
            "Change":    f"{r.get('change_pct',0):+.1f}%",
        } for r in filtered]

        df = pd.DataFrame(rows)
        hl_map = {"🟢":"#1a6b3a","🔴":"#8b1a1a","⚪":"#2a3a4a"}
        def hl(s):
            row = rows[s.name]
            clr = hl_map.get(row.get("Status",""), "#2a3a4a")
            if row.get("Trend","").startswith("↓") and clr == "#1a6b3a": clr = "#7a5c00"
            return [f"background-color:{clr}; color:#ffffff" for _ in s]
        st.dataframe(
            df.style.apply(hl, axis=1),
            use_container_width=True, hide_index=True)
        st.caption("Employees are compared against the UPH target set in Dept Goals. 🟢 Avg UPH ≥ target · 🔴 Avg UPH < target · 🟡 On goal but UPH is declining week-over-week · ⚪ No target set for this department")
        _gs_buf = io.BytesIO()
        df.to_excel(_gs_buf, index=False, engine="openpyxl")
        _gs_buf.seek(0)
        st.download_button("⬇️ Download Goal Status.xlsx", _gs_buf.read(),
                           f"Goal_Status_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_goal_status")

        st.divider()
        st.subheader("🚩 Flag employees for performance tracking")
        emp_labels      = [
            f"{_normalize_label_text(r.get('Employee Name',''))} — "
            f"{_normalize_label_text(r.get('Department',''), max_len=24)} "
            f"({_normalize_label_text(r.get('Shift',''), max_len=18)})"
            for r in filtered
        ]
        active_flag_ids = set(_cached_active_flags().keys())
        if emp_labels:
            _flag_tab1, _flag_tab2 = st.tabs(["Flag individual", "Bulk flag / unflag"])
            with _flag_tab1:
                if "flag_reason_val" not in st.session_state:
                    st.session_state.flag_reason_val = ""
                fc3, fc4, fc5 = st.columns([3,2,1])
                _below_labels = [l for l, r in zip(emp_labels, filtered) if r.get("goal_status") == "below_goal"]
                _flag_default  = emp_labels.index(_below_labels[0]) if _below_labels else 0
                sel_emp = fc3.selectbox("Employee", emp_labels, index=_flag_default, key="flag_sel")
                reason  = fc4.text_input("Reason", value=st.session_state.flag_reason_val,
                                          key="flag_reason", placeholder="Optional reason")
                if fc5.button("Flag", type="primary", use_container_width=True):
                    idx    = emp_labels.index(sel_emp)
                    emp    = filtered[idx]
                    emp_id = str(emp.get("EmployeeID", emp.get("Employee Name","")))
                    flag_employee(emp_id, emp.get("Employee Name",""), emp.get("Department",""), reason.strip())
                    _audit("FLAG", f"{emp.get('Employee Name','')} | dept={emp.get('Department','')} | reason={reason.strip()}")
                    _raw_cached_active_flags.clear()
                    _raw_cached_all_coaching_notes.clear()
                    _raw_cached_coaching_notes_for.clear()
                    st.session_state.flag_reason_val = ""
                    st.session_state.pop("flag_reason", None)
                    st.toast(f"✓ {emp.get('Employee Name','')} flagged", icon="🚩")
                    st.rerun()
            with _flag_tab2:
                # ── Currently flagged summary ──────────────────────────────
                _currently_flagged = [
                    (_lbl, _r) for _lbl, _r in zip(emp_labels, filtered)
                    if str(_r.get("EmployeeID", _r.get("Employee Name",""))) in active_flag_ids
                ]
                if _currently_flagged:
                    st.markdown(
                        f"<div style='background:#FFF3E0;border-left:4px solid #E65100;"
                        f"border-radius:0 6px 6px 0;padding:10px 14px;margin-bottom:12px;'>"
                        f"<b style='color:#BF360C;font-size:12px;text-transform:uppercase;"
                        f"letter-spacing:.05em;'>🚩 {len(_currently_flagged)} Active Flag(s)</b><br>"
                        + "".join(
                            f"<span style='color:#BF360C;font-size:13px;'>• {_html_mod.escape(_lbl)}</span><br>"
                            for _lbl, _ in _currently_flagged
                        )
                        + "</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.info("No employees currently flagged.")

                st.caption("Select employees below — 🚩 marks those already flagged.")
                # Build display labels with 🚩 prefix for already-flagged employees
                _bulk_display = []
                for _lbl, _r in zip(emp_labels, filtered):
                    _eid = str(_r.get("EmployeeID", _r.get("Employee Name", "")))
                    _bulk_display.append(f"🚩 {_lbl}" if _eid in active_flag_ids else _lbl)

                bulk_sel_disp = st.multiselect("Select employees", _bulk_display, key="bulk_flag_sel")
                # Strip the 🚩 prefix to recover the plain label for lookup
                bulk_sel      = [l.removeprefix("🚩 ") for l in bulk_sel_disp]

                bulk_reason = st.text_input("Reason (applies to all)", key="bulk_flag_reason", placeholder="Optional")
                bc1, bc2    = st.columns(2)
                if bc1.button("🚩 Flag all selected", type="primary", use_container_width=True, key="bulk_flag_btn"):
                    if bulk_sel:
                        for lbl in bulk_sel:
                            _bi = emp_labels.index(lbl)
                            _be = filtered[_bi]
                            _beid = str(_be.get("EmployeeID", _be.get("Employee Name","")))
                            flag_employee(_beid, _be.get("Employee Name",""), _be.get("Department",""), bulk_reason.strip())
                        _raw_cached_active_flags.clear(); _raw_cached_all_coaching_notes.clear(); _raw_cached_coaching_notes_for.clear()
                        st.toast(f"✓ {len(bulk_sel)} employee(s) flagged", icon="🚩"); st.rerun()
                if bc2.button("✓ Unflag all selected", type="secondary", use_container_width=True, key="bulk_unflag_btn"):
                    if bulk_sel:
                        for lbl in bulk_sel:
                            _bi = emp_labels.index(lbl)
                            _be = filtered[_bi]
                            _beid = str(_be.get("EmployeeID", _be.get("Employee Name","")))
                            if _beid in active_flag_ids:
                                unflag_employee(_beid)
                        _raw_cached_active_flags.clear()
                        st.toast(f"✓ {len(bulk_sel)} employee(s) unflagged", icon="✅"); st.rerun()

    # ── TRENDS ────────────────────────────────────────────────────────────────
    elif chosen_prod == "📈 Trends":
        if not _plan_gate("pro", "Trend Analysis"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        trends = st.session_state.dept_trends
        if not trends: st.info("No trend data yet."); return
        df_t = pd.DataFrame(trends)
        # Date range filter
        months = sorted(df_t["Month"].unique()) if "Month" in df_t.columns else []
        if months:
            import re as _re
            min_m, max_m = months[0], months[-1]
            # Persist applied range in session state
            # Reset range if data was refreshed (new pipeline run)
            _data_key = f"trends_data_key_{min_m}_{max_m}"
            if st.session_state.get("_trends_data_key") != _data_key:
                st.session_state.trend_applied_from = min_m
                st.session_state.trend_applied_to   = max_m
                st.session_state["_trends_data_key"] = _data_key
            elif "trend_applied_from" not in st.session_state:
                st.session_state.trend_applied_from = min_m
            if "trend_applied_to" not in st.session_state:
                st.session_state.trend_applied_to   = max_m

            tr1, tr2, tr3 = st.columns([3, 3, 1])
            from_raw = tr1.text_input("From (YYYY-MM)", value=st.session_state.trend_applied_from,
                                       key="trend_from", placeholder="e.g. 2024-03")
            to_raw   = tr2.text_input("To (YYYY-MM)",   value=st.session_state.trend_applied_to,
                                       key="trend_to",   placeholder="e.g. 2025-03")
            if tr3.button("Apply", type="primary", use_container_width=True, key="trend_apply"):
                fm = from_raw.strip()
                tm = to_raw.strip()
                st.session_state.trend_applied_from = fm if _re.match(r"\d{4}-\d{2}", fm) else min_m
                st.session_state.trend_applied_to   = tm if _re.match(r"\d{4}-\d{2}", tm) else max_m
                st.rerun()

            from_month = st.session_state.trend_applied_from
            to_month   = st.session_state.trend_applied_to
            mask = (df_t["Month"] >= from_month) & (df_t["Month"] <= to_month)
            df_t = df_t[mask]
        st.subheader("Average UPH over time by department")
        st.caption("Each line shows the average UPH for that department across all employees for each month. Use From/To to zoom into a date range.")
        try:
            import math as _math2
            df_t_clean = df_t[df_t["Average UPH"].apply(
                lambda x: _math2.isfinite(float(x)) if x is not None else False)]
            if not df_t_clean.empty:
                st.line_chart(df_t_clean.pivot(index="Month", columns="Department", values="Average UPH"),
                              use_container_width=True)
        except Exception: pass
        st.dataframe(df_t, use_container_width=True, hide_index=True)
        _tr_buf = io.BytesIO()
        df_t.to_excel(_tr_buf, index=False, engine="openpyxl")
        _tr_buf.seek(0)
        st.download_button("⬇️ Download Trends.xlsx", _tr_buf.read(),
                           f"Trends_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_trends_prod")

    # ── ROLLING AVG ───────────────────────────────────────────────────────────
    elif chosen_prod == "📉 Rolling Avg":
        if not _plan_gate("pro", "Rolling Averages"):
            return
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        rolling = st.session_state.get("employee_rolling_avg", [])
        if not rolling: st.info("No rolling average data yet."); return
        df_r = pd.DataFrame(rolling)

        # Employee filter
        employees = sorted(df_r["Employee"].unique()) if "Employee" in df_r.columns else []
        if employees:
            emp_sel = st.multiselect("Filter employees", employees, key="roll_emp_sel")
            if emp_sel:
                df_r = df_r[df_r["Employee"].isin(emp_sel)]

        st.subheader("7-Day and 14-Day Rolling Average UPH per Employee")
        st.caption("Each point shows the rolling average UPH for that employee on that date.")

        # Select rolling period
        roll_period = st.selectbox("Rolling Period", ["7-Day", "14-Day"], key="roll_period")

        # Chart
        try:
            col_name = "7DayRollingAvg" if roll_period == "7-Day" else "14DayRollingAvg"
            df_r_clean = df_r.dropna(subset=[col_name])
            if not df_r_clean.empty:
                st.line_chart(df_r_clean.pivot(index="Date", columns="Employee", values=col_name),
                              use_container_width=True)
        except Exception: pass
        st.dataframe(df_r, use_container_width=True, hide_index=True)
        _r_buf = io.BytesIO()
        df_r.to_excel(_r_buf, index=False, engine="openpyxl")
        _r_buf.seek(0)
        st.download_button("⬇️ Download RollingAvg.xlsx", _r_buf.read(),
                           f"RollingAvg_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_rolling_prod")

    # ── RISK ASSESSMENT ───────────────────────────────────────────────────────
    elif chosen_prod == "⚠️ Risk Assessment":
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        risk = st.session_state.get("employee_risk", [])
        if not risk: st.info("No risk assessment data yet."); return
        df_risk = pd.DataFrame(risk)

        # Filter by risk level
        risk_levels = ["high", "medium", "low"]
        sel_levels = st.multiselect("Filter by Risk Level", risk_levels, default=risk_levels, key="risk_level_sel")
        if sel_levels:
            df_risk = df_risk[df_risk["Risk Level"].isin(sel_levels)]

        st.subheader("Employee Risk Assessment")
        st.caption("Risk levels based on under goal streak, downward trend, and below department average.")

        # Summary
        level_counts = df_risk["Risk Level"].value_counts()
        col1, col2, col3 = st.columns(3)
        col1.metric("High Risk", level_counts.get("high", 0))
        col2.metric("Medium Risk", level_counts.get("medium", 0))
        col3.metric("Low Risk", level_counts.get("low", 0))

        st.dataframe(df_risk, use_container_width=True, hide_index=True)
        _risk_buf = io.BytesIO()
        df_risk.to_excel(_risk_buf, index=False, engine="openpyxl")
        _risk_buf.seek(0)
        st.download_button("⬇️ Download RiskAssessment.xlsx", _risk_buf.read(),
                           f"RiskAssessment_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_risk_prod")

    # ── WEEKLY ────────────────────────────────────────────────────────────────
    elif chosen_prod == "📅 Weekly":
        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()
        weekly = st.session_state.weekly_summary
        if not weekly: st.info("No weekly data yet."); return
        df_w = pd.DataFrame(weekly)

        # Date range filter
        all_from = sorted(df_w["From"].dropna().unique()) if "From" in df_w.columns else []
        all_to   = sorted(df_w["To"].dropna().unique())   if "To"   in df_w.columns else []
        all_dates = sorted(set(all_from + all_to))

        wk1, wk2 = st.columns(2)
        if all_dates:
            d_from = wk1.date_input("From", value=datetime.strptime(all_dates[0], "%Y-%m-%d").date(),
                                    key="wk_from")
            d_to   = wk2.date_input("To",   value=datetime.strptime(all_dates[-1], "%Y-%m-%d").date(),
                                    key="wk_to")
            mask = pd.Series([True] * len(df_w))
            if "From" in df_w.columns:
                mask &= df_w["From"] >= d_from.isoformat()
            if "To" in df_w.columns:
                mask &= df_w["To"] <= d_to.isoformat()
            df_w = df_w[mask]

        st.subheader("Total units per week by department")
        st.caption("Each row represents one department's output for a given week. Use the date range to filter.")

        # Chart
        try:
            df_w_clean = df_w.dropna(subset=["Total Units"])
            if not df_w_clean.empty:
                chart_label = df_w_clean["Week Range"] if "Week Range" in df_w_clean.columns else df_w_clean["Week"]
                chart_df = df_w_clean.copy()
                chart_df["Period"] = chart_label
                st.line_chart(chart_df.pivot(index="Period", columns="Department", values="Total Units"),
                              use_container_width=True)
        except Exception: pass

        # Table grouped by week
        display_cols = [c for c in ["Week Range", "Department", "Avg UPH", "Total Units", "Record Count"]
                        if c in df_w.columns]
        if not display_cols:
            display_cols = [c for c in df_w.columns if c not in ("From", "To")]

        weeks_sorted = sorted(df_w["Week"].unique()) if "Week" in df_w.columns else []
        for wk in weeks_sorted:
            wk_slice = df_w[df_w["Week"] == wk]
            if wk_slice.empty:
                continue
            wk_range = wk_slice["Week Range"].iloc[0] if "Week Range" in wk_slice.columns else wk
            st.markdown(f"**{wk}** &nbsp;&nbsp; {wk_range}")
            show_cols = [c for c in display_cols if c != "Week Range"]
            st.dataframe(wk_slice[show_cols], use_container_width=True, hide_index=True)

        _wk_buf = io.BytesIO()
        df_w[display_cols].to_excel(_wk_buf, index=False, engine="openpyxl")
        _wk_buf.seek(0)
        st.download_button("⬇️ Download Weekly.xlsx", _wk_buf.read(),
                           f"Weekly_{date.today()}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_weekly")

    elif chosen_prod == "💰 Labor Cost":
        if not _plan_gate("pro", "Labor Cost Impact"):
            return

        st.subheader("Labor Cost Impact Analysis")
        st.caption("See the dollar impact of employee performance vs targets. Enter your average hourly wage to calculate.")

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        _hourly_wage = st.number_input("Average hourly wage ($)", min_value=0.0, value=18.0, step=0.50, key="labor_wage")
        _targets_map = _cached_targets()

        # Build labor cost table
        _lc_rows = []
        for r in gs:
            name = r.get("Employee", r.get("Employee Name", ""))
            dept = r.get("Department", "")
            uph  = r.get("Avg UPH", r.get("Average UPH"))
            target = r.get("Target UPH")
            hours = r.get("Hours Worked") or r.get("HoursWorked")

            if target in ("—", None, "", 0):
                target = _targets_map.get(dept, 0)

            if not uph or target in ("—", None, "", 0):
                continue
            try:
                uph = float(uph)
                target = float(target)
            except (ValueError, TypeError):
                continue

            if target <= 0:
                continue

            # Estimate hours from data or default to 40
            try:
                hours = float(hours) if hours and hours not in ("—", None, "") else 40.0
            except (ValueError, TypeError):
                hours = 40.0

            expected_units = target * hours
            actual_units = uph * hours
            unit_diff = actual_units - expected_units
            # Cost per unit = wage / target_uph
            cost_per_unit = _hourly_wage / target if target > 0 else 0
            dollar_impact = unit_diff * cost_per_unit

            _lc_rows.append({
                "Employee": name,
                "Department": dept,
                "Avg UPH": round(uph, 1),
                "Target": round(target, 1),
                "UPH Diff": round(uph - target, 1),
                "Est. Hours": round(hours, 1),
                "Unit Diff": round(unit_diff, 0),
                "$ Impact": round(dollar_impact, 2),
            })

        if not _lc_rows:
            st.info("No employees with both UPH and targets set.")
            return

        df_lc = pd.DataFrame(_lc_rows).sort_values("$ Impact")

        # Summary metrics
        total_loss = sum(r["$ Impact"] for r in _lc_rows if r["$ Impact"] < 0)
        total_gain = sum(r["$ Impact"] for r in _lc_rows if r["$ Impact"] > 0)
        net_impact = total_loss + total_gain

        # TOP 3 UNDERPERFORMERS — Personal & Actionable
        st.subheader("⚠️ Top Cost Impact")
        underperformers = [r for r in _lc_rows if r["$ Impact"] < 0]
        underperformers.sort(key=lambda x: x["$ Impact"])  # Most negative first
        top_3 = underperformers[:3]
        
        if top_3:
            top_3_cost = sum(emp["$ Impact"] for emp in top_3)
            st.error(f"🔴 Top 3 underperformers costing **${abs(top_3_cost):,.0f}** this week")
            
            for i, emp in enumerate(top_3, 1):
                col1, col2, col3 = st.columns([2, 1.5, 1.5])
                col1.write(f"**{i}. {emp['Employee']}** ({emp['Department']})")
                col2.metric("UPH Gap", f"{emp['UPH Diff']:.1f}", "Below target")
                col3.metric("Weekly Cost", f"${abs(emp['$ Impact']):,.0f}", "Lost productivity")

        st.divider()
        
        # Summary metrics
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("Lost from Underperformance", f"-${abs(total_loss):,.0f}", delta_color="inverse")
        mc2.metric("Gained from Overperformance", f"+${total_gain:,.0f}")
        mc3.metric("Net Impact", f"${net_impact:,.0f}",
                    delta=f"${net_impact:,.0f}", delta_color="normal")

        st.markdown("---")

        # Color code: red for negative, green for positive
        st.subheader("Employee Breakdown")
        st.caption("Negative = costing you money vs target. Positive = saving you money.")

        def _color_impact(val):
            try:
                v = float(val)
                if v < 0: return "color: #FF4444"
                if v > 0: return "color: #44AA44"
            except: pass
            return ""

        styled = df_lc.style.map(_color_impact, subset=["$ Impact", "UPH Diff"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Department summary
        st.subheader("Department Summary")
        dept_summary = {}
        for r in _lc_rows:
            d = r["Department"]
            if d not in dept_summary:
                dept_summary[d] = {"Department": d, "Employees": 0, "Total $ Impact": 0}
            dept_summary[d]["Employees"] += 1
            dept_summary[d]["Total $ Impact"] += r["$ Impact"]
        df_dept = pd.DataFrame(dept_summary.values())
        df_dept["Total $ Impact"] = df_dept["Total $ Impact"].round(2)
        st.dataframe(df_dept, use_container_width=True, hide_index=True)

    # ── PRIORITY LIST ──────────────────────────────────────────────────────────────
    elif chosen_prod == "📋 Priority List":
        st.subheader("📋 Priority List")
        st.caption("Employees below goal ranked by risk (combines trend + streak + variance).")

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        history = st.session_state.get("history", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        # Filter for below_goal employees
        below_goal = [r for r in gs if r.get("goal_status") == "below_goal"]
        if not below_goal:
            st.success("✅ All employees are meeting their goals!")
            return

        # Calculate risk for all below-goal employees
        def _calc_priority_risk_level(emp, history):
            """Calculate performance risk level."""
            risk_score = 0.0
            details = {"trend_score": 0, "streak_score": 0, "variance_score": 0}
            
            trend = emp.get("trend", "insufficient_data")
            if trend == "down":
                risk_score += 4
                details["trend_score"] = 4
            elif trend == "flat":
                risk_score += 1
                details["trend_score"] = 1
            elif trend == "up":
                risk_score -= 2
                details["trend_score"] = -2
            
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            target_uph = emp.get("Target UPH", "—")
            under_goal_streak = 0
            
            if history and target_uph != "—":
                try:
                    target = float(target_uph)
                    emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                    if emp_history:
                        sorted_hist = sorted(emp_history, key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                        for r in reversed(sorted_hist):
                            try:
                                uph_val = float(r.get("UPH", 0) or 0)
                                if uph_val < target:
                                    under_goal_streak += 1
                                else:
                                    break
                            except (ValueError, TypeError):
                                break
                except (ValueError, TypeError):
                    pass
            
            if under_goal_streak >= 7:
                risk_score += 5
                details["streak_score"] = 5
            elif under_goal_streak >= 5:
                risk_score += 4
                details["streak_score"] = 4
            elif under_goal_streak >= 3:
                risk_score += 2.5
                details["streak_score"] = 2.5
            elif under_goal_streak >= 1:
                risk_score += 0.5
                details["streak_score"] = 0.5
            
            uph_values = []
            if history:
                emp_history = [r for r in history if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                for r in emp_history:
                    try:
                        uph_val = float(r.get("UPH", 0) or 0)
                        if uph_val > 0:
                            uph_values.append(uph_val)
                    except (ValueError, TypeError):
                        pass
            
            if len(uph_values) >= 3:
                avg_uph = sum(uph_values) / len(uph_values)
                variance = sum((x - avg_uph) ** 2 for x in uph_values) / len(uph_values)
                std_dev = variance ** 0.5
                coeff_variation = (std_dev / avg_uph * 100) if avg_uph > 0 else 0
                
                if coeff_variation > 30:
                    risk_score += 3
                    details["variance_score"] = 3
                elif coeff_variation > 20:
                    risk_score += 1.5
                    details["variance_score"] = 1.5
                elif coeff_variation > 10:
                    risk_score += 0.5
                    details["variance_score"] = 0.5
                
                details["variance_pct"] = round(coeff_variation, 1)
            
            details["under_goal_streak"] = under_goal_streak
            details["total_score"] = round(risk_score, 1)
            
            if risk_score >= 7:
                return "🔴 High", risk_score, details
            elif risk_score >= 4:
                return "🟡 Medium", risk_score, details
            else:
                return "🟢 Low", risk_score, details

        # Score all below-goal employees
        priority_list = []
        for emp in below_goal:
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            risk_level, risk_score, risk_details = _calc_priority_risk_level(emp, history)
            priority_list.append({
                "name": emp.get("Employee", emp.get("Employee Name", "Unknown")),
                "department": emp.get("Department", ""),
                "emp_id": emp_id,
                "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
                "target_uph": emp.get("Target UPH", "—"),
                "trend": emp.get("trend", "—"),
                "change_pct": emp.get("change_pct", 0.0),
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_details": risk_details,
            })

        # Sort by risk score (highest first)
        priority_list.sort(key=lambda x: x["risk_score"], reverse=True)

        # Summary
        high_risk = len([r for r in priority_list if r["risk_level"] == "🔴 High"])
        med_risk = len([r for r in priority_list if r["risk_level"] == "🟡 Medium"])
        low_risk = len([r for r in priority_list if r["risk_level"] == "🟢 Low"])
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Below Goal", len(priority_list))
        col2.metric("🔴 High Risk", high_risk)
        col3.metric("🟡 Medium Risk", med_risk)
        col4.metric("🟢 Low Risk", low_risk)

        st.divider()

        # Display table
        table_data = []
        for r in priority_list:
            table_data.append({
                "Risk": r["risk_level"],
                "Name": r["name"],
                "Department": r["department"],
                "Current": r["avg_uph"],
                "Target": r["target_uph"],
                "Trend": r["trend"],
                "Streak": f"{r['risk_details']['under_goal_streak']}d",
                "Score": r["risk_score"],
            })

        df_prio = pd.DataFrame(table_data)
        
        def _color_risk(val):
            if "🔴" in str(val):
                return "background-color: #ffcccc; color: #8b0000"
            elif "🟡" in str(val):
                return "background-color: #fff9e6; color: #ff6600"
            elif "🟢" in str(val):
                return "background-color: #e6ffe6; color: #008000"
            return ""

        compact_prio = st.checkbox("Compact table (small screens)", value=False, key="prio_compact")
        if compact_prio:
            df_show = df_prio[["Risk", "Name", "Current", "Streak"]]
        else:
            df_show = df_prio

        styled = df_show.style.map(_color_risk, subset=["Risk"])
        st.dataframe(styled, use_container_width=True, hide_index=True)

        st.markdown("##### Quick actions")
        for i, r in enumerate(priority_list[:12]):
            qa1, qa2, qa3, qa4 = st.columns([3, 2, 1, 1])
            qa1.write(r["name"])
            qa2.caption(r["department"] or "No dept")
            if qa3.button("Coach", key=f"prio_coach_{i}"):
                st.session_state["goto_page"] = "employees"
                st.session_state["emp_view"] = "Performance Journal"
                st.rerun()
            if qa4.button("History", key=f"prio_hist_{i}"):
                st.session_state["goto_page"] = "employees"
                st.session_state["emp_view"] = "Employee History"
                st.rerun()

        # Download
        csv_buf = df_prio.to_csv(index=False)
        st.download_button("⬇️ Download priority list", csv_buf, f"priority_list_{date.today()}.csv", "text/csv")

    # ── COACHING CORNER ───────────────────────────────────────────────────────────
    elif chosen_prod == "🧑‍🏫 Coaching":
        if not _plan_gate("pro", "Coaching"):
            return
        st.subheader("🧑‍🏫 Who Needs Coaching?")
        st.caption("Top 3 employees who need performance coaching based on trend + goal status + context.")

        # Define available context tags
        CONTEXT_TAGS = [
            "New employee",
            "Cross-training",
            "Equipment issues",
            "Shift change",
            "Short staffed",
        ]

        if not st.session_state.pipeline_done and not st.session_state.get("_archived_loaded"):
            _build_archived_productivity()

        gs = st.session_state.get("goal_status", [])
        history = st.session_state.get("history", [])
        if not gs:
            st.info("No productivity data. Import data first.")
            return

        # Filter for below_goal employees
        coaching_candidates = [r for r in gs if r.get("goal_status") == "below_goal"]
        if not coaching_candidates:
            st.success("✅ All employees are meeting their goals! Great work!")
            return

        # Load active flags and their context tags
        active_flags_data = get_active_flags()

        # Calculate coaching score for each candidate


        def _calc_coaching_score(emp, context_tags):
            """Score an employee for coaching need. Higher = more urgent.
            Context tags reduce urgency (e.g., new employees need coaching but lower priority).
            Also returns whether they meet strict auto-flag criteria."""
            score = 0.0
            reasons = []

            # Factor 1: How far below goal
            target = emp.get("Target UPH") or 0
            avg_uph = emp.get("Average UPH") or 0
            under_goal = False
            if target and target != "—":
                try:
                    target = float(target)
                    avg_uph = float(avg_uph)
                    pct_below = ((target - avg_uph) / target * 100) if target > 0 else 0
                    score += abs(pct_below) * 0.5  # 0.5 points per percent below
                    reasons.append(f"{pct_below:.1f}% below target")
                    under_goal = True
                except (ValueError, TypeError):
                    pass

            # Factor 2: Negative trend (direction + magnitude)
            trend = emp.get("trend", "insufficient_data")
            change_pct = emp.get("change_pct", 0.0)
            trending_down = False
            try:
                change_pct = float(change_pct)
            except (ValueError, TypeError):
                change_pct = 0.0

            if trend == "down":
                score += 5  # Declining is a red flag
                reasons.append(f"Declining trend ({change_pct:+.1f}%)")
                trending_down = True
            elif trend == "flat":
                score += 2  # Flat is concerning when below goal
                if change_pct <= -3 or change_pct >= 3:
                    reasons.append(f"Flat trend ({change_pct:+.1f}%)")
                else:
                    reasons.append("Flat trend")
            elif change_pct < 0:
                score += 1  # Slight decline
                reasons.append(f"Slight decline ({change_pct:+.1f}%)")

            # Factor 3: Streak (consecutive entries below their average)
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            streak = 0
            emp_avg = avg_uph
            if history and emp_avg:
                emp_history = [r for r in history
                               if str(r.get("EmployeeID", r.get("Employee Name", ""))) == emp_id]
                if emp_history:
                    # Sort by date/week if available
                    sorted_hist = sorted(emp_history,
                                        key=lambda r: (r.get("Date", "") or r.get("Week", "")))
                    # Count consecutive entries below the employee's average
                    for r in reversed(sorted_hist):
                        try:
                            uph_val = float(r.get("UPH", 0) or 0)
                            if uph_val < emp_avg:
                                streak += 1
                            else:
                                break
                        except (ValueError, TypeError):
                            break
                    if streak >= 3:
                        score += streak * 0.3
                        reasons.append(f"{streak} consecutive entries below avg")

            # Strict auto-flag criteria: under goal AND trending down AND multi-day streak
            meets_auto_flag_criteria = (under_goal and trending_down and streak >= 3)

            # Factor 4: Context modifier — reduce urgency based on situational factors
            context_impact = []
            if "New employee" in context_tags:
                score *= 0.5  # New employees are less urgent
                context_impact.append("(New — expect ramp-up period)")
            if "Cross-training" in context_tags:
                score *= 0.6  # Cross-training reduces urgency
                context_impact.append("(In cross-training)")
            if "Equipment issues" in context_tags:
                score *= 0.4  # Equipment issues are not coaching-related
                context_impact.append("(Address equipment first)")
            if "Shift change" in context_tags:
                score *= 0.5  # Shift changes take time to adjust
                context_impact.append("(Recent shift change)")
            if "Short staffed" in context_tags:
                score *= 0.7  # Staffing is a workload not coaching issue
                context_impact.append("(Team capacity issue)")

            return score, reasons, context_impact, meets_auto_flag_criteria

        # Score and rank all candidates
        scored = []
        for emp in coaching_candidates:
            emp_id = str(emp.get("EmployeeID", emp.get("Employee Name", "")))
            # Get context tags for this employee from their flag
            emp_context = []
            if emp_id in active_flags_data:
                flag_info = active_flags_data[emp_id]
                emp_context = flag_info.get("context_tags", [])

            score, reasons, context_impact, meets_auto_flag_criteria = _calc_coaching_score(emp, emp_context)
            risk_level, risk_score, risk_details = _calc_risk_level(emp, history)
            
            emp_name = emp.get("Employee", emp.get("Employee Name", "Unknown"))
            scored.append({
                "employee": emp_name,
                "department": emp.get("Department", ""),
                "emp_id": emp_id,
                "avg_uph": round(float(emp.get("Average UPH", 0) or 0), 2),
                "target_uph": emp.get("Target UPH", "—"),
                "trend": emp.get("trend", "—"),
                "change_pct": emp.get("change_pct", 0.0),
                "score": score,
                "reasons": reasons,
                "context_tags": emp_context,
                "context_impact": context_impact,
                "meets_auto_flag_criteria": meets_auto_flag_criteria,
                "risk_level": risk_level,
                "risk_score": risk_score,
                "risk_details": risk_details,
            })

        # Sort by score (descending) and take top 3
        scored.sort(key=lambda x: x["score"], reverse=True)
        top_3 = scored[:3]

        if not top_3:
            st.success("✅ All employees are meeting their goals!")
            return

        # Display each in an expandable card
        for idx, emp in enumerate(top_3, 1):
            with st.container(border=True):
                col1, col2, col3 = st.columns([2, 1, 1])
                col1.markdown(f"### #{idx} — {emp['employee']}")
                col2.markdown(f"**{emp['department']}**")

                # Status badge
                if emp['context_tags']:
                    status_text = "Under goal BUT " + ", ".join(emp['context_tags']).lower()
                    col3.markdown(f"*{status_text}*")
                else:
                    col3.markdown("⚠️ **Needs coaching**")

                st.divider()

                # Current performance
                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("Current UPH", f"{emp['avg_uph']}", delta=f"{emp['change_pct']:+.1f}%")
                mc2.metric("Target UPH", emp['target_uph'] if emp['target_uph'] != "—" else "—")
                mc3.metric("Trend", emp['trend'].replace("_", " ").title())
                mc4.markdown(f"<div style='text-align: center; padding: 10px;'><div style='font-size: 2em;'>{emp['risk_level'].split()[0]}</div><div style='font-size: 0.8em;'>Risk</div></div>", unsafe_allow_html=True)

                # Risk breakdown (collapsed)
                with st.expander("📊 Risk breakdown", expanded=False):
                    rd = emp['risk_details']
                    rc1, rc2, rc3 = st.columns(3)
                    rc1.metric("Trend score", rd['trend_score'])
                    rc2.metric("Under-goal streak", f"{rd['under_goal_streak']} days")
                    rc3.metric("Variance", f"{rd.get('variance_pct', 0):.1f}%")
                    st.caption(f"Total risk score: {rd['total_score']}")

                # Context tags display
                if emp['context_tags']:
                    st.markdown("#### Context")
                    tag_cols = st.columns(len(emp['context_tags']))
                    for tag_col, tag in zip(tag_cols, emp['context_tags']):
                        tag_col.markdown(f"🏷️ **{tag}**", help="This context reduces coaching priority")

                # Why they need coaching
                st.markdown("#### Performance Issues")
                for reason in emp["reasons"]:
                    st.markdown(f"• {reason}")

                # Context impact notes
                if emp['context_impact']:
                    st.markdown("#### Context Notes")
                    for note in emp['context_impact']:
                        st.info(note, icon="📌")

                # Suggested actions based on trend and gap
                st.markdown("#### Suggested Actions")
                actions = []

                # Equipment-specific action
                if "Equipment issues" in emp['context_tags']:
                    actions.append("**Fix the equipment first** — Resolve tool/system issues before coaching on performance.")

                # New employee specific action
                if "New employee" in emp['context_tags']:
                    actions.append("**Structured onboarding check** — Ensure they have proper training and mentoring.")

                # Cross-training specific action
                if "Cross-training" in emp['context_tags']:
                    actions.append("**Support the transition** — Provide extra mentoring during skill-building phase.")

                # Shift change specific action
                if "Shift change" in emp['context_tags']:
                    actions.append("**Allow adjustment time** — Follow up in 2 weeks to see if new rhythm improves performance.")

                # Staffing specific action
                if "Short staffed" in emp['context_tags']:
                    actions.append("**Increase team capacity** — Hiring or task redistribution may solve this faster than coaching.")

                # Trend-based actions (only if not covered by context)
                if not any(tag in emp['context_tags'] for tag in ["Equipment issues", "Short staffed"]):
                    if emp["trend"] == "down":
                        actions.append("**Identify obstacles** — Ask what's changed. Look for workload spikes or personal issues.")
                    elif emp["trend"] == "flat":
                        actions.append("**Break the plateau** — Try new training, task rotation, or peer mentoring.")

                # Gap-based actions
                try:
                    target = float(emp["target_uph"]) if emp["target_uph"] != "—" else 0
                    if target > 0:
                        gap = target - emp["avg_uph"]
                        if gap > 5 and "New employee" not in emp['context_tags']:
                            actions.append("**Major gap** — Structured improvement plan with weekly check-ins.")
                        elif gap > 2 and "New employee" not in emp['context_tags']:
                            actions.append("**1-on-1 coaching** — Discuss goals and what support they need.")
                except (ValueError, TypeError):
                    pass

                # Default actions
                if not actions:
                    actions.append("**1-on-1 conversation** — Discuss performance, barriers, and support.")

                for action in actions:
                    st.markdown(f"→ {action}")

                st.divider()
                
                # ── SOFT ACTIONS FOR MEDIUM-RISK EMPLOYEES ──
                _render_soft_action_buttons(emp['emp_id'], emp['employee'], emp['risk_level'], emp['context_tags'])
                
                st.divider()

                # Context tagging section
                st.markdown("#### Add Context")
                st.caption("Select tags that explain underperformance (coaching urgency will adjust):")
                selected_context = st.multiselect(
                    "Context tags:",
                    CONTEXT_TAGS,
                    default=emp['context_tags'],
                    key=f"context_{emp['emp_id']}",
                    label_visibility="collapsed",
                )

                # Save context tags if changed
                if selected_context != emp['context_tags']:
                    try:
                        flags = get_active_flags()
                        if emp['emp_id'] in flags:
                            flags[emp['emp_id']]["context_tags"] = selected_context
                            # Save back to goals
                            goals_data = load_goals()
                            if emp['emp_id'] in goals_data["flagged_employees"]:
                                goals_data["flagged_employees"][emp['emp_id']]["context_tags"] = selected_context
                                save_goals(goals_data)
                            st.success("✓ Context tags updated")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error saving context: {e}")

                # Add quick coaching note button
                note_text = st.text_input(
                    f"Quick note for {emp['employee']}:",
                    key=f"coach_note_{emp['emp_id']}",
                    placeholder="e.g. 'Discussed new tools, plan to follow up Friday'",
                )
                if note_text:
                    if st.button(f"Log note", key=f"coach_save_{emp['emp_id']}", use_container_width=True):
                        try:
                            add_note(emp['emp_id'], note_text)
                            st.success(f"✅ Note saved for {emp['employee']}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error saving note: {e}")

                # Flag employee
                is_flagged = emp['emp_id'] in active_flags_data
                
                # Show auto-flag criteria status
                if not is_flagged:
                    if emp['meets_auto_flag_criteria']:
                        st.success("✅ Meets auto-flag criteria: under goal + trending down + multi-day streak", icon="✅")
                    else:
                        # Explain what's missing
                        missing = []
                        if emp['trend'] != "down":
                            missing.append("not trending down")
                        if not any(r for r in emp['reasons'] if "consecutive" in r):
                            missing.append("no multi-day streak")
                        missing_text = " + ".join(missing) if missing else "one or more criteria"
                        st.warning(f"⚠️ Doesn't meet auto-flag criteria ({missing_text}). Coaching may still help.", icon="⚠️")

                flag_col1, flag_col2 = st.columns([1, 1])
                if is_flagged:
                    if flag_col1.button(f"🚩 Unflag", key=f"unflag_{emp['emp_id']}", use_container_width=True):
                        try:
                            unflag_employee(emp['emp_id'])
                            st.success(f"✓ {emp['employee']} unflagged")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error unflagging: {e}")
                else:
                    if flag_col1.button(f"🚩 Flag for tracking", key=f"flag_{emp['emp_id']}", use_container_width=True):
                        try:
                            flag_employee(emp['emp_id'], emp['employee'], emp['department'],
                                         reason="Coaching priority: " + ", ".join(emp["reasons"][:2]))
                            # Save context tags with the flag
                            goals_data = load_goals()
                            if emp['emp_id'] in goals_data["flagged_employees"]:
                                goals_data["flagged_employees"][emp['emp_id']]["context_tags"] = selected_context
                                save_goals(goals_data)
                            st.success(f"✓ {emp['employee']} flagged for tracking")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error flagging: {e}")



# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EMAIL SETUP (preserved from previous version)
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_period_dates(period: str):
    """Convert a named period string to (start_date, end_date)."""
    from datetime import timedelta
    today = date.today()
    if period == "Prior day":
        return today - timedelta(days=1), today - timedelta(days=1)
    elif period == "Current week":
        return today - timedelta(days=today.weekday()), today
    elif period == "Prior week":
        end = today - timedelta(days=today.weekday() + 1)
        return end - timedelta(days=6), end
    elif period == "Prior month":
        first_of_this = today.replace(day=1)
        end = first_of_this - timedelta(days=1)
        return end.replace(day=1), end
    return today - timedelta(days=1), today - timedelta(days=1)


def _build_period_report(d_start, d_end, dept_choice: str, depts: list,
                          gs: list, targets: dict, tenant_id: str = "",
                          plan_name: str = "starter"):
    """
    Build Excel report bytes, subject, and HTML body for a date range.
    d_start / d_end are date objects. Returns (xl_bytes, subj, body).
    """
    from collections import defaultdict as _dc
    today = date.today()

    from_iso     = d_start.isoformat()
    to_iso       = d_end.isoformat()
    period_label = from_iso if from_iso == to_iso else f"{from_iso} – {to_iso}"
    dept_label   = dept_choice if dept_choice != "All departments" else "All Departments"
    subj         = f"Performance Report — {dept_label} — {period_label}"

    try:
        _sb = _get_db_client()
        _uph_q = _sb.table("uph_history").select(
            "emp_id, units, hours_worked, uph, work_date, department"
        ).gte("work_date", from_iso).lte("work_date", to_iso)
        _emp_q = _sb.table("employees").select("emp_id, name, department")
        if tenant_id:
            _uph_q = _uph_q.eq("tenant_id", tenant_id)
            _emp_q = _emp_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq3
            _uph_q = _tq3(_uph_q)
            _emp_q = _tq3(_emp_q)
        subs = (_uph_q.execute().data or [])
        emps_lookup = {e["emp_id"]: e for e in (_emp_q.execute().data or [])}
    except Exception:
        subs = []
        emps_lookup = {e["emp_id"]: e for e in (_cached_employees() or [])}
    if dept_choice != "All departments":
        subs = [s for s in subs
                if (s.get("department","") == dept_choice or
                    emps_lookup.get(s.get("emp_id",""),{}).get("department","") == dept_choice)]

    if not subs:
        body = (f"<h2>{dept_label} — {period_label}</h2>"
                f"<p>No work data was found for <strong>{dept_label}</strong> "
                f"between <strong>{from_iso}</strong> and <strong>{to_iso}</strong>.</p>"
                f"<p>Employees may not have had submissions during this period, "
                f"or the date range falls outside your imported data.</p>")
        return None, subj, body

    emp_agg = _dc(lambda: {"units": 0.0, "hours": 0.0, "dept": ""})
    for s in subs:
        eid = s.get("emp_id","")
        emp_agg[eid]["units"] += float(s.get("units") or 0)
        emp_agg[eid]["hours"] += float(s.get("hours_worked") or 0)
        emp_agg[eid]["dept"]   = emps_lookup.get(eid,{}).get("department","")

    try:
        _sb = _get_db_client()
        from datetime import timedelta
        _recent = (d_end - timedelta(days=30)).isoformat()
        _hist_q = _sb.table("uph_history").select("emp_id, uph, work_date").gte("work_date", _recent)
        if tenant_id:
            _hist_q = _hist_q.eq("tenant_id", tenant_id)
        else:
            from database import _tq as _tq_hist
            _hist_q = _tq_hist(_hist_q)
        _email_hist = _hist_q.execute().data or []
    except Exception:
        _email_hist = []

    gs_by_emp = {str(r.get("EmployeeID", "")): r for r in (gs or []) if r.get("EmployeeID")}

    def _trend_snapshot(emp_id: str) -> tuple[str, float]:
        if emp_id in gs_by_emp:
            _row = gs_by_emp[emp_id]
            return _row.get("trend", "insufficient_data"), float(_row.get("change_pct", 0) or 0)
        emp_hist = []
        for row in _email_hist:
            if row.get("emp_id") == emp_id:
                try:
                    emp_hist.append((row.get("work_date", ""), float(row.get("uph") or 0)))
                except Exception:
                    pass
        if len(emp_hist) < 4:
            return "insufficient_data", 0.0
        emp_hist.sort(key=lambda item: item[0])
        recent = [v for _, v in emp_hist[-3:]]
        previous = [v for _, v in emp_hist[-6:-3]]
        if not previous:
            return "insufficient_data", 0.0
        prev_avg = sum(previous) / len(previous)
        recent_avg = sum(recent) / len(recent)
        if prev_avg <= 0:
            return "insufficient_data", 0.0
        pct = round(((recent_avg - prev_avg) / prev_avg) * 100, 1)
        if pct >= 5:
            return "up", pct
        if pct <= -5:
            return "down", pct
        return "flat", pct

    scope_gs = []
    for eid, agg in emp_agg.items():
        emp_info = emps_lookup.get(eid, {})
        dept     = agg["dept"]
        uph      = round(agg["units"] / agg["hours"], 2) if agg["hours"] > 0 else 0
        tgt      = float(targets.get(dept, 0) or 0)
        trend, change_pct = _trend_snapshot(eid)
        scope_gs.append({
            "Employee Name": emp_info.get("name", eid),
            "EmployeeID":    eid,
            "Department":    dept,
            "Average UPH":   uph,
            "Total Units":   round(agg["units"]),
            "Hours Worked":  round(agg["hours"], 2),
            "Target UPH":    tgt if tgt else "—",
            "goal_status":   ("on_goal" if tgt and uph >= tgt
                              else "below_goal" if tgt else "no_goal"),
            "trend":         trend,
            "change_pct":    change_pct,
            "flagged":       False,
        })
    scope_gs.sort(key=lambda r: float(r.get("Average UPH", 0) or 0), reverse=True)

    on_g  = sum(1 for r in scope_gs if r["goal_status"] == "on_goal")
    below = sum(1 for r in scope_gs if r["goal_status"] == "below_goal")

    # Calculate risk levels for email (quick version)
    def _email_risk_level(emp, history_all):
        """Quick risk calculation for email."""
        risk_score = 0.0
        trend = emp.get("trend", "insufficient_data")
        if trend == "down":
            risk_score += 4
        elif trend == "flat":
            risk_score += 1
        elif trend == "up":
            risk_score -= 2
        
        # Approximation: use goal_status to estimate streak
        emp_id = emp.get("EmployeeID", "")
        if history_all and emp.get("Target UPH") != "—":
            try:
                target = float(emp.get("Target UPH"))
                emp_hist = [r for r in history_all if r.get("emp_id") == emp_id]
                streak = 0
                for r in reversed(emp_hist[-10:]):  # Check last 10 entries only
                    try:
                        uph = float(r.get("uph") or 0)
                        if uph < target:
                            streak += 1
                        else:
                            break
                    except:
                        break
                if streak >= 3:
                    risk_score += 3
            except:
                pass
        
        if risk_score >= 7:
            return "🔴 High"
        elif risk_score >= 4:
            return "🟡 Medium"
        else:
            return "🟢 Low"

    # Top 3 performers
    _top3 = scope_gs[:3]
    _top3_html = ""
    if _top3:
        _top3_html = "<h3>🏆 Top Performers</h3><ol>"
        for t in _top3:
            _top3_html += f"<li><strong>{t['Employee Name']}</strong> ({t['Department']}) — {t['Average UPH']} UPH</li>"
        _top3_html += "</ol>"

    # Bottom 3 performers (with targets & risk)
    _bottom = [r for r in reversed(scope_gs) if r["goal_status"] == "below_goal"][:3]
    _bottom_html = ""
    if _bottom:
        _bottom_html = "<h3>⚠️ Needs Coaching</h3><table style='width:100%; border-collapse: collapse;'>"
        _bottom_html += "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Employee</th><th style='text-align: center; border: 1px solid #ddd;'>Risk</th><th style='text-align: right; border: 1px solid #ddd;'>UPH</th></tr>"
        for b in _bottom:
            _risk = _email_risk_level(b, _email_hist)
            _tgt = b.get('Target UPH', '—')
            _diff = ""
            try:
                _diff = f" (vs {_tgt})"
            except (ValueError, TypeError):
                pass
            _bottom_html += f"<tr><td style='padding: 8px; border: 1px solid #ddd;'><strong>{b['Employee Name']}</strong><br/><span style='font-size: 0.9em; color: #666;'>{b['Department']}</span></td><td style='text-align: center; border: 1px solid #ddd;'>{_risk}</td><td style='text-align: right; border: 1px solid #ddd;'>{b['Average UPH']}{_diff}</td></tr>"
        _bottom_html += "</table>"

    # ACTION SECTION 1: TODAY's critical attention (🔴 High risk)
    _critical_html = ""
    _critical_list = []
    for r in scope_gs:
        if r["goal_status"] == "below_goal":
            _risk = _email_risk_level(r, _email_hist)
            if "🔴" in _risk:
                _critical_list.append(r)
    
    if _critical_list:
        _critical_html = "<h3 style='color: #8b0000;'>🔴 PRIORITY: Needs Attention TODAY</h3>"
        _critical_html += "<div style='background: #ffe6e6; border-left: 4px solid #8b0000; padding: 12px; margin-bottom: 16px;'>"
        for c in _critical_list[:3]:  # Top 3 critical
            _crit_risk = _email_risk_level(c, _email_hist)
            _action = ""
            try:
                _cur = float(c.get("Average UPH") or 0)
                _tgt = float(c.get("Target UPH") or 0)
                _gap = _tgt - _cur
                if _gap > 5:
                    _action = "👉 <strong>Action:</strong> Schedule 1-on-1. Discuss specific goals & support needed."
                else:
                    _action = "👉 <strong>Action:</strong> Check in on any blockers or issues."
            except:
                _action = "👉 <strong>Action:</strong> Check in immediately."
            
            _critical_html += f"<div style='margin-bottom: 12px;'><strong>{c['Employee Name']}</strong> ({c['Department']}) — {c['Average UPH']} UPH<br/>{_action}</div>"
        _critical_html += "</div>"

    # ACTION SECTION 2: Who improved this week (trend = up)
    _improved_html = ""
    _improved_list = [r for r in scope_gs if r.get("trend") == "up" and r.get("change_pct", 0) > 0]
    if _improved_list:
        _improved_html = "<h3 style='color: #008000;'>🟢 Recognition: Who Improved This Week</h3>"
        _improved_html += "<div style='background: #e6ffe6; border-left: 4px solid #008000; padding: 12px; margin-bottom: 16px;'>"
        for imp in _improved_list[:3]:  # Top 3 improvers
            _pct = imp.get("change_pct", 0)
            _improved_html += f"<div style='margin-bottom: 8px;'><strong>✓ {imp['Employee Name']}</strong> ({imp['Department']}) — <span style='color: green;'>+{_pct:.1f}%</span> trending up<br/>👉 <strong>Action:</strong> Recognize this improvement. Ask what's working & how to sustain it.</div>"
        _improved_html += "</div>"

    # ACTION SECTION 3: Who is trending down (early warning)
    _trending_down_html = ""
    _trending_down = [r for r in scope_gs if r.get("trend") == "down" and r.get("goal_status") != "no_goal"]
    if _trending_down:
        _trending_down_html = "<h3 style='color: #ff6600;'>⚠️ Early Warning: Trending Down</h3>"
        _trending_down_html += "<div style='background: #fff9e6; border-left: 4px solid #ff6600; padding: 12px; margin-bottom: 16px;'>"
        _trending_down_html += f"<p><strong>{len(_trending_down)} employee(s) showing declining performance:</strong></p>"
        for td in _trending_down[:5]:  # Show up to 5
            _pct = td.get("change_pct", 0)
            _status = "at risk" if td.get("goal_status") == "below_goal" else "still on track"
            _trending_down_html += f"<div style='margin-bottom: 8px;'><strong>• {td['Employee Name']}</strong> ({td['Department']}) — {_pct:.1f}% down ({_status})<br/>👉 <strong>Action:</strong> Proactive check-in. Address trend early before it worsens.</div>"
        _trending_down_html += "</div>"

    # Average UPH across all employees
    _avg_all = round(sum(r["Average UPH"] for r in scope_gs) / len(scope_gs), 1) if scope_gs else 0

    _dept_health_rows = []
    for _dept in sorted({r.get("Department", "") for r in scope_gs if r.get("Department")}):
        _dept_rows = [r for r in scope_gs if r.get("Department") == _dept]
        _dept_on = sum(1 for r in _dept_rows if r.get("goal_status") == "on_goal")
        _dept_total = len(_dept_rows)
        _dept_pct = round((_dept_on / _dept_total) * 100) if _dept_total else 0
        _dept_health_rows.append(
            f"<tr><td style='padding:8px;border:1px solid #ddd;'><strong>{_dept}</strong></td>"
            f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_dept_on}/{_dept_total}</td>"
            f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_dept_pct}%</td></tr>"
        )
    _dept_health_html = ""
    if _dept_health_rows:
        _dept_health_html = (
            "<h3>📊 Department Health</h3>"
            "<table style='width:100%; border-collapse: collapse;'>"
            "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Department</th>"
            "<th style='text-align: center; border: 1px solid #ddd;'>On Goal</th>"
            "<th style='text-align: center; border: 1px solid #ddd;'>Health</th></tr>"
            + "".join(_dept_health_rows)
            + "</table>"
        )

    _top_risks_html = ""
    if plan_name in ("pro", "business"):
        _risk_rows = []
        for _row in [r for r in scope_gs if r.get("goal_status") == "below_goal"][:5]:
            _risk_rows.append(
                f"<tr><td style='padding:8px;border:1px solid #ddd;'><strong>{_row['Employee Name']}</strong></td>"
                f"<td style='padding:8px;border:1px solid #ddd;'>{_row['Department']}</td>"
                f"<td style='padding:8px;border:1px solid #ddd;text-align:right;'>{_row['Average UPH']}</td>"
                f"<td style='padding:8px;border:1px solid #ddd;text-align:center;'>{_email_risk_level(_row, _email_hist)}</td></tr>"
            )
        if _risk_rows:
            _top_risks_html = (
                "<h3>🔴 Top Risks — Action Required Today</h3>"
                "<table style='width:100%; border-collapse: collapse;'>"
                "<tr style='background: #f5f5f5;'><th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Employee</th>"
                "<th style='padding: 8px; text-align: left; border: 1px solid #ddd;'>Dept</th>"
                "<th style='padding: 8px; text-align: right; border: 1px solid #ddd;'>UPH</th>"
                "<th style='padding: 8px; text-align: center; border: 1px solid #ddd;'>Risk</th></tr>"
                + "".join(_risk_rows)
                + "</table>"
            )

    _cost_html = ""
    if plan_name in ("pro", "business"):
        try:
            from settings import Settings as _ImpactS
            _impact_settings = _ImpactS(tenant_id=tenant_id) if tenant_id else _ImpactS()
            _avg_wage = float(_impact_settings.get("avg_hourly_wage", 18.0))
        except Exception:
            _avg_wage = 18.0
        _cost_total = 0.0
        for _row in scope_gs:
            try:
                _target = float(_row.get("Target UPH") or 0)
                _avg = float(_row.get("Average UPH") or 0)
                _hours = float(_row.get("Hours Worked") or 0)
                if _target > 0 and 0 < _avg < _target:
                    _cost_total += max(((_target - _avg) / _target) * _hours * _avg_wage, 0)
            except Exception:
                pass
        _cost_html = f"<h3>💰 Cost Impact</h3><p>Estimated labor cost impact for below-goal performance in this period: <strong>${_cost_total:,.0f}</strong></p>"

    body  = (f"<h2>{dept_label} — {period_label}</h2>"
             f"<p><strong>{len(scope_gs)}</strong> employees · "
             f"Avg UPH: <strong>{_avg_all}</strong> · "
             f"<strong style='color:green'>{on_g} on goal</strong> · "
             f"<strong style='color:red'>{below} below goal</strong></p>"
             f"<hr/>"
             f"{_dept_health_html}"
             f"{_top3_html}"
             f"{_top_risks_html}"
             f"{_critical_html}"
             f"{_improved_html}"
             f"{_trending_down_html}"
             f"{_bottom_html}"
             f"{_cost_html}"
             f"<hr/>"
             f"<p style='font-size: 0.9em; color: #666;'>See the attached Excel report for full details and department breakdown.</p>")

    xl_data = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        HDR_FILL  = PatternFill("solid", fgColor="FF0F2D52")
        HDR_FONT  = Font(bold=True, color="FFFFFFFF", size=10, name="Arial")
        GRN_FILL  = PatternFill("solid", fgColor="FFD4EDDA")
        RED_FILL  = PatternFill("solid", fgColor="FFF8D7DA")
        DARK_FONT = Font(color="FF1A2D42", size=10, name="Arial")
        STATUS_LABELS = {"on_goal":"On Goal","below_goal":"Below Goal","no_goal":"No Target"}

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Summary"
        ws["A1"] = f"{dept_label} — {period_label}"
        ws["A1"].font = Font(bold=True, size=13, name="Arial", color="FF0F2D52")
        ws["A2"] = f"Generated: {today.isoformat()}   |   {on_g} on goal · {below} below goal"
        ws["A2"].font = Font(size=10, name="Arial", color="FF5A7A9C")

        hdrs = ["Employee","Department","Total Units","Hours Worked","Avg UPH","Target UPH","Status"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(4, ci, h); c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(scope_gs, 5):
            fill = (GRN_FILL if r["goal_status"] == "on_goal"
                    else RED_FILL if r["goal_status"] == "below_goal" else None)
            for ci, v in enumerate([r["Employee Name"], r["Department"],
                                     r["Total Units"], r["Hours Worked"],
                                     r["Average UPH"], r["Target UPH"],
                                     STATUS_LABELS.get(r["goal_status"],"—")], 1):
                c = ws.cell(ri, ci, v); c.font = DARK_FONT
                c.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
                if fill: c.fill = fill

        if scope_gs:
            chart = BarChart(); chart.type = "bar"; chart.style = 10
            chart.title = f"Avg UPH — {period_label}"
            chart.width = 20; chart.height = max(8, len(scope_gs) * 0.55)
            last_r = 4 + len(scope_gs)
            chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=last_r), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last_r))
            chart.series[0].graphicalProperties.solidFill = "0F2D52"
            ws.add_chart(chart, "I5")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or "")) for c in col), default=8) + 3, 40)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); xl_data = buf.read()
    except Exception:
        pass

    return xl_data, subj, body


